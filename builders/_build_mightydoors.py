# Builds ToF_v1.xlsx — the 'ToF' config sheet for Mighty Doors / Tower of Fortune, the push-your-luck
# event from design_pdfs/DRBL-Mighty Doors (DB Tower of Fortune)-010926-212845.pdf.
# See source_docs/mighty-doors.md for the mechanics this encodes.
#
# WHAT IS AUTHORED HERE vs WHAT IS DERIVED
#   The deck specifies the STRUCTURE completely (60 stages, 4 doors, 1 Pig per standard stage, safe
#   stages every 5th plus milestones at 30/60, six reward tiers of 10 stages) and gives NOT ONE
#   numeric reward value, continue cost, or duration. So this sheet ships with the structure fully
#   populated and every VALUE at 0 — the same convention the D19/D21 pack ladders shipped under: a
#   zero is "not authored yet", which is correct rather than a plumbing failure.
#
# WHY THE STAGE TABLE CARRIES FORMULAS
#   Stage type, pig count, tier, survival and reach are all COMPUTABLE from the RUN CONFIG panel, so
#   per the project rule they are computed, not typed. Change 'Safe Stage Every N' or 'Total Stages'
#   in the panel and all 60 rows re-derive. Only the yellow cells are inputs.
#
# THE ACCUMULATION MODEL (why 'Reach p' is not enough on its own)
#   Rewards banked during a run are LOST if the player meets a Pig and declines to continue (deck
#   p7). Expected payout is therefore NOT sum(reach x reward) — it is
#       E[payout] = SUM_s P(cash out at s) x CumReward(s) + P(complete final stage) x CumReward(last)
#   A player who never cashes out and never continues expects ~nothing: P(reach 60) unaided is
#   0.75^47 ~ 1.3e-6. Cash-Out is the ONLY thing that converts a run into income, which is why the
#   Cash-Out % column is an input per stage rather than a single policy constant.
#
# CONTINUES (user decision 2026-09-02: unlimited for now, limit later)
#   Continuing removes the Pig and guarantees a reward from the remaining doors, so with take-up c
#   the effective survival is s + (1-s)*c. At c = 1 the Pig stops mattering entirely and every run
#   walks to the final stage — depth is set by take-up, not by the 3/4 odds.
#
# PLAYER REACH SIMULATION (added 2026-09-02 after review)
#   Every collection/milestone _v2 sheet in the workbook carries a 'Player Reach Simulation (per
#   event day) - SIMULATED' block anchored at AH1 - HH_v2, BB_v2, J_v2, Ph_v2, RM_1st_v2, RM_2nd_v2
#   all have it; the leaderboards (Ki, Race, F), TaD and NS do not, because rank and daily-reset
#   events have no progress to show. Mighty Doors is a progression ladder, so it belongs in the
#   first family and gets the block, with the SAME grammar: AH2/AI2 Percentile, AH3/AI3 Payer,
#   header row 5 = Day_of_event + <segment>_ms / <segment>_reward pairs, one row per event day.
#   For this event '_ms' is the STAGE reached (the milestone analogue) and '_reward' is the bundle
#   banked at that stage, rendered as the same '{Coins: 15, Red: 1}' dict string the others use.
#
# WHY BEHAVIOUR IS PER SEGMENT
#   The five _ms columns only say something if the segments differ. For a collection event they
#   differ through measured accrual; Mighty Doors has no accrual - a run is walked in one sitting -
#   so what separates a whale from a new player here is how often they PAY TO CONTINUE and how long
#   they push before cashing out. The deck's own business goal is 'additional spending opportunities
#   for whales and highly engaged players' (p2), so continue take-up is exactly the segment axis.
#   Global assumptions would render five identical columns, which would be worse than none.
#
# Sheet name is 'MD' (user decision: the calendar lane is 'MD', not the full event name). Duplicate
# the imported sheet as 'MD_v2' — the engine reads the _v2 sheet for the redesign side, exactly as
# NS_v2 shipped as a verbatim clone of NS.
import json
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter as CL

DISPLAY = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'display')

# The 21-column reward grammar every config sheet in the workbook shares (engine REWARD_COLUMNS).
# Punch-card rule: ALL currency columns are present even when this event never pays them.
REWARD_COLS = ['Coins', 'SPT', 'SPT x2', 'Red', 'Chuck', 'Bomb', 'Slingshot', 'Shuffle', 'Comet',
               'Unlimited Lives', 'Unlimited Red', 'Unlimited Chuck', 'Unlimited Bomb',
               'COOP Token', 'Avatar', '1-star Dly', '2-star Dly', '3-star Dly', '4-star Dly',
               '5-star Dly', '6-star Dly']

# The reward ladder is keyed by ENGINE resource names; the sheet's columns use the config-sheet
# spelling. One map, so the two lists cannot drift apart silently.
RES_ALIAS = {'Coins': 'HC', 'SPT x2': 'SPTx2',
             'Unlimited Red': 'UL Red', 'Unlimited Chuck': 'UL Chuck',
             'Unlimited Bomb': 'UL Bomb',
             '1-star Dly': '1-star Pack', '2-star Dly': '2-star Pack',
             '3-star Dly': '3-star Pack', '4-star Dly': '4-star Pack',
             '5-star Dly': '5-star Pack', '6-star Dly': '6-star Pack'}

TOTAL_STAGES = 60
N_TIERS = 6

F_BAR, F_HDR, F_IN, F_OUT, F_DATA = 'FF000000', 'FFF7CB4D', 'FFFFF2CC', 'FFE2EFDA', 'FFCFE2F3'
fill = lambda rgb: PatternFill('solid', fgColor=rgb)
ARIAL = lambda **kw: Font(name='Arial', **kw)
thin = Side(style='thin')
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'ToF'      # renamed from 'MD' 2026-09-02: ToF is the preferred name
ws.sheet_view.showGridLines = False


def bar(r, text, ncol):
    """Black section bar. Blocks are located by this column-A label, never by row number."""
    for c in range(1, ncol + 1):
        ws.cell(r, c).fill = fill(F_BAR)
    ws.cell(r, 1, text).font = ARIAL(size=11, bold=True, color='FFFFFFFF')


def hdr(r, labels):
    for i, t in enumerate(labels, start=1):
        c = ws.cell(r, i, t)
        c.font = ARIAL(size=10, bold=True)
        c.fill = fill(F_HDR)
        c.border = BOX
        c.alignment = Alignment(horizontal='center', wrap_text=True, vertical='center')


def note(r, text):
    ws.cell(r, 1, text).font = ARIAL(size=9, italic=True, color='FF666666')


def put(r, c, v, kind='out', numfmt=None):
    """kind: 'in' = author types here (yellow) · 'out' = derived (green) · 'data' = fixed (blue)."""
    cell = ws.cell(r, c, v)
    cell.font = ARIAL(size=10)
    cell.fill = fill({'in': F_IN, 'out': F_OUT, 'data': F_DATA}[kind])
    cell.border = BOX
    if numfmt:
        cell.number_format = numfmt
    return cell


LAD = json.load(open(os.path.join(os.path.dirname(os.path.abspath(__file__)), '_md_ladder.json')))
LADDER = LAD['ladder']

# Guard against the failure this map exists to prevent: the first build silently dropped every
# UL Bomb / UL Chuck / UL Red payout because those three aliases were missing, and a zero column
# looks exactly like "not authored yet". Fail the build instead.
_sheet_keys = {RES_ALIAS.get(c, c) for c in REWARD_COLS}
_lost = sorted({k for node in LADDER for k in node} - _sheet_keys)
if _lost:
    raise SystemExit('reward ladder produces resources with no sheet column: %s' % _lost)

ws['A1'] = 'MIGHTY DOORS (Tower of Fortune)'
ws['A1'].font = ARIAL(size=14, bold=True)

r = 3

# ---------------------------------------------------------------- RUN CONFIG
bar(r, 'RUN CONFIG', 4); r += 1
hdr(r, ['Parameter', 'Value', 'What it means']); r += 1
RUN_FIRST = r
RUN = {}
# Five parameters were removed 2026-09-02 because nothing read them and nothing could:
#   Total Stages          the ladder length IS the number of STAGES rows; a second number could
#                         only disagree with it
#   Tickets per Run       always 1 (deck p9); the engine assumes 1 when it is absent
#   Empty Outcomes        the Empty Slots column is 0 throughout and no reader consults the flag
#   Ticket Recharge       tickets are EARNED, not recharged (user decision) -- there is no recharge
#   Runs per Player       was derived from the recharge; runs now come from the ticket budget
# Cash-Out Variant STAYS: it is a real design switch in the deck (A p9/p19, B p23) and the engine
# reads it. Default A -- bank at any successful node.
run_rows = [
    ('Choices per Stage (default)', 4, 'in', 'How many doors are shown at a node. Min 2, max 4.'),
    ('Failure Outcomes per Stage', 1, 'in', 'Base number of Pigs. The STAGES table raises this on later tiers.'),
    ('Safe Stage Every N Stages', 5, 'in', 'Every Nth node carries no Pig at all.'),
    ('Major Milestone Stage', 30, 'in', 'A safe node with a large bundle.'),
    ('Aspirational Milestone Stage', 60, 'in', 'The final node. Reaching it auto claims everything.'),
    ('Reward Tier Size (stages)', 10, 'in', 'Only used to label the Tier column.'),
    ('Starting Tickets', LAD['start_tickets'], 'in',
     'Tickets the player already holds when the event opens. Everything after this is EARNED.'),
    ('Cash-Out Variant', 'A', 'in', 'A: bank at any node. B: bank only on safe nodes.'),
    ('Event Duration (days)', 3, 'in', 'Length of ONE instance. The event is always-on today: '
                                       'cal_new row 22 is a single merged 33-day lane.'),
]
for name, val, kind, meaning in run_rows:
    ws.cell(r, 1, name).font = ARIAL(size=10)
    put(r, 2, val, kind)
    ws.cell(r, 3, meaning).font = ARIAL(size=9, italic=True, color='FF666666')
    RUN[name] = '$B$%d' % r
    r += 1
# 'Runs per Player per Instance' was removed with the recharge it derived from. Runs are no longer
# a config number at all: the engine walks the calendar day by day, banks whatever ToF_Ticket the
# other sources paid that day, and spends up to 'Runs per Active Day'. Read the answer off
# ECOGAINS_TOF(payer,"RUN") below rather than typing one here.
RUNS_ROW = r
r += 1

# ---------------------------------------------------------------- SEGMENT BEHAVIOUR
# MAX is not a real engagement segment: it is the ceiling case -- a player with effectively
# unlimited tickets and coins who never voluntarily stops. It exists to price the deep ladder, which
# no real segment reaches (the deepest measured cash-out is stage 20, and 80% of the ladder's value
# sits on stages 30 and 60). Cash-Out Stage 0 means "never cash out", which the engine already reads
# as "run to the last stage".
SEGMENTS = ['0-9', '10-19', '20-39', '40-99', '100+', 'MAX']
bar(r, 'SEGMENT BEHAVIOUR', 7); r += 1
hdr(r, ['Segment', 'Continue Take-Up', 'Cash-Out Stage', 'Runs per Active Day',
        'Max Continues per Run', 'Coin Balance override',
        'What this row says about this player type']); r += 1
SEG_FIRST = r
BEH_NOTE = {
    '0-9': 'Rarely pays to revive, banks early at the first safe node.',
    '10-19': 'Occasionally revives, pushes one safe node further.',
    '20-39': 'Revives one time in five, comfortable going past the halfway safe node.',
    '40-99': 'Often revives, pushes deep because the coins are affordable.',
    '100+': 'Revives more often than not, so the Pig rarely stops the run.',
    'MAX': 'CEILING CASE, not a real player: always revives, never banks early, and holds enough '
           'coins that price never stops them. Shows what the deep ladder is worth to someone who '
           'actually reaches it.',
}
# Runs per Active Day is a CAP, not the driver. The engine spends whatever tickets the calendar
# paid, up to this many runs a day; 2 matches the old 12-hour recharge cadence.
RUNS_PER_DAY = {'0-9': 2, '10-19': 2, '20-39': 2, '40-99': 2, '100+': 2, 'MAX': 99}
MAX_BEH = {'take': 1.0, 'stop': 0, 'balance': 100000}
for seg in SEGMENTS:
    put(r, 1, seg, 'data')
    put(r, 2, MAX_BEH['take'] if seg == 'MAX' else LAD['take'][seg], 'in', '0%')
    put(r, 3, MAX_BEH['stop'] if seg == 'MAX' else LAD['stop'][seg], 'in')
    put(r, 4, RUNS_PER_DAY[seg], 'in', '0')
    put(r, 5, 0, 'in')
    put(r, 6, MAX_BEH['balance'] if seg == 'MAX' else '', 'in', '#,##0')
    ws.cell(r, 7, BEH_NOTE[seg]).font = ARIAL(size=9, italic=True, color='FF666666')
    r += 1
SEG_ROW = {seg: SEG_FIRST + i for i, seg in enumerate(SEGMENTS)}
# Column meanings, written once under the block so nobody has to guess.
for txt in [
    'Continue Take-Up: the chance this player PAYS COINS to revive when a Pig ends their run. '
    'It is not a retention number. At 100% the Pig never stops them and every run reaches the end.',
    'Cash-Out Stage: the node where this player voluntarily stops and banks what they have. '
    'They only get paid if they actually REACH it, because a Pig before then loses everything.',
    'Runs per Active Day: a CAP, not the driver. Tickets are EARNED, so the engine walks the '
    'calendar day by day, banks whatever ToF_Ticket the other sources paid, and spends up to this '
    'many runs. Unspent tickets carry forward with no cap.',
    'Max Continues per Run: 0 means "as many as the CONTINUE COST LADDER has rungs" - the ladder '
    'length is the real cap, because a rung with no price cannot be sold. A number here lowers it.',
    'Coin Balance override: normally BLANK, and the sim reads the four hc_balance percentiles '
    '(p25/p50/p75/p90) for this segment straight from data_econ, walking the run at each and '
    'averaging. A number here replaces all four for this segment - which is how MAX gets a wallet '
    'no real player has.',
]:
    ws.cell(r, 1, txt).font = ARIAL(size=9, italic=True, color='FF666666')
    r += 1
r += 1

# ---------------------------------------------------------------- CONTINUE COST LADDER (INPUT)
# Moved above the SIM PART divider: it is something you author, not something the sheet works out.
bar(r, 'CONTINUE COST LADDER', 4); r += 1
for txt in [
    'What it is: the COIN PRICE of reviving after a Pig, within a single run. The first revive in '
    'a run costs row 1, the second costs row 2, and so on. Costs must rise each time (deck p8).',
    'This is an INPUT you author. It is a coin SINK, the only one this event has, and it is how '
    'the feature is meant to make money.',
]:
    ws.cell(r, 1, txt).font = ARIAL(size=9, italic=True, color='FF666666')
    r += 1
# The first four rungs are ANCHORED to real wallet percentiles from data_econ, so each one is the
# rung a named slice of the population can still just afford. Pooled across all ten segment/payer
# rows: p25 ~ 50 coins, p50 ~ 99, p75 ~ 258, p90 ~ 716.
#     rung 1 cum   25  - under every segment's p25, so the FIRST continue is affordable to everyone
#     rung 2 cum  100  - the median player
#     rung 3 cum  260  - p75
#     rung 4 cum  720  - p90
# Past that the population has run out, so growth is a plain multiplier and the rungs exist for the
# MAX case. The CAP is expressed as a rung INDEX, not a price: change the multiplier and the ceiling
# re-derives itself instead of needing a new number invented for it.
CONT_ANCHORED = [25, 75, 160, 460]      # percentile-anchored, authored
CONT_MULT     = 2.5                     # growth per rung past the anchored four
CONT_CAP_RUNG = 7                       # growth STOPS at this rung; later rungs repeat its price
CONT_RUNGS    = 10
for txt in [
    'The first four rungs are priced against real wallets (data_econ hc_balance): rung 1 sits under '
    'every segment p25 so the first continue always feels affordable, and rungs 2-4 sit at the p50, '
    'p75 and p90 cumulative. Rung 5 onward is where the population has run out.',
    'Growth multiplier and "growth stops after rung" are INPUTS. The cap is a rung INDEX rather '
    'than a price, so changing the multiplier moves the ceiling with it and no new cap has to be '
    'invented. Rungs past the cap repeat the capped price - a revive never gets more expensive '
    'than the cap, it just stays there.',
]:
    ws.cell(r, 1, txt).font = ARIAL(size=9, italic=True, color='FF666666')
    r += 1
ws.cell(r, 1, 'Growth multiplier (rung 5+)').font = ARIAL(size=10)
put(r, 2, CONT_MULT, 'in', '0.0"x"'); MULT_REF = '$B$%d' % r; r += 1
ws.cell(r, 1, 'Growth stops after rung #').font = ARIAL(size=10)
put(r, 2, CONT_CAP_RUNG, 'in'); CAP_REF = '$B$%d' % r; r += 1
r += 1
hdr(r, ['Continue # in this run', 'Cost (Coins)', 'Cumulative if they revive this many times']); r += 1
CONT_FIRST = r
for i in range(1, CONT_RUNGS + 1):
    put(r, 1, i, 'data')
    if i <= len(CONT_ANCHORED):
        put(r, 2, CONT_ANCHORED[i - 1], 'in')
    else:
        # Grow from the LAST ANCHORED rung by an exponent that stops climbing at the cap. Chaining
        # off the row above instead (=IF(rung>cap, <capped row>, prev*mult)) makes the capped rung's
        # own formula name its own cell, which Sheets reports as a circular reference even though
        # the branch is never taken. This form references only fixed cells above it, and both the
        # multiplier AND the cap rung stay live inputs.
        #   price_i = anchor_last x mult ^ clamp(i - n_anchored, 0, cap - n_anchored)
        anchor_row = CONT_FIRST + len(CONT_ANCHORED) - 1
        n = len(CONT_ANCHORED)
        put(r, 2, '=ROUND($B${ar}*{mult}^MAX(0,MIN($A{row}-{n},{cap}-{n})),0)'.format(
            ar=anchor_row, mult=MULT_REF, row=r, n=n, cap=CAP_REF), 'out', '#,##0')
    put(r, 3, '=SUM($B${a}:$B{r})'.format(a=CONT_FIRST, r=r), 'out', '#,##0')
    r += 1
CONT_LAST = r - 1
r += 1

# ---------------------------------------------------------------- STAGES
NODE_LEAD = ['Stage', 'Type', 'Tier', 'Choices', 'Reward Slots', 'Pig Slots', 'Empty Slots',
             'Survive p', 'P(reward | survived)']
bar(r, 'STAGES', len(NODE_LEAD) + len(REWARD_COLS)); r += 1
for txt in [
    'One row per node. Pig Slots rises with depth (1 for stages 1 to 20, 2 for 21 to 40, 3 for 41 '
    'to 60) and never goes above 3, so later nodes are far more dangerous. Safe nodes carry none.',
    'Which door hides the Pig is drawn at random when the player arrives, so only the COUNT is '
    'configurable, never the position.',
    'Rewards are tuned so the event pays each segment roughly 75% of what Rainbow Maker pays, '
    'allowing for 3 days against Rainbow Maker 4. See builders/_md_ladder.py for the solve.',
]:
    ws.cell(r, 1, txt).font = ARIAL(size=9, italic=True, color='FF666666')
    r += 1
hdr(r, NODE_LEAD + REWARD_COLS); r += 1
NODE_FIRST = r
NODE_C0 = len(NODE_LEAD) + 1
for st in range(1, TOTAL_STAGES + 1):
    row = NODE_FIRST + st - 1
    put(row, 1, st, 'data')
    put(row, 2, ('=IF(A{r}=1,"Safe (Start)",IF(OR(A{r}={maj},A{r}={asp}),"Milestone",'
                 'IF(AND({every}>0,MOD(A{r},{every})=0),"Safe","Standard")))').format(
        r=row, maj=RUN['Major Milestone Stage'], asp=RUN['Aspirational Milestone Stage'],
        every=RUN['Safe Stage Every N Stages']))
    put(row, 3, '=MIN({nt},MAX(1,ROUNDUP(A{r}/{ts},0)))'.format(
        r=row, nt=N_TIERS, ts=RUN['Reward Tier Size (stages)']))
    put(row, 4, '=%s' % RUN['Choices per Stage (default)'], 'in')
    # Pig ramp: 1 / 2 / 3 by tier pair, never above 3, and never on a safe node.
    put(row, 6, '=IF($B{r}<>"Standard",0,MIN(3,MAX({p},IF($C{r}<=2,1,IF($C{r}<=4,2,3)))))'.format(
        r=row, p=RUN['Failure Outcomes per Stage']))
    put(row, 7, 0, 'in')
    put(row, 5, '=MAX(0,$D{r}-$F{r}-$G{r})'.format(r=row))
    put(row, 8, '=IF($D{r}<=0,0,($D{r}-$F{r})/$D{r})'.format(r=row), 'out', '0.0%')
    put(row, 9, '=IF($D{r}-$F{r}<=0,0,$E{r}/($D{r}-$F{r}))'.format(r=row), 'out', '0.0%')
    payout = LADDER[st - 1]
    for j, res in enumerate(REWARD_COLS):
        put(row, NODE_C0 + j, payout.get(RES_ALIAS.get(res, res), 0), 'in')
NODE_LAST = NODE_FIRST + TOTAL_STAGES - 1
NODE_NCOL = len(NODE_LEAD) + len(REWARD_COLS)
ws.conditional_formatting.add(
    'A%d:%s%d' % (NODE_FIRST, CL(NODE_NCOL), NODE_LAST),
    FormulaRule(formula=['$F%d=0' % NODE_FIRST], fill=fill('FFE8F5E9')))
ws.conditional_formatting.add(
    'F%d:F%d' % (NODE_FIRST, NODE_LAST),
    CellIsRule(operator='greaterThanOrEqual', formula=['2'], fill=fill('FFF4CCCC')))
ws.conditional_formatting.add(
    '%s%d:%s%d' % (CL(NODE_C0), NODE_FIRST, CL(NODE_NCOL), NODE_LAST),
    CellIsRule(operator='greaterThan', formula=['0'], fill=fill('FFD9EAD3'),
               font=ARIAL(size=10, bold=True)))
r = NODE_LAST + 2

# ================================================================ SIM PART
bar(r, 'SIM PART', 12); r += 1

# ---------------------------------------------------------------- PER-SEGMENT PROGRESSION
SEG_SUB = ['Survive (w/ cont.)', 'Reach p', 'P(end here)', 'Cum P(ended)']
SEG_C0 = 2
bar(r, 'PER-SEGMENT PROGRESSION', 1 + len(SEGMENTS) * len(SEG_SUB)); r += 1
for txt in [
    'Survive (w/ cont.): chance this player gets past this node once you allow for them paying to '
    'revive. Higher than the raw odds, because a revive turns a Pig into a guaranteed reward.',
    'Reach p: chance they ever arrive at this node at all, having survived every node before it '
    'and not banked early.',
    'P(end here): chance their run finishes at this node, either because they banked or because a '
    'Pig got them and they refused to pay.',
    'Cum P(ended): running total of the column to its left, so it climbs to 1.0 by the last node. '
    'The reach simulation reads this to find the typical stopping point.',
]:
    ws.cell(r, 1, txt).font = ARIAL(size=9, italic=True, color='FF666666')
    r += 1
hdr(r, ['Stage'])
for i, seg in enumerate(SEGMENTS):
    for j, lbl in enumerate(SEG_SUB):
        cell = ws.cell(r, SEG_C0 + i * len(SEG_SUB) + j, seg + ' ' + lbl)
        cell.font = ARIAL(size=9, bold=True); cell.fill = fill(F_HDR); cell.border = BOX
        cell.alignment = Alignment(horizontal='center', wrap_text=True, vertical='center')
r += 1
PRG_FIRST = r
for st in range(1, TOTAL_STAGES + 1):
    row = PRG_FIRST + st - 1
    nrow = NODE_FIRST + st - 1
    put(row, 1, '=$A{n}'.format(n=nrow), 'data')
    for i, seg in enumerate(SEGMENTS):
        c0 = SEG_C0 + i * len(SEG_SUB)
        C = lambda k: CL(c0 + k)
        take, stop = '$B$%d' % SEG_ROW[seg], '$C$%d' % SEG_ROW[seg]
        put(row, c0 + 0, '=$H{n}+(1-$H{n})*{t}'.format(n=nrow, t=take), 'out', '0.0%')
        if st == 1:
            put(row, c0 + 1, 1, 'out', '0.000%')
        else:
            p = row - 1
            put(row, c0 + 1, '={c1}{p}*{c0}{p}*IF(AND({s}>0,$A{p}>={s}),0,1)'.format(
                c1=C(1), c0=C(0), p=p, s=stop), 'out', '0.000%')
        put(row, c0 + 2, ('={c1}{r}*(IF(AND({s}>0,$A{r}>={s}),{c0}{r},0)+(1-$H{n})*(1-{t}))').format(
            c1=C(1), c0=C(0), r=row, n=nrow, s=stop, t=take), 'out', '0.000%')
        put(row, c0 + 3, '=SUM({c2}${f}:{c2}{r})'.format(c2=C(2), f=PRG_FIRST, r=row), 'out', '0.000%')
PRG_LAST = PRG_FIRST + TOTAL_STAGES - 1
r = PRG_LAST + 2

# ---------------------------------------------------------------- RUN EXPECTATION
bar(r, 'RUN EXPECTATION BY PLAYER TYPE (derived)', 1 + len(SEGMENTS)); r += 1
ws.cell(r, 1, 'Rewards banked in a run are LOST if a Pig ends it and the player will not pay, so '
              'reaching a node is not the same as being paid for it.').font = ARIAL(
    size=9, italic=True, color='FF666666'); r += 1
hdr(r, ['Metric'] + SEGMENTS); r += 1
EXP_FIRST = r
SC = lambda i, k: CL(SEG_C0 + i * len(SEG_SUB) + k)
exp_rows = [
    ('P(reach Major Milestone)', lambda i, s: '=IFERROR(INDEX({c}{a}:{c}{b},MATCH({m},$A{a}:$A{b},0)),0)'.format(
        c=SC(i, 1), a=PRG_FIRST, b=PRG_LAST, m=RUN['Major Milestone Stage']), '0.000%'),
    ('P(reach Final Stage)', lambda i, s: '=IFERROR(INDEX({c}{a}:{c}{b},MATCH({m},$A{a}:$A{b},0)),0)'.format(
        c=SC(i, 1), a=PRG_FIRST, b=PRG_LAST, m=RUN['Aspirational Milestone Stage']), '0.00000%'),
    ('P(run pays NOTHING)', lambda i, s: '=SUMPRODUCT({c}{a}:{c}{b},(1-$H${nf}:$H${nl}),(1-$B${sr}))'.format(
        c=SC(i, 1), a=PRG_FIRST, b=PRG_LAST, nf=NODE_FIRST, nl=NODE_LAST, sr=SEG_ROW[s]), '0.0%'),
    ('P(run pays ANYTHING)', lambda i, s: '=1-{c}{r}'.format(c=CL(2 + i), r=EXP_FIRST + 2), '0.0%'),
    ('Expected stage reached', lambda i, s: '=SUM({c}{a}:{c}{b})'.format(
        c=SC(i, 1), a=PRG_FIRST, b=PRG_LAST), '0.00'),
    ('Expected continues per run', lambda i, s: '=SUMPRODUCT({c}{a}:{c}{b},(1-$H${nf}:$H${nl}),$B${sr})'.format(
        c=SC(i, 1), a=PRG_FIRST, b=PRG_LAST, nf=NODE_FIRST, nl=NODE_LAST, sr=SEG_ROW[s]), '0.000'),
    ('Coins spent on continues per run', lambda i, s: '=IFERROR(INDEX($C${cf}:$C${cl},'
        'MIN({n},MAX(1,ROUND({c}{er},0)))),0)*{c}{er}/MAX({c}{er},1)'.format(
        cf=CONT_FIRST, cl=CONT_LAST, n=CONT_RUNGS, c=CL(2 + i), er=EXP_FIRST + 5), '0.0'),
    ('Stage reached at percentile', lambda i, s: '=MIN({t},COUNTIF({c}{a}:{c}{b},"<"&$AI$2)+1)'.format(
        t=TOTAL_STAGES, c=SC(i, 3), a=PRG_FIRST, b=PRG_LAST), '0'),
]
for name, fn, numfmt in exp_rows:
    ws.cell(r, 1, name).font = ARIAL(size=10, bold=True)
    for i, seg in enumerate(SEGMENTS):
        put(r, 2 + i, fn(i, seg), 'out', numfmt)
    r += 1
PCT_ROW = r - 1
r += 1

# ---------------------------------------------------------------- CUMULATIVE REWARD
bar(r, 'CUMULATIVE REWARD AT THE REACHED STAGE (derived)', 1 + len(SEGMENTS)); r += 1
ws.cell(r, 1, 'Everything this player type has banked by the node they stop at, counting every '
              'node up to it.').font = ARIAL(size=9, italic=True, color='FF666666'); r += 1
hdr(r, ['Resource'] + SEGMENTS); r += 1
CUM_FIRST = r
for k, res in enumerate(REWARD_COLS):
    ws.cell(r, 1, res).font = ARIAL(size=10)
    col = CL(NODE_C0 + k)
    for i, seg in enumerate(SEGMENTS):
        m = '{c}${pr}'.format(c=CL(2 + i), pr=PCT_ROW)
        put(r, 2 + i, '=SUMPRODUCT(($A${a}:$A${b}<={m})*$I${a}:$I${b}*{c}${a}:{c}${b})'.format(
            a=NODE_FIRST, b=NODE_LAST, m=m, c=col), 'out', '0.00')
    r += 1
CUM_LAST = r - 1

# ------------------------------------------- PLAYER REACH SIMULATION (AH1, house grammar)
RS_C = 34
ws.cell(1, RS_C, 'Player Reach Simulation (per event day) - SIMULATED').font = ARIAL(size=11, bold=True)
ws.cell(2, RS_C, 'Percentile').font = ARIAL(size=10, bold=True)
put(2, RS_C + 1, 0.5, 'in', '0.00')
ws.cell(3, RS_C, 'Payer').font = ARIAL(size=10, bold=True)
put(3, RS_C + 1, 'NONPAYER', 'in')
rs_hdr = ['Day_of_event']
for seg in SEGMENTS:
    lbl = seg[:-1] if seg.endswith('+') else seg
    rs_hdr += [lbl + '_ms', lbl + '_reward']
for j, t in enumerate(rs_hdr):
    c = ws.cell(4, RS_C + j, t)
    c.font = ARIAL(size=10, bold=True); c.fill = fill(F_HDR); c.border = BOX
    c.alignment = Alignment(horizontal='center', wrap_text=True, vertical='center')
dur = RUN['Event Duration (days)']
for d in range(1, 15):
    row = 4 + d
    put(row, RS_C, d, 'data')
    for i, seg in enumerate(SEGMENTS):
        put(row, RS_C + 1 + i * 2, '=IF(AND({dur}>0,$AH{r}>{dur}),"",{c}${pr})'.format(
            dur=dur, r=row, c=CL(2 + i), pr=PCT_ROW), 'out')
        put(row, RS_C + 2 + i * 2,
            ('=IF({msc}{r}="","","{{"&TEXTJOIN(", ",TRUE,IF({cc}${cf}:{cc}${cl}>0,'
             '$A${cf}:$A${cl}&": "&{cc}${cf}:{cc}${cl},""))&"}}")').format(
                msc=CL(RS_C + 1 + i * 2), r=row, cc=CL(2 + i), cf=CUM_FIRST, cl=CUM_LAST), 'out')

ws.column_dimensions['A'].width = 30.0
ws.column_dimensions['B'].width = 20.0
ws.column_dimensions['C'].width = 18.0
for c in range(4, 4 + len(REWARD_COLS) + 30):
    ws.column_dimensions[CL(c)].width = 12.0

out = os.path.join(DISPLAY, 'ToF_v1.xlsx')
wb.save(out)
print('written ToF_v1.xlsx  (sheet "ToF")')
print('  RUN CONFIG        rows %d..%d' % (RUN_FIRST, RUNS_ROW))
print('  SEGMENT BEHAVIOUR rows %d..%d' % (SEG_FIRST, SEG_FIRST + len(SEGMENTS) - 1))
print('  CONTINUE COSTS    rows %d..%d   (INPUT, above the SIM PART divider)' % (CONT_FIRST, CONT_LAST))
print('  STAGES            rows %d..%d' % (NODE_FIRST, NODE_LAST))
print('  PER-SEG PROGRESS  rows %d..%d' % (PRG_FIRST, PRG_LAST))
print('  RUN EXPECTATION   rows %d..%d' % (EXP_FIRST, PCT_ROW))
print('  CUMULATIVE REWARD rows %d..%d' % (CUM_FIRST, CUM_LAST))
print('  REWARD POOL MAP removed (glossary, referenced by nothing)')
