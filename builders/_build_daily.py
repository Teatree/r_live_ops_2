# Builds EcoGainsSim_Daily_v2.xlsx — per-day view of the 33-day sim, driven by ECOGAINS_DAILY
# (EcoGainsSim_Daily.gs). Rows = days 1-33 (+DoW), blocks = CURRENT / NEW / DIFF x 11 resources
# + since v2 the per-earner NET blocks (ACTUAL SPEND / CURRENT NET / NEW NET spills reading
# data_econ_daily, plus a net-Δ formula block) — NET is blank unless Source = ALL and until the
# data_econ_daily sheet exists. Filters = Payer / Segment / Source dropdowns.
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter as CL

DISPLAY = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'display')

# 13 since 2026-07-10 (append-only): SPT/SPTx2 = season pass tokens (D16)
RES = ['HC', 'Slingshot', 'Shuffle', 'Comet', 'Red', 'Chuck', 'Bomb',
       'UL Bomb', 'UL Chuck', 'UL Red', 'Unlimited Lives', 'SPT', 'SPTx2']
CATS = ['Ads', 'Bomb Challenge', "Bomb's Ballet", 'Chuck Challenge', 'Core', 'Daily Gift',
        'Daily Night Sky Prize', 'Flock Flurry', 'Hatchling Hideaway', 'Jigsaw', 'Kite Festival',
        'Level Race', 'Other', 'Photoshoot', 'Red Challenge', 'River Rush', 'Saga',
        'Season Pass (Free)', 'Target Day', 'Team Event', 'Team Race', 'Flash Race',
        'FlowerCoop', 'Rainbow Maker', 'IAPs']
DOW = ['Wed', 'Thu', 'Fri', 'Sat', 'Sun', 'Mon', 'Tue']

F_CUR, F_NEW, F_DIF = 'FF2E75B6', 'FF548235', 'FF666666'
F_HCUR, F_HNEW, F_HDIF = 'FFCFE2F3', 'FFE2EFDA', 'FFEFEFEF'
F_NET, F_HNET = 'FF7030A0', 'FFE9DEF3'          # NET bands: same purple family as Sim per Segment
F_YELLOW, F_HDRSEG = 'FFFFF2CC', 'FFD9D9D9'
F_WEEKEND = 'FFFEF8E3'
fill = lambda rgb: PatternFill('solid', fgColor=rgb)
thin = Side(style='thin')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
FMT_VAL, FMT_DIFF = '#,##0.00', '#,##0.00;-#,##0.00'
FMT_NET = '#,##0.00;[Red]-#,##0.00'             # net can be negative (spend > gain)

# 13-wide blocks (since 2026-07-10) + 1 gap col: pitch 14
BLOCKS = [(4, 'CURRENT', '◄ CURRENT (cal_curr, measured) ►', F_CUR, F_HCUR, FMT_VAL),
          (18, 'NEW', '◄ NEW (cal_new, simulated) ►', F_NEW, F_HNEW, FMT_VAL),
          (32, 'DIFF', 'Δ  NEW − CURRENT (same day)', F_DIF, F_HDIF, FMT_DIFF)]
# NET blocks (v2): three ECOGAINS_DAILY spills reading data_econ_daily (per earner, actual data)
NET_SPILL = [(46, 'SPEND', '◄ ACTUAL SPEND / earner (data_econ_daily) ►', F_NET, F_HNET, FMT_VAL),
             (60, 'CURNET', '◄ CURRENT NET / earner (gain − spend) ►', F_NET, F_HNET, FMT_NET),
             (74, 'NEWNET', '◄ NEW NET / earner (gain + sim Δ − spend) ►', F_NET, F_HNET, FMT_NET)]
# net Δ = NEWNET − CURNET, plain sheet formulas (== the DIFF block when NET is filled)
NETD = (88, 'NETD', 'Δ  NET (new − cur; == sim Δ)', F_DIF, F_HDIF, FMT_DIFF)
STYLE_BLOCKS = BLOCKS + NET_SPILL + [NETD]      # everything that gets band/header/day styling
GAPS = ['Q', 'AE', 'AS', 'BG', 'BU', 'CI', 'CW']
DROP_COL = 102                                  # CX — source dropdown list
D0, DN = 9, 41            # data rows (33 days)
TOT = 42

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'EcoGainsSim_Daily'
ws.sheet_view.showGridLines = False
ws.column_dimensions['A'].width = 1.75
ws.column_dimensions['B'].width = 5.5
ws.column_dimensions['C'].width = 5.0
for c0, *_ in STYLE_BLOCKS:
    for j in range(len(RES)):
        ws.column_dimensions[CL(c0 + j)].width = 8.0
for g in GAPS:
    ws.column_dimensions[g].width = 2.88
ws.column_dimensions[CL(DROP_COL)].width = 21.0

ws['B1'] = 'DAILY RESOURCE GAINS + NET — 33-day calendar, per earner (day 1 = Wednesday)'
ws['B1'].font = Font(name='Arial', size=13, bold=True)

# ---- filters ----
filters = [('Payer', 'NONPAYER', '"NONPAYER,PAYER"'),
           ('Segment', '0-9', '"0-9,10-19,20-39,40-99,100+"'),
           ('Source', 'ALL', f'=${CL(DROP_COL)}$2:${CL(DROP_COL)}$27')]
for i, (label, default, f1) in enumerate(filters):
    r = 3 + i
    ws.cell(r, 2, label).font = Font(name='Arial', size=10, bold=True)
    v = ws.cell(r, 3, default)
    v.font = Font(name='Arial', size=10, bold=True)
    v.fill = fill(F_YELLOW)
    v.alignment = Alignment(horizontal='center')
    v.border = BORDER
    dv = DataValidation(type='list', formula1=f1, allow_blank=False, showDropDown=False)
    ws.add_data_validation(dv)
    dv.add(ws.cell(r, 3))
ws['D3'] = '◀ filters (dropdowns) — the whole table re-simulates on change'
ws['D3'].font = Font(name='Arial', size=9, color='FF808080')

# source dropdown list (range-based: 26 items exceed the 255-char inline validation limit)
ws.cell(1, DROP_COL, 'source dropdown list').font = Font(name='Arial', size=8, color='FF808080')
for i, s in enumerate(['ALL'] + CATS):
    ws.cell(2 + i, DROP_COL, s).font = Font(name='Arial', size=8, color='FF808080')

# ---- band row (7) + header row (8) ----
for c0, key, txt, band, hdr, fmt in STYLE_BLOCKS:
    for j in range(len(RES)):
        cell = ws.cell(7, c0 + j)
        cell.fill = fill(band)
        cell.font = Font(name='Arial', size=11, bold=True, color='FFFFFFFF')
    ws.cell(7, c0).value = txt
    for j, r in enumerate(RES):
        cell = ws.cell(8, c0 + j, r)
        cell.font = Font(name='Arial', size=9, bold=True)
        cell.fill = fill(hdr)
        cell.alignment = Alignment(horizontal='center')
        cell.border = BORDER
for col, label in ((2, 'Day'), (3, 'DoW')):
    cell = ws.cell(8, col, label)
    cell.font = Font(name='Arial', size=9, bold=True)
    cell.fill = fill(F_HDRSEG)
    cell.alignment = Alignment(horizontal='center')
    cell.border = BORDER

# ---- day rows ----
for d in range(1, 34):
    r = D0 + d - 1
    b = ws.cell(r, 2, d)
    c = ws.cell(r, 3, DOW[(d - 1) % 7])
    for cell in (b, c):
        cell.font = Font(name='Arial', size=10)
        cell.alignment = Alignment(horizontal='center')
    for c0, key, txt, band, hdr, fmt in STYLE_BLOCKS:
        for j in range(len(RES)):
            cell = ws.cell(r, c0 + j)
            cell.font = Font(name='Arial', size=10)
            cell.alignment = Alignment(horizontal='center')
            cell.number_format = fmt

# spill anchors (CURRENT/NEW/DIFF + the three NET spills). The trailing sim_refresh!$A$1 is the
# engine's refresh NONCE (ignored by the function; changing it is what re-runs the sim — the
# engine no longer clears/re-sets formulas, which is what used to wipe them periodically).
for c0, key, *_ in BLOCKS + NET_SPILL:
    ws.cell(D0, c0).value = (f'=LET(payer,$C$3, segment,$C$4, source,$C$5, '
                             f'ECOGAINS_DAILY(payer, segment, source, "{key}", sim_refresh!$A$1))')

# net Δ block: per-cell formulas NEWNET − CURNET. IFERROR is load-bearing: when NET is blank
# (source ≠ ALL or no data_econ_daily) the spills emit '' text cells, the subtraction errors,
# and the Δ shows blank instead of a false 0.
for r in range(D0, DN + 1):
    for j in range(len(RES)):
        ws.cell(r, NETD[0] + j).value = \
            f'=IFERROR({CL(NET_SPILL[2][0]+j)}{r}-{CL(NET_SPILL[1][0]+j)}{r},"")'

# ---- TOTAL row ----
tc = ws.cell(TOT, 2, 'TOTAL')
tc.font = Font(name='Arial', size=9, bold=True)
tc.fill = fill(F_HDRSEG)
tc.border = BORDER
ws.cell(TOT, 3).fill = fill(F_HDRSEG)
ws.cell(TOT, 3).border = BORDER
for c0, key, txt, band, hdr, fmt in STYLE_BLOCKS:
    is_net = key in ('SPEND', 'CURNET', 'NEWNET', 'NETD')
    for j in range(len(RES)):
        col = CL(c0 + j)
        # NET totals blank out when the block is blank ('' cells → COUNT 0); a plain SUM would
        # show 0 and read as "net is zero" instead of "no data / not applicable".
        f = (f'=IF(COUNT({col}{D0}:{col}{DN})=0,"",SUM({col}{D0}:{col}{DN}))' if is_net
             else f'=SUM({col}{D0}:{col}{DN})')
        cell = ws.cell(TOT, c0 + j, f)
        cell.font = Font(name='Arial', size=10, bold=True)
        cell.fill = fill(hdr)
        cell.alignment = Alignment(horizontal='center')
        cell.number_format = fmt
        cell.border = BORDER

# ---- conditional formatting ----
# red/green on the DIFF and net-Δ blocks first (stopIfTrue so they win over the weekend tint)
for rng in (f'{CL(BLOCKS[2][0])}{D0}:{CL(BLOCKS[2][0] + len(RES) - 1)}{DN}',
            f'{CL(NETD[0])}{D0}:{CL(NETD[0] + len(RES) - 1)}{DN}'):
    ws.conditional_formatting.add(rng, CellIsRule(
        operator='lessThan', formula=['0'], stopIfTrue=True,
        font=Font(color='FF990000'), fill=fill('FFF4CCCC')))
    ws.conditional_formatting.add(rng, CellIsRule(
        operator='greaterThan', formula=['0'], stopIfTrue=True,
        font=Font(color='FF006100'), fill=fill('FFD9EAD3')))
# weekend tint across the whole table (Fri/Sat/Sun)
ws.conditional_formatting.add(f'B{D0}:{CL(NETD[0] + len(RES) - 1)}{DN}', FormulaRule(
    formula=[f'OR($C{D0}="Fri",$C{D0}="Sat",$C{D0}="Sun")'], fill=fill(F_WEEKEND)))

# ---- legend ----
legend = [
    'LEGEND / ALLOCATION RULES (claim-day realistic — EcoGainsSim_Daily.gs)',
    'Window totals are exactly the main sim’s numbers (CURRENT = measured on cal_curr, NEW = simulated on cal_new); this view only places them on days — column TOTALs reconcile with EcoGainsSim_HC.',
    'Leaderboard events (Bomb/Chuck/Red Challenge, Level Race, Flash Race, Target Day, Kite) pay on each instance’s LAST day → expect end-day spikes. Instance slices are split by reach.',
    'Collections (Hatchling Hideaway, Bomb’s Ballet, Jigsaw, Photoshoot) spread across instance days by the accrual curve’s marginal share; Rainbow Maker spreads ∝ active rate within its instances (no curve — flagged).',
    'Core / Saga / Daily Gift pay daily ∝ weekday/weekend active rate; Night Sky over its daily instances. Season Pass (Free) spreads ∝ active rate across its season-lane instances (tier claims are continuous — same treatment as Rainbow Maker). Non-calendar sources (Ads, Teams, Other, IAPs; River Rush current side) are flat ÷33 — their DIFF is uniform.',
    'Weekend rows (Fri/Sat/Sun) are tinted. Filters: Payer / Segment / Source — Source=ALL sums every category; pick one to isolate its daily contribution.',
    'NET blocks (AT:CH, per EARNER; 13 resources incl. SPT/SPTx2): ACTUAL SPEND and the gain side come from the data_econ_daily sheet (real per-day telemetry, window-earner denominator). CURRENT NET = actual gain − spend; NEW NET = actual gain + the sim’s day shift (NEW − CURRENT) − spend, so spend is held constant and net Δ == the DIFF block.',
    'NET blocks are BLANK when Source ≠ ALL (spend is game-wide, not attributable to one event) and until the data_econ_daily sheet exists in the workbook. Their TOTAL row blanks out too (no data ≠ zero net).',
]
for i, txt in enumerate(legend):
    cell = ws.cell(TOT + 2 + i, 2, txt)
    cell.font = Font(name='Arial', size=9, bold=(i == 0), color='FF000000' if i == 0 else 'FF808080')

wb.save(os.path.join(DISPLAY, 'EcoGainsSim_Daily_v2.xlsx'))
print('written EcoGainsSim_Daily_v2.xlsx — blocks at',
      '/'.join(CL(c0) for c0, *_ in STYLE_BLOCKS), '(13 res each), data rows',
      D0, '-', DN, ', TOTAL', TOT)
