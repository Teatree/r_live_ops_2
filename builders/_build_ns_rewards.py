# Builds display/NS_Rewards_v1.xlsx — the 'NS Proposal' tab for LiveOps_v2_AB_Summary.xlsx.
#
# Purpose: size a Night Sky COIN uplift that pulls the A/B's HC-gain regression back into the
# -4%..0% band and steepens the streak ladder, and show the resulting config.
#
# EVERYTHING observed is a live formula over the A/B workbook's own 'Raw Data' sheet, in the same
# SUMPRODUCT/SUMIFS idiom the 'By Bucket' tab uses, so the sheet re-prices itself on the next data
# pull. The proposed ladder is an INPUT block (#FFF2CC): retype a reward and the HC delta, the new
# Δ%, the band check and the marginal-HC-per-win table all move.
#
# NOTE ON THE STANDALONE FILE: the 'Raw Data' formulas cannot resolve until the tab is copied into
# LiveOps_v2_AB_Summary.xlsx (this file has no Raw Data sheet — same pattern as every other display
# sheet in display/, which references data_* sheets that live in the target workbook).
#
# Style is lifted from the A/B workbook, not from the simulation-sheet palette: Arial, no gridlines,
# 305496/4472C4 header bands, D9E1F2 category rows, +0.00%;-0.00% deltas, red/green CF.
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, '..', 'display', 'NS_Rewards_v1.xlsx')

# ---- the config under discussion -------------------------------------------------------------
# Requirements are UNCHANGED (this proposal moves coins only). Current == the live NS ladder,
# identical in NEW_LIVEOPS_CALENDAR_ECO (13) and (14) and identical in NS_v2 — i.e. Night Sky was
# NOT part of what the A/B tested, which is what makes it a clean additive lever on the result.
SEGS = [  # display name, 'Raw Data' avg_completions_7d_bucket key
    ('1-9',   'B. 1-9'),
    ('10-19', 'C. 10-19'),
    ('20-39', 'D. 20-39'),
    ('40-99', 'E. 40-99'),
    ('100+',  'F. 100+'),
]
REQ  = {'1-9': [2, 5, 10],   '10-19': [6, 13, 26],   '20-39': [11, 26, 42],
        '40-99': [28, 60, 100], '100+': [80, 175, 280]}
CUR  = {'1-9': [0, 10, 15],  '10-19': [10, 30, 60],  '20-39': [15, 50, 100],
        '40-99': [50, 120, 250], '100+': [100, 300, 400]}
PROP = {'1-9': [0, 15, 30],  '10-19': [10, 45, 100], '20-39': [15, 70, 160],
        '40-99': [60, 160, 300], '100+': [120, 450, 700]}

# ---- static cross-check from the economy sim (EcoGainsSim_v4.gs over workbook (14)) ------------
# R = E(NS_v2)/E(NS) through the survival model (data_streaks max-streak percentiles x
# NS_STREAK_N 1.25) under the PROPOSED ladder, and the resulting 33-day NS row (HC per earner,
# NONPAYER). Static by nature — it comes from the other workbook, not from this one.
SIM_R    = {'1-9': 1.677, '10-19': 1.393, '20-39': 1.374, '40-99': 1.252, '100+': 1.358}
SIM_NOW  = {'1-9': 14.07, '10-19': 43.51, '20-39': 86.85, '40-99': 180.67, '100+': 201.74}
SIM_NEW  = {'1-9': 23.61, '10-19': 60.62, '20-39': 119.36, '40-99': 226.11, '100+': 274.05}

# ---- 'Raw Data' column letters (see sqls/daily_gains.py HEADER_NOTES for meanings) -------------
C_GROUP, C_BUCKET, C_PLAYERS = 'B', 'C', 'D'
C_HC_TOTAL = 'Q'                       # avg_hc_gain_total
C_STREAK   = {'p50': 'AF', 'p75': 'AG', 'p90': 'AH'}
C_PASS     = ['AI', 'AJ', 'AK']        # pct_player_days_passed_R1/R2/R3 (percent units)
C_ACTIVE   = 'AM'                      # actual_liveops_active_rate_pct
C_FIN      = ['AN', 'AO', 'AP']        # actual_pct_finished_R1/R2/R3

# ---- style vocabulary (read off LiveOps_v2_AB_Summary.xlsx) -----------------------------------
NAVY, BAND, SUBHDR = 'FF1F3864', 'FF4472C4', 'FF305496'
CAT_FILL, WHITE, INPUT_FILL = 'FFD9E1F2', 'FFFFFFFF', 'FFFFF2CC'
KEY_FILL, NOTE_GREY = 'FFF2F2F2', 'FF7F7F7F'
THIN = Side(style='thin', color='FFBFBFBF')
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
F_TITLE = Font(name='Arial', size=14, bold=True, color=NAVY)
F_SUB   = Font(name='Arial', size=9, italic=True, color=NOTE_GREY)
F_HDR   = Font(name='Arial', size=11, bold=True, color='FFFFFFFF')
F_KEY   = Font(name='Arial', size=8, color=NOTE_GREY)
F_CAT   = Font(name='Arial', size=10, bold=True, color=NAVY)
F_ROW   = Font(name='Arial', size=10)
F_ROWB  = Font(name='Arial', size=10, bold=True)

FMT_PCT   = '+0.00%;-0.00%;0.00%'      # true fractions (Δ%)
FMT_NUM   = '#,##0'
FMT_HC    = '#,##0.00'
FMT_RATE  = '0.00'                     # percent-unit values carried as-is from Raw Data
FMT_X     = '0.000'

wb = Workbook()
ws = wb.active
ws.title = 'NS Proposal'
ws.sheet_view.showGridLines = False

FIRST_COL = 3                          # C
COLS = {name: get_column_letter(FIRST_COL + i) for i, (name, _) in enumerate(SEGS)}


def put(row, col, value, font=F_ROW, fill=None, fmt=None, align=None, border=True):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font
    if fill:
        c.fill = PatternFill('solid', fgColor=fill)
    if fmt:
        c.number_format = fmt
    if align:
        c.alignment = Alignment(horizontal=align)
    if border:
        c.border = BOX
    return c


def category(row, label):
    """A D9E1F2 band spanning the full width, like the A/B sheets' Category rows."""
    put(row, 1, label, font=F_CAT, fill=CAT_FILL)
    for col in range(2, FIRST_COL + len(SEGS)):
        put(row, col, None, fill=CAT_FILL)


def metric(row, label, formulas, fmt=FMT_HC, font=F_ROW, fill=WHITE, note=None):
    """One metric row: label in B, one value per segment column."""
    put(row, 1, None, fill=fill)
    put(row, 2, label, font=font, fill=fill)
    for name, _ in SEGS:
        col = FIRST_COL + [s[0] for s in SEGS].index(name)
        put(row, col, formulas(name), font=font, fill=fill, fmt=fmt, align='right')
    if note:
        c = ws.cell(row=row, column=FIRST_COL + len(SEGS) + 1, value=note)
        c.font = F_SUB


def wavg(col_letter, group, bucket_cell):
    """Player-weighted mean of a Raw Data column for one arm x bucket — the By Bucket idiom."""
    r = lambda c: "'Raw Data'!${0}$2:${0}$9999".format(c)
    return ('=IFERROR(SUMPRODUCT(({g}="{arm}")*({b}={key})*{p}*{v})'
            '/SUMIFS(\'Raw Data\'!${pc}:${pc}, \'Raw Data\'!${gc}:${gc}, "{arm}", '
            '\'Raw Data\'!${bc}:${bc}, {key}),"")').format(
        g=r(C_GROUP), b=r(C_BUCKET), p=r(C_PLAYERS), v=r(col_letter),
        arm=group, key=bucket_cell, pc=C_PLAYERS, gc=C_GROUP, bc=C_BUCKET)


# ============================== HEADER =========================================================
ws['A1'] = 'Night Sky coin uplift - proposal and HC impact'
ws['A1'].font = F_TITLE
ws['A2'] = ('Sizes a Night Sky COIN increase that pulls HC gain back into the -4%..0% band and re-steepens '
            'the streak ladder. Observed values are live formulas over \'Raw Data\' (copy this tab INTO '
            'LiveOps_v2_AB_Summary.xlsx - they cannot resolve standalone). Yellow cells are inputs: retype a '
            'reward and everything below re-prices. Requirements are unchanged; coins only.')
ws['A2'].font = F_SUB

HDR_ROW, KEY_ROW = 4, 5
put(HDR_ROW, 1, 'Block', font=F_HDR, fill=SUBHDR, align='center')
put(HDR_ROW, 2, 'Metric', font=F_HDR, fill=SUBHDR, align='center')
for name, key in SEGS:
    col = FIRST_COL + [s[0] for s in SEGS].index(name)
    put(HDR_ROW, col, name, font=F_HDR, fill=BAND, align='center')
    put(KEY_ROW, col, key, font=F_KEY, fill=KEY_FILL, align='center')
put(KEY_ROW, 1, None, fill=KEY_FILL)
put(KEY_ROW, 2, "'Raw Data' bucket key", font=F_KEY, fill=KEY_FILL)
ws.freeze_panes = 'C6'

r = 6
# ============================== 1. CONFIG ======================================================
category(r, 'NS CONFIG'); r += 1
CFG_REQ_ROW = r
for i, rd in enumerate(['R1', 'R2', 'R3']):
    metric(r, 'Cum streak req - %s' % rd, lambda n, i=i: REQ[n][i], fmt=FMT_NUM,
           fill=INPUT_FILL, note='unchanged by this proposal' if i == 0 else None)
    r += 1
CFG_CUR_ROW = r
for i, rd in enumerate(['R1', 'R2', 'R3']):
    metric(r, 'Current HC reward - %s' % rd, lambda n, i=i: CUR[n][i], fmt=FMT_NUM, fill=INPUT_FILL,
           note='live NS ladder; identical in NS_v2, so NS was NOT changed by the A/B' if i == 0 else None)
    r += 1
CFG_NEW_ROW = r
for i, rd in enumerate(['R1', 'R2', 'R3']):
    metric(r, 'PROPOSED HC reward - %s' % rd, lambda n, i=i: PROP[n][i], fmt=FMT_NUM,
           font=F_ROWB, fill=INPUT_FILL)
    r += 1
CFG_DELTA_ROW = r
for i in range(3):
    metric(r, 'Delta HC - R%d' % (i + 1),
           lambda n, i=i: '={c}{new}-{c}{cur}'.format(c=COLS[n], new=CFG_NEW_ROW + i, cur=CFG_CUR_ROW + i),
           fmt=FMT_NUM)
    r += 1

# ============================== 2. OBSERVED ====================================================
r += 0
category(r, 'OBSERVED (variant arm)'); r += 1
PASS_ROW = r
for i, rd in enumerate(['R1', 'R2', 'R3']):
    metric(r, 'Claim rate / player-day - %s (%%)' % rd,
           lambda n, i=i: wavg(C_PASS[i], 'Variant', '%s$%d' % (COLS[n], KEY_ROW)), fmt=FMT_RATE,
           note='pct_player_days_passed_R1/2/3 - the per-day claim probability' if i == 0 else None)
    r += 1
STREAK_ROW = r
for pct in ['p50', 'p75', 'p90']:
    metric(r, 'Daily streak %s' % pct,
           lambda n, pct=pct: wavg(C_STREAK[pct], 'Variant', '%s$%d' % (COLS[n], KEY_ROW)), fmt=FMT_RATE)
    r += 1
metric(r, 'NS active rate (%)',
       lambda n: wavg(C_ACTIVE, 'Variant', '%s$%d' % (COLS[n], KEY_ROW)), fmt=FMT_RATE); r += 1
for i, rd in enumerate(['R1', 'R2', 'R3']):
    metric(r, '%s finished (%%) [actual]' % rd,
           lambda n, i=i: wavg(C_FIN[i], 'Variant', '%s$%d' % (COLS[n], KEY_ROW)), fmt=FMT_RATE)
    r += 1

# ============================== 3. WHERE EACH ROUND SITS =======================================
category(r, 'WHERE EACH ROUND SITS'); r += 1
for i, rd in enumerate(['R1', 'R2', 'R3']):
    metric(r, '%s req vs streak distribution' % rd,
           lambda n, i=i: '=IF({c}{q}<{c}{p50},"below p50 - no streak pull",'
                          'IF({c}{q}<{c}{p75},"p50-p75 - reachable stretch",'
                          'IF({c}{q}<{c}{p90},"p75-p90 - stretch",">p90 - aspiration")))'.format(
               c=COLS[n], q=CFG_REQ_ROW + i, p50=STREAK_ROW, p75=STREAK_ROW + 1, p90=STREAK_ROW + 2),
           fmt='General', note='raising a below-p50 rung is pure HC inflation' if i == 0 else None)
    r += 1

# ============================== 4. NS HC PER PLAYER-DAY ========================================
category(r, 'NS HC PER PLAYER-DAY'); r += 1
NS_NOW_ROW = r
metric(r, 'NS HC now (sum of claim rate x reward)',
       lambda n: '=SUMPRODUCT({c}{p}:{c}{p2}/100,{c}{v}:{c}{v2})'.format(
           c=COLS[n], p=PASS_ROW, p2=PASS_ROW + 2, v=CFG_CUR_ROW, v2=CFG_CUR_ROW + 2)); r += 1
NS_NEW_ROW = r
metric(r, 'NS HC with proposal',
       lambda n: '=SUMPRODUCT({c}{p}:{c}{p2}/100,{c}{v}:{c}{v2})'.format(
           c=COLS[n], p=PASS_ROW, p2=PASS_ROW + 2, v=CFG_NEW_ROW, v2=CFG_NEW_ROW + 2)); r += 1
NS_DIFF_ROW = r
metric(r, 'Delta NS HC / player-day',
       lambda n: '={c}{a}-{c}{b}'.format(c=COLS[n], a=NS_NEW_ROW, b=NS_NOW_ROW), font=F_ROWB); r += 1
NS_UPLIFT_ROW = r
metric(r, 'NS uplift %',
       lambda n: '=IFERROR({c}{a}/{c}{b}-1,"")'.format(c=COLS[n], a=NS_NEW_ROW, b=NS_NOW_ROW),
       fmt=FMT_PCT); r += 1

# ============================== 5. TOTAL HC GAIN ===============================================
category(r, 'TOTAL HC GAIN / PLAYER'); r += 1
HC_CTRL_ROW = r
metric(r, 'Control',
       lambda n: wavg(C_HC_TOTAL, 'Control', '%s$%d' % (COLS[n], KEY_ROW))); r += 1
HC_VAR_ROW = r
metric(r, 'Variant (as tested)',
       lambda n: wavg(C_HC_TOTAL, 'Variant', '%s$%d' % (COLS[n], KEY_ROW))); r += 1
HC_NEWV_ROW = r
metric(r, 'Variant + proposal',
       lambda n: '={c}{v}+{c}{d}'.format(c=COLS[n], v=HC_VAR_ROW, d=NS_DIFF_ROW), font=F_ROWB); r += 1
NS_SHARE_ROW = r
metric(r, 'NS share of HC gain (now)',
       lambda n: '=IFERROR({c}{ns}/{c}{v},"")'.format(c=COLS[n], ns=NS_NOW_ROW, v=HC_VAR_ROW),
       fmt='0.0%'); r += 1
D_NOW_ROW = r
metric(r, 'Delta % as tested',
       lambda n: '=IFERROR({c}{v}/{c}{ctl}-1,"")'.format(c=COLS[n], v=HC_VAR_ROW, ctl=HC_CTRL_ROW),
       fmt=FMT_PCT); r += 1
D_NEW_ROW = r
metric(r, 'Delta % with proposal', font=F_ROWB,
       formulas=lambda n: '=IFERROR({c}{v}/{c}{ctl}-1,"")'.format(c=COLS[n], v=HC_NEWV_ROW, ctl=HC_CTRL_ROW),
       fmt=FMT_PCT); r += 1
BAND_ROW = r
metric(r, 'In target band (-4% .. 0%)?',
       lambda n: '=IF({c}{d}="","",IF(AND({c}{d}>=-0.04,{c}{d}<=0),"YES","NO"))'.format(c=COLS[n], d=D_NEW_ROW),
       fmt='General', font=F_ROWB, note='static estimate - behavioural response is upside on top'); r += 1

# ============================== 6. MARGINAL HC PER WIN =========================================
category(r, 'MARGINAL HC PER EXTRA WIN'); r += 1
MARG_ROW = r
for i, rd in enumerate(['R1', 'R2', 'R3']):
    for which, row0, fnt in (('now', CFG_CUR_ROW, F_ROW), ('proposed', CFG_NEW_ROW, F_ROWB)):
        if i == 0:
            f = lambda n, row0=row0: '=IFERROR({c}{v}/{c}{q},"")'.format(
                c=COLS[n], v=row0, q=CFG_REQ_ROW)
        else:
            f = lambda n, row0=row0, i=i: '=IFERROR(({c}{v}-{c}{vp})/({c}{q}-{c}{qp}),"")'.format(
                c=COLS[n], v=row0 + i, vp=row0 + i - 1, q=CFG_REQ_ROW + i, qp=CFG_REQ_ROW + i - 1)
        metric(r, '%s - %s' % (rd, which), f, fmt='0.00', font=fnt)
        r += 1
# Top-rung pull = R3's marginal rate as a multiple of R2's. Above 1.00 the deepest push is the
# best-paid win in the ladder; below 1.00 the player is better off stopping at R2. Shown as a
# ratio rather than a pass/fail verdict because the useful signal is the MOVEMENT: the proposal
# roughly doubles it at 1-9 and 100+ without every ladder reaching 1.00.
INV_NOW_ROW = r
metric(r, 'Top-rung pull (R3/R2 per win) - now',
       lambda n: '=IFERROR({c}{r3}/{c}{r2},"")'.format(c=COLS[n], r3=MARG_ROW + 4, r2=MARG_ROW + 2),
       fmt='0.00', font=F_ROWB,
       note='<1.00 = the last rung pays less per win than the one before it'); r += 1
INV_NEW_ROW = r
metric(r, 'Top-rung pull - proposed',
       lambda n: '=IFERROR({c}{r3}/{c}{r2},"")'.format(c=COLS[n], r3=MARG_ROW + 5, r2=MARG_ROW + 3),
       fmt='0.00', font=F_ROWB); r += 1

# ============================== 7. CROSS-CHECK =================================================
category(r, 'CROSS-CHECK vs ECONOMY SIM (static)'); r += 1
metric(r, 'R(HC) - survival model, EcoGainsSim_v4 (static)',
       lambda n: SIM_R[n], fmt=FMT_X,
       note='E(NS_v2)/E(NS) over data_streaks x NS_STREAK_N 1.25, workbook (14)'); r += 1
SIM_R_ROW = r - 1
metric(r, 'R(HC) - implied by A/B claim rates',
       lambda n: '=IFERROR({c}{a}/{c}{b},"")'.format(c=COLS[n], a=NS_NEW_ROW, b=NS_NOW_ROW),
       fmt=FMT_X); r += 1
metric(r, 'Agreement (A/B vs model)',
       lambda n: '=IFERROR({c}{a}/{c}{b}-1,"")'.format(c=COLS[n], a=r - 1, b=SIM_R_ROW),
       fmt=FMT_PCT, note='two independent methods - agreement is the validation'); r += 1
metric(r, 'Sim NS row now (HC/earner, 33d, NONPAYER) (static)', lambda n: SIM_NOW[n]); r += 1
metric(r, 'Sim NS row with proposal (static)', lambda n: SIM_NEW[n], font=F_ROWB); r += 1

# ============================== NOTES ==========================================================
r += 1
notes = [
    'How this is sized: NS HC per player-day = SUM over rounds of (claim rate x reward), taken from '
    "'Raw Data' pct_player_days_passed_R1/2/3 - the same units as avg_hc_gain_total, so no basis "
    'conversion is involved. The proposal solves for a landing near -3%, deliberately short of neutral.',
    'Why coins on the deep rungs: a rung below the segment median streak is already cleared on most '
    'active days (claim rates 60-74%), so raising it buys HC with no behavioural pull. The increase '
    'goes to the rung players are stretching for, which also re-steepens HC-per-extra-win.',
    'Top-rung pull: the proposal roughly doubles it at 1-9 (0.30 -> 0.60) and 100+ (0.45 -> 0.69) and '
    'lifts 20-39 (1.34 -> 1.53), but only 20-39 and 40-99 end above 1.00 - at 1-9, 10-19 and 100+ the '
    'R2 rung still pays the best marginal rate. Pushing them past 1.00 costs little (100+ R3 815 '
    'instead of 700 => -3.55% instead of -3.73%); type it into the input block to see it.',
    'Upside not modelled (this is why the target is -3%, not -1%): if streak chasing responds, the '
    'claim rates themselves rise, and more attempts feed Core HC. Both push the result toward 0.',
    'FLAGGED - rewards may not be the whole fix: R2/R3 completions fell 20-49% in the mid/high buckets '
    'while streaks fell only 3-8%, so the requirement ladder also got harder relative to the population. '
    'Coins fix the HC gap and the marginal incentive, not reachability.',
    'FLAGGED - 100+: the modelled per-day claim rate (~30% clear R1) sits well below the telemetry share '
    'of players who finished R1 at least once (~56%; different denominators). If the true daily rate is '
    'higher, the 100+ HC bill runs above this estimate - which is why 100+ is targeted at the -4% edge.',
    'Static cells are marked "(static)" and come from EcoGainsSim_v4.gs over NEW_LIVEOPS_CALENDAR_ECO (14); '
    'every other value on this sheet is computed from \'Raw Data\'.',
]
for n in notes:
    c = ws.cell(row=r, column=1, value=n)
    c.font = F_SUB
    c.alignment = Alignment(wrap_text=False)
    r += 1

# ============================== conditional formatting =========================================
last = get_column_letter(FIRST_COL + len(SEGS) - 1)
green = dict(font=Font(color='FF006100'), fill=PatternFill(bgColor='FFC6EFCE'))
red = dict(font=Font(color='FF9C0006'), fill=PatternFill(bgColor='FFFFC7CE'))
amber = dict(font=Font(color='FF9C6500'), fill=PatternFill(bgColor='FFFFEB9C'))
# 'as tested' — plain worse/better
for row in (D_NOW_ROW,):
    rng = 'C%d:%s%d' % (row, last, row)
    ws.conditional_formatting.add(rng, CellIsRule(operator='greaterThan', formula=['0'], **green))
    ws.conditional_formatting.add(rng, CellIsRule(operator='lessThan', formula=['0'], **red))
# 'with proposal' — the band is the goal: inside = green, below = red, ABOVE 0 = amber (overshoot)
rng = 'C%d:%s%d' % (D_NEW_ROW, last, D_NEW_ROW)
ws.conditional_formatting.add(rng, CellIsRule(operator='between', formula=['-0.04', '0'], **green))
ws.conditional_formatting.add(rng, CellIsRule(operator='lessThan', formula=['-0.04'], **red))
ws.conditional_formatting.add(rng, CellIsRule(operator='greaterThan', formula=['0'], **amber))
# band check (row numbers captured at build time — never re-derived from `r`)
rng = 'C%d:%s%d' % (BAND_ROW, last, BAND_ROW)
ws.conditional_formatting.add(rng, CellIsRule(operator='equal', formula=['"YES"'], **green))
ws.conditional_formatting.add(rng, CellIsRule(operator='notEqual', formula=['"YES"'], **amber))
# top-rung pull: >= 1 means the deepest push is the best-paid win in the ladder
for row in (INV_NOW_ROW, INV_NEW_ROW):
    rng = 'C%d:%s%d' % (row, last, row)
    ws.conditional_formatting.add(rng, CellIsRule(operator='greaterThanOrEqual', formula=['1'], **green))
    ws.conditional_formatting.add(rng, CellIsRule(operator='lessThan', formula=['1'], **amber))
# NS uplift — informational, always a gain
rng = 'C%d:%s%d' % (NS_UPLIFT_ROW, last, NS_UPLIFT_ROW)
ws.conditional_formatting.add(rng, CellIsRule(operator='greaterThan', formula=['0'], **green))

# ============================== geometry =======================================================
ws.column_dimensions['A'].width = 24
ws.column_dimensions['B'].width = 42
for name, _ in SEGS:
    ws.column_dimensions[COLS[name]].width = 13
ws.column_dimensions[get_column_letter(FIRST_COL + len(SEGS) + 1)].width = 4

wb.save(OUT)
print('written', os.path.abspath(OUT))
print('rows used:', r, '| segments:', ', '.join(n for n, _ in SEGS))
