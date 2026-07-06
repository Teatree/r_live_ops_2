# Builds EcoGainsSim_HC_v4.xlsx — display sheet for EcoGainsSim_v4.gs:
# 5 segment blocks + A. 0 appendix block (25 categories each) + assumption legend.
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule

DISPLAY = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'display')

CATS = ['Ads', 'Bomb Challenge', "Bomb's Ballet", 'Chuck Challenge', 'Core', 'Daily Gift',
        'Daily Night Sky Prize', 'Flock Flurry', 'Hatchling Hideaway', 'Jigsaw', 'Kite Festival',
        'Level Race', 'Other', 'Photoshoot', 'Red Challenge', 'River Rush', 'Saga',
        'Season Pass (Free)', 'Target Day', 'Team Event', 'Team Race', 'Flash Race',
        'FlowerCoop', 'Rainbow Maker', 'IAPs']
RES = ['HC', 'Slingshot', 'Shuffle', 'Comet', 'Red', 'Chuck', 'Bomb',
       'UL Bomb', 'UL Chuck', 'UL Red', 'Unlimited Lives']
ALWAYS_ON = {'Daily Gift', 'Saga', 'Daily Night Sky Prize'}   # yellow label (always-on/daily sim)
EVENT_SIM = {'Bomb Challenge', "Bomb's Ballet", 'Chuck Challenge', 'Hatchling Hideaway', 'Jigsaw',
             'Kite Festival', 'Level Race', 'Photoshoot', 'Red Challenge', 'River Rush',
             'Target Day', 'Flash Race', 'Rainbow Maker'}     # pink label (calendar event sim)
A0_APPLIED = {'Saga': 'FFFFF2CC', 'Daily Gift': 'FFFFF2CC', 'River Rush': 'FFF4CCCC'}

BLOCKS = [('0-9',   '0-9 Segment - Over Time Period  (uses 1-9 gains data)', 'main'),
          ('10-19', '10-19 Segment - Over Time Period', 'main'),
          ('20-39', '20-39 Segment - Over Time Period', 'main'),
          ('40-99', '40-99 Segment - Over Time Period', 'main'),
          ('100+',  '100+ Segment - Over Time Period', 'main'),
          ('A. 0',  'A. 0 Appendix - carried & annotated (not simulated)', 'appendix')]

F_BLUE, F_GRAY = 'FFDDEBF7', 'FFEFEFEF'
F_YELLOW, F_PINK, F_ORANGE, F_HDR = 'FFFFF2CC', 'FFF4CCCC', 'FFFCE5CD', 'FFD9D9D9'
fill = lambda rgb: PatternFill('solid', fgColor=rgb)
thin = Side(style='thin')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
FMT_SIM, FMT_DIFF = '#,##0.00;[Red]-#,##0.00', '#,##0.00;-#,##0.00'

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'EcoGainsSim_HC'
ws.sheet_view.showGridLines = False
for col, w in {'A': 1.75, 'B': 19.25, 'C': 8.25, 'Z': 7.63}.items():
    ws.column_dimensions[col].width = w

ws['B2'] = 'Per-earner gains over the 33-day'
ws['B2'].font = Font(name='Arial', size=12, bold=True)
ws['B3'] = 'Payer'
ws['B3'].font = Font(name='Arial', size=10, bold=True)
ws['C3'] = 'PAYER'
ws['C3'].font = Font(name='Arial', size=10, bold=True)
ws['C3'].fill = fill(F_YELLOW)
ws['D3'] = '◀ input (global)'
ws['D3'].font = Font(name='Arial', size=9, color='FF808080')

N = len(CATS)                       # 25
PITCH = N + 4                       # 29 rows per block
first_hdr = 6
sim_rows = []
for t, (seg, title, kind) in enumerate(BLOCKS):
    hdr = first_hdr + t * PITCH     # 6, 35, 64, 93, 122, 151
    cols_row, d0 = hdr + 1, hdr + 2
    tag = ws.cell(hdr, 2, seg)
    tag.font = Font(name='Arial', size=10, bold=True)
    tag.fill = fill(F_ORANGE)
    tag.alignment = Alignment(horizontal='center')
    ws.cell(hdr, 3, title).font = Font(name='Arial', size=11, bold=True)
    dtitle = (title.split('  (')[0] + ' - Difference') if kind == 'main' else 'A. 0 Appendix - Difference (config-only changes)'
    ws.cell(hdr, 15, dtitle).font = Font(name='Arial', size=11, bold=True)
    hf = Font(name='Arial', size=9, bold=True)
    ws.cell(cols_row, 2, 'Source').font = hf
    ws.cell(cols_row, 2).fill = fill(F_HDR)
    ws.cell(cols_row, 2).border = BORDER
    for j, r in enumerate(RES):
        for base in (3, 15):
            cell = ws.cell(cols_row, base + j, r)
            cell.font, cell.fill, cell.border = hf, fill(F_HDR), BORDER
            cell.alignment = Alignment(horizontal='center')
    for i, cat in enumerate(CATS):
        row = d0 + i
        lab = ws.cell(row, 2, cat)
        lab.font = Font(name='Arial', size=9, bold=True)
        lab.border = BORDER
        if kind == 'appendix':
            sim_cell = cat in A0_APPLIED
            if sim_cell:
                lab.fill = fill(A0_APPLIED[cat])
            data_fill = fill(F_GRAY) if sim_cell else fill(F_BLUE)
        else:
            if cat in ALWAYS_ON:
                lab.fill = fill(F_YELLOW)
            elif cat in EVENT_SIM:
                lab.fill = fill(F_PINK)
            data_fill = fill(F_GRAY) if (cat in ALWAYS_ON or cat in EVENT_SIM) else fill(F_BLUE)
        for j in range(len(RES)):
            c1 = ws.cell(row, 3 + j)
            c1.fill, c1.border, c1.number_format = data_fill, BORDER, FMT_SIM
            c2 = ws.cell(row, 15 + j)
            c2.border, c2.number_format = BORDER, FMT_DIFF
        if t == 0 and (cat in ALWAYS_ON or cat in EVENT_SIM):
            sim_rows.append(row)
    ws.cell(d0, 3).value = f'=LET(payer, $C$3, segment, $B${hdr}, ECOGAINS_SIM(payer, segment))'
    ws.cell(d0, 15).value = f'=LET(payer, $C$3, segment, $B${hdr}, ECOGAINS_DIFF(payer, segment))'
    diff_rng = f'O{d0}:Y{d0 + N - 1}'
    ws.conditional_formatting.add(diff_rng, CellIsRule(
        operator='lessThan', formula=['0'],
        font=Font(color='FF990000', bold=False), fill=fill(F_PINK)))
    ws.conditional_formatting.add(diff_rng, CellIsRule(
        operator='greaterThan', formula=['0'],
        font=Font(color='FF006100', bold=False), fill=fill('FFD9EAD3')))
    for r in range(hdr, d0 + N):
        ws.row_dimensions[r].height = 15.75

ws['J3'] = '=JOIN(",", {' + ','.join(f'B{r}' for r in sim_rows) + '})'

# ---- assumption legend (static — flags from SIMULATION_PLAN) ----
leg0 = first_hdr + len(BLOCKS) * PITCH + 1   # row after last block + gap
legend = [
    'LEGEND / ASSUMPTIONS (EcoGainsSim_v4)',
    'Colors: blue = measured (carried) · gray = simulated · yellow label = always-on/daily sim · pink label = calendar-event sim · diff: red < 0 < green.',
    'Night Sky: CARRIED by default (NS_SIMULATE = false in EcoGainsSim_v4.gs - the bottom-up model overestimates actual NS gains; open question). Flag true -> full-rollout sim (segment ladder x survival over data_streaks max-streak p25-p90 x 1.25 effective-streak factor x active days); CURRENT is A/B-diluted, so DIFF would read as ROLLOUT EFFECT.',
    'Rainbow Maker & Night Sky magnitudes are tail-sensitive: milestones past p90 are priced by linear extrapolation (conservative S=0-beyond-p90 bounds available in the offline harness).',
    "Target Day: pure leaderboard (milestones pay 0 by design); D=1 rank-invariance assumed — same assumption as the bird challenges. Rainbow Maker's clipped 2-day instance scales the matchables axis x(2/4).",
    'Photoshoot: single instance in both calendars, so T is placement-sensitive. River Rush: simulated 0 because cal_new has no instances (removal) — re-add instances and it re-prices.',
    "A. 0 block: no behaviour data (these players barely play) — everything carried except config-only changes: Saga HC x ratio, Daily Gift HC x ratio (0-9 streaks as PROXY, slightly overstates), River Rush -> 0. RM/NS not simulated for A. 0.",
    'Reward-config edits FLOW (since 2026-07-06): editing rewards or requirements on any _v2 sheet reprices its row (R = v2/base ladder at the measured rank / progress distribution). Base sheets = the measured world; leave them untouched. TaD milestone rewards are the exception (base pays 0 -> no anchor -> carried).',
    'Kite Festival is priced as a LEADERBOARD (since 2026-07-06): payouts are zero-sum per league of 60, so duration does not move them; D=1 and the row GROWS ~x1.3 via cadence.',
    'After editing a calendar: menu EcoGainsSim ▸ Precompute calendars (writes hidden cal_parsed; engine prefers it). Sanity canary: the Kite Festival row must GROW ~x1.3 vs measured. If every event row equals measured, the calendar read failed.',
]
for i, txt in enumerate(legend):
    c = ws.cell(leg0 + i, 2, txt)
    c.font = Font(name='Arial', size=9, bold=(i == 0), color='FF000000' if i == 0 else 'FF808080')

wb.save(os.path.join(DISPLAY, 'EcoGainsSim_HC_v4.xlsx'))
print('written EcoGainsSim_HC_v4.xlsx')
print('block hdr rows:', [first_hdr + t * PITCH for t in range(len(BLOCKS))])
print('J3 sim rows:', sim_rows)
print('legend at row', leg0)
