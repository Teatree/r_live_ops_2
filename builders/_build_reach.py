# Builds EventReach_v1.xlsx — replicas of J_v2 / HH_v2 / BB_v2 / Ph_v2 / RM (values + styles)
# from NEW_LIVEOPS_CALENDAR_ECO (5).xlsx, each extended with a "player reach" table at column AH
# driven purely by Google Sheets formulas (data_event_inst / data_event_accrual / data_RM lookups).
from copy import copy
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter as CL

HERE = os.path.dirname(os.path.abspath(__file__))
SRC = os.path.join(HERE, '..', 'workbooks', 'NEW_LIVEOPS_CALENDAR_ECO (5).xlsx')
OUT = os.path.join(HERE, '..', 'display', 'EventReach_v1.xlsx')
SEGS = ['0-9', '10-19', '20-39', '40-99', '100+']
SEG_HDR = ['0-9', '10-19', '20-39', '40-99', '100']       # header labels per user spec ('100_ms')

F_YELLOW, F_GRAY, F_HDRGRAY = 'FFFFF2CC', 'FFEFEFEF', 'FFD9D9D9'
fill = lambda rgb: PatternFill('solid', fgColor=rgb)
thin = Side(style='thin')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

# percentile -> data_event_accrual share column (Avg=mean I, p25=K, p50=J, p75=L)
ACCR_COLS = [('Avg', 'I'), ('p25', 'K'), ('p50', 'J'), ('p75', 'L')]
# percentile -> data_RM column (avg C, p25 E, p50 F, p75 G, p90 H)
RM_COLS = [('Avg', 'C'), ('p25', 'E'), ('p50', 'F'), ('p75', 'G'), ('p90', 'H')]

CFG = {
    'J_v2': dict(kind='token', event='Jigsaw', dur='$B$3',
                 req='$B$11:$B$22', rew='$C$11:$W$22', hdr='$C$10:$W$10', completion=False),
    'HH_v2': dict(kind='token', event='Hatchling Hideaway', dur='$B$3',
                  req='$AV$5:$AV$9', rew='$B$12:$V$16', hdr='$B$11:$V$11', completion=False,
                  helper='hh'),
    'BB_v2': dict(kind='token', event='Bombs Ballet', dur='$B$3',
                  req='$B$9:$B$23', rew='$C$9:$W$24', hdr='$C$8:$W$8', completion=True),
    'Ph_v2': dict(kind='token', event='Photoshoot', dur='$B$3',
                  req='$AU$5:$AU$34', rew='$H$25:$AB$54', hdr='$H$24:$AB$24', completion=False,
                  helper='ph'),
    'RM': dict(kind='rm', dur='$B$3',
               req='$C$16:$C$45', rew='$E$16:$Y$45', hdr='$E$15:$Y$15', completion=False),
}
NCOLS = 21   # every reward block spans 21 columns (Coins .. 6-star Dly)

src = openpyxl.load_workbook(SRC)          # formulas view (these sheets are values anyway)
out = openpyxl.Workbook()
out.remove(out.active)

def copy_sheet(name):
    s, t = src[name], out.create_sheet(name)
    t.sheet_view.showGridLines = s.sheet_view.showGridLines
    for row in s.iter_rows():
        for c in row:
            if c.value is None and not c.has_style:
                continue
            n = t.cell(c.row, c.column, c.value)
            if c.has_style:
                n.font = copy(c.font); n.fill = copy(c.fill); n.border = copy(c.border)
                n.alignment = copy(c.alignment); n.number_format = c.number_format
    for k, dim in s.column_dimensions.items():
        if dim.width: t.column_dimensions[k].width = dim.width
        if dim.hidden: t.column_dimensions[k].hidden = True
    for k, dim in s.row_dimensions.items():
        if dim.height: t.row_dimensions[k].height = dim.height
    for m in s.merged_cells.ranges:
        t.merge_cells(str(m))
    return t

def accr_lookup(col, event, day_ref):
    return (f"AVERAGEIFS(data_event_accrual!${col}:${col},"
            f"data_event_accrual!$A:$A,\"{event}\","
            f"data_event_accrual!$B:$B,$AI$3,"
            f"data_event_accrual!$C:$C,\"{{SEG}}\","
            f"data_event_accrual!$E:$E,{day_ref})")

def rm_lookup(col):
    return (f"SUMIFS(data_RM!${col}:${col},data_RM!$A:$A,\"{{SEG}}\",data_RM!$B:$B,$AI$3)")

def build_block(ws, cfg):
    # title + inputs
    ws['AH1'] = 'Player Reach Simulation (per event day) — SIMULATED'
    ws['AH1'].font = Font(name='Arial', size=11, bold=True)
    pcts = [p for p, _ in (RM_COLS if cfg['kind'] == 'rm' else ACCR_COLS)]
    for r, (label, default, dv_list) in enumerate([
            ('Percentile', 'Avg', '"' + ','.join(pcts) + '"'),
            ('Payer', 'NONPAYER', '"NONPAYER,PAYER"')]):
        lab = ws.cell(2 + r, 34, label)
        lab.font = Font(name='Arial', size=9, bold=True)
        v = ws.cell(2 + r, 35, default)
        v.font = Font(name='Arial', size=9, bold=True)
        v.fill = fill(F_YELLOW)
        v.alignment = Alignment(horizontal='center')
        v.border = BORDER
        dv = DataValidation(type='list', formula1=dv_list, allow_blank=False, showDropDown=False)
        ws.add_data_validation(dv)
        dv.add(v)
    # header row 5
    hdr_cells = ['Day_of_event']
    for s in SEG_HDR:
        hdr_cells += [f'{s}_ms', f'{s}_reward']
    for j, txt in enumerate(hdr_cells):
        c = ws.cell(5, 34 + j, txt)
        c.font = Font(name='Arial', size=9, bold=True)
        c.fill = fill(F_HDRGRAY)
        c.alignment = Alignment(horizontal='center')
        c.border = BORDER
    # helper ladders (HH / Ph)
    if cfg.get('helper') == 'hh':
        ws['AT2'] = 'accuracy'
        ws['AT2'].font = Font(name='Arial', size=9, bold=True)
        au2 = ws['AU2']; au2.value = 0.8
        au2.fill = fill(F_YELLOW); au2.border = BORDER; au2.number_format = '0%'
        au2.font = Font(name='Arial', size=9, bold=True)
        for j, txt in enumerate(['Gate', 'Tiles', 'Cum tokens req']):
            c = ws.cell(4, 46 + j, txt)
            c.font = Font(name='Arial', size=9, bold=True); c.fill = fill(F_HDRGRAY); c.border = BORDER
        for i, tiles in enumerate([16, 25, 36, 49, 64]):
            r = 5 + i
            ws.cell(r, 46, i + 1).border = BORDER
            tc = ws.cell(r, 47, tiles)
            tc.fill = fill(F_YELLOW); tc.border = BORDER
            fc = ws.cell(r, 48, f'=SUM($AU$5:$AU{r})/$AU$2')
            fc.fill = fill(F_GRAY); fc.border = BORDER; fc.number_format = '#,##0.0'
        note = ws.cell(11, 46, 'treasure-dig: player finds gems ~80% of taps → needs tiles ÷ accuracy tokens per gate')
        note.font = Font(name='Arial', size=8, color='FF808080')
    if cfg.get('helper') == 'ph':
        c = ws.cell(4, 47, 'Cum item price')
        c.font = Font(name='Arial', size=9, bold=True); c.fill = fill(F_HDRGRAY); c.border = BORDER
        for i in range(30):
            r = 5 + i
            ws.cell(r, 46, i + 1).border = BORDER
            fc = ws.cell(r, 47, f'=SUM($F$25:$F{25 + i})')
            fc.fill = fill(F_GRAY); fc.border = BORDER; fc.number_format = '#,##0'
        note = ws.cell(36, 46, 'simplified: items bought sequentially in item-id order (choice ignored)')
        note.font = Font(name='Arial', size=8, color='FF808080')
    # day rows 6..12
    for i in range(7):
        r = 6 + i
        dc = ws.cell(r, 34, f'=IF({i + 1}<={cfg["dur"]},{i + 1},"")')
        dc.fill = fill(F_GRAY); dc.border = BORDER
        dc.alignment = Alignment(horizontal='center')
        for si, seg in enumerate(SEGS):
            ms_col, rw_col = 35 + si * 2, 36 + si * 2
            ms_ref = f'{CL(ms_col)}{r}'
            if cfg['kind'] == 'token':
                shares = ','.join(accr_lookup(col, cfg['event'], f'$AH{r}') for _, col in ACCR_COLS)
                pl = ','.join(f'"{p}"' for p, _ in ACCR_COLS)
                ms_f = (f'=IF($AH{r}="","",LET('
                        f'bal,SUMIFS(data_event_inst!$N:$N,data_event_inst!$A:$A,"{cfg["event"]}",'
                        f'data_event_inst!$B:$B,$AI$3,data_event_inst!$C:$C,"{seg}"),'
                        f'shr,MIN(1,IFERROR(CHOOSE(MATCH($AI$2,{{{pl}}},0),{shares}),1)),'
                        f'COUNTIF({cfg["req"]},"<="&bal*shr)))').replace('{SEG}', seg)
            else:
                mats = ','.join(rm_lookup(col) for _, col in RM_COLS)
                pl = ','.join(f'"{p}"' for p, _ in RM_COLS)
                ms_f = (f'=IF($AH{r}="","",LET('
                        f'mat,IFERROR(CHOOSE(MATCH($AI$2,{{{pl}}},0),{mats}),0)*$AH{r}/{cfg["dur"]},'
                        f'COUNTIF({cfg["req"]},"<="&mat)))').replace('{SEG}', seg)
            mc = ws.cell(r, ms_col, ms_f)
            mc.fill = fill(F_GRAY); mc.border = BORDER; mc.number_format = '0'
            mc.alignment = Alignment(horizontal='center')
            m_expr = f'IF({ms_ref}=15,16,{ms_ref})' if cfg['completion'] else ms_ref
            rw_f = (f'=IF($AH{r}="","",IF({ms_ref}=0,"{{}}",LET(m,{m_expr},'
                    f's,MMULT(SEQUENCE(1,m,1,0),ARRAY_CONSTRAIN({cfg["rew"]},m,{NCOLS})),'
                    f'"{{"&TEXTJOIN(", ",TRUE,ARRAYFORMULA(IF(s>0,{cfg["hdr"]}&": "&ROUND(s,2),"")))&"}}")))')
            rc = ws.cell(r, rw_col, rw_f)
            rc.fill = fill(F_GRAY); rc.border = BORDER
            rc.alignment = Alignment(horizontal='left')
    # widths
    ws.column_dimensions['AG'].width = 2.88
    ws.column_dimensions['AH'].width = 11.0
    for si in range(5):
        ws.column_dimensions[CL(35 + si * 2)].width = 7.0
        ws.column_dimensions[CL(36 + si * 2)].width = 34.0
    ws.column_dimensions['AT'].width = 6.0
    ws.column_dimensions['AU'].width = 8.0
    ws.column_dimensions['AV'].width = 13.0

for name, cfg in CFG.items():
    ws = copy_sheet(name)
    build_block(ws, cfg)

out.save(OUT)
print('written', OUT, '— sheets:', out.sheetnames)
