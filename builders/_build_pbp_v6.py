# Builds EcoGainsSim_PlybyPly_v6.xlsx - the production play-by-play sheet in the workbook (6)
# style the user hand-applied to the live EcoGainsSim_PlybyPly sheet (green = simulation):
#   548235 section bars w/ white bold text; E2EFDA green labels/headers; FFF2CC input values;
#   F3F3F3 profile values; EFEFEF spill areas; 999999 italic notes; 000000 assumptions bar with
#   808080 legend lines; Arial; 9pt chrome / 11pt ledger; no gridlines, no merges, no em dashes.
# Supersedes _build_pbp_v5.py (TaD_v2 black/gold style). Wired to EcoGainsSim_PBP.gs (v3
# signature, openingInv removed):
#   A14: ECOGAINS_PBP_PROFILE   A23: ECOGAINS_PBP_EVENTS   A51: ECOGAINS_PBP
# Spill regions are left EMPTY but pre-styled (incl. the Session Summary rows, which inherit
# the ledger styling - user feedback). Formulas show #NAME? in Excel - expected.
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter as CL

OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'display',
                   'EcoGainsSim_PlybyPly_v6.xlsx')
F_BAR, F_LABEL, F_INPUT, F_VAL, F_SPILL = 'FF548235', 'FFE2EFDA', 'FFFFF2CC', 'FFF3F3F3', 'FFEFEFEF'
fill = lambda rgb: PatternFill('solid', fgColor=rgb)
bar_f = Font(name='Arial', bold=True, color='FFFFFFFF')
lbl_f = Font(name='Arial', size=9, bold=True)
in_f = Font(name='Arial', size=9)
note_f = Font(name='Arial', size=9, italic=True, color='FF999999')
leg_f = Font(name='Arial', size=9, color='FF808080')

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'EcoGainsSim_PlybyPly'
ws.sheet_view.showGridLines = False

def bar(row, text, c0, c1):
    for c in range(c0, c1 + 1):
        cell = ws.cell(row, c)
        cell.fill = fill(F_BAR)
        cell.font = bar_f
    ws.cell(row, c0, text)

def cfg_row(row, label, value, note):
    a = ws.cell(row, 1, label); a.fill = fill(F_LABEL); a.font = lbl_f
    b = ws.cell(row, 2, value); b.fill = fill(F_INPUT); b.font = in_f
    c = ws.cell(row, 3, note);  c.font = note_f

# ---------------- title + config panel ----------------
ws.cell(1, 1, 'EcoGainsSim PlybyPly').font = Font(name='Arial', size=14, bold=True)
bar(2, 'Config Panel', 1, 3)
CONFIG = [
    ('Calendar', 'cal_new',
     'Which calendar the day comes from: cal_curr = measured live calendar, cal_new = redesign'),
    ('Day to Simulate (1-33)', 9,
     'Day of the 33-day window (day 1 = Wednesday); decides which events are running'),
    ('Player Segment', '10-19',
     'Engagement segment: saga completions in the last 7 days (0-9 .. 100+)'),
    ('Payer', 'NONPAYER',
     'Lifetime payer flag: PAYER = has ever spent money, NONPAYER = never'),
    ('Mode', 'Sampled',
     'Expected = one representative day (seed ignored); Sampled = one random day drawn with Seed'),
    ('Luck (progress + placement)', 'p50',
     'Percentile used for event progress and leaderboard rank: p25 unlucky, p50 typical, p75 lucky'),
    ('Seed (Sampled mode)', 32,
     'Any integer; the same seed always reproduces the same sampled day'),
    ('Levels Played (blank = auto)', None,
     'Override for the number of level attempts; blank = attempts/day from data_streaks'),
    ('Starting Saga Level (blank = rnd 100-400)', None,
     'Absolute saga level the player walks in on; anchors Saga node payouts'),
]
for i, (label, value, note) in enumerate(CONFIG):
    cfg_row(3 + i, label, value, note)

dv = [('B3', '"cal_curr,cal_new"'), ('B5', '"0-9,10-19,20-39,40-99,100+"'),
      ('B6', '"NONPAYER,PAYER"'), ('B7', '"Expected,Sampled"'), ('B8', '"p25,p50,p75"')]
for addr, lst in dv:
    v = DataValidation(type='list', formula1=lst, allow_blank=True)
    ws.add_data_validation(v)
    v.add(addr)

# ---------------- player profile (7 rows x 3 cols, notes come from the engine) --------------
bar(13, 'Player Profile (what the simulated player looks like)', 1, 3)
ws.cell(14, 1, '=ECOGAINS_PBP_PROFILE($B$5,$B$6)')
for r in range(14, 21):
    ws.cell(r, 1).fill = fill(F_LABEL); ws.cell(r, 1).font = lbl_f
    ws.cell(r, 2).fill = fill(F_VAL);   ws.cell(r, 2).font = in_f
    ws.cell(r, 3).font = note_f

# ---------------- active events (header + up to 13 rows x 6 cols) ---------------------------
bar(22, 'Active Events on This Day (sim - spills one row per running event + per session-start claim)', 1, 6)
ws.cell(23, 1, '=ECOGAINS_PBP_EVENTS($B$3,$B$4,$B$5,$B$6,$B$8)')
for c in range(1, 7):
    ws.cell(23, c).fill = fill(F_LABEL); ws.cell(23, c).font = lbl_f
for r in range(24, 37):
    ws.cell(r, 1).fill = fill(F_LABEL); ws.cell(r, 1).font = lbl_f
    for c in range(2, 7):
        ws.cell(r, c).fill = fill(F_SPILL); ws.cell(r, c).font = in_f

# ---------------- assumptions & flags ----------------
bar(38, 'Assumptions & Flags (model spec: SIMULATION_METHODOLOGY.md par. 14)', 1, 11)
ws.cell(38, 1).fill = fill('FF000000')
for c in range(1, 12): ws.cell(38, c).fill = fill('FF000000')
LEGEND = [
    '1. N plays / win rate / streak persistence from data_streaks; win draws are a 2-state Markov chain. The sim conditions on the player being active and participating in the running events.',
    '2. The player is NOT fresh: they walk in at the Starting Saga Level (input, or a seeded random 100-400); the Level column is the ABSOLUTE saga level. Event progress is likewise mid-instance.',
    '3. Event progress at session start = measured final_balance percentile (Luck) x accrual-curve share. Milestones banked before today never appear in the ledger.',
    '4. Per-win earning is MECHANICAL where documented: Hatchling Hideaway 1.5 tokens/win (config 1/2/3 by difficulty, level-mix average); Bomb\'s Ballet 5 Notes on FIRST-TRY wins only; Jigsaw Completion Bonus tiers 3/5/7/10 (2021 Valentine\'s origin doc; win steps the tier up, loss steps down, session starts Copper - flagged); Rainbow Maker whole matchables per win.',
    '5. Photoshoot: first-try streak multiplier ladder x1/2/4/6/10 from config; the per-win base is undocumented so it is calibrated to the measured day total (shape = mechanics, level = measurement).',
    '6. Hatchling Hideaway gate unlock requirements = EventReach helper column (board cost x 1.25 bad-tile buffer): the buffer models imperfect tile picks and affects unlock timing ONLY, never the gains.',
    '7. Score events (Kite Festival, Target Day) are streak-driven off their config ladders; per-play increments are scaled so the day total hits the measured target (user-approved).',
    '8. Leaderboard payouts land on row E only for instances ending today; rank = Luck percentile of measured position_p25/50/75 (+/- 0.25 quantile jitter in Sampled mode).',
    '9. Saga pays the FULL node bundle (HC + boosters + Unlimited minutes) from c_saga / c_saga_v2 at node boundaries anchored to the absolute level. Core chapter chests are NOT simulated (cadence unknown).',
    '10. Daily Gift is ALWAYS claimed at session start: cycle day from login streak p50, bundle = ONE config variant (Expected: Variant 1; Sampled: seeded pick) - never an average. Flock Flurry opt-in = 60 min Unlimited Lives (design-PDF constant).',
    '11. Night Sky pays at day end: effective streak = max win streak x 1.25 (Expected: the p50 from data_streaks; Sampled: the longest run actually produced by the play trace). EVERY milestone whose Cum Streak Req is cleared pays, each on its own row; unreached milestones never pay.',
    '12. SPT / COOP Token / Avatar / star-daily rewards are outside the 11-resource universe and are not tracked.',
]
for i, txt in enumerate(LEGEND):
    ws.cell(39 + i, 1, txt).font = leg_f

# ---------------- play-by-play ledger (22 cols, one row per claim; summary inherits style) ---
bar(51, 'Play-by-Play Sim (one row per play; extra claims spill onto their own rows; Session Summary at the end)', 1, 22)
ws.cell(52, 1, '=ECOGAINS_PBP($B$3,$B$4,$B$5,$B$6,$B$7,$B$8,$B$9,$B$10,$B$11)')
for c in range(1, 23):
    ws.cell(52, c).fill = fill(F_LABEL); ws.cell(52, c).font = lbl_f
led_f = Font(name='Arial', size=11)
claims_f = Font(name='Arial', size=11, italic=True)
for r in range(53, 313):     # up to ~147 plays + claim rows + Session Summary, all one style
    for c in range(1, 23):
        cell = ws.cell(r, c)
        cell.fill = fill(F_SPILL)
        cell.font = claims_f if c == 7 else led_f

# ---------------- widths ----------------
widths = {'A': 31.8, 'B': 12.2, 'C': 10.5, 'D': 12.2, 'E': 11.4, 'F': 15.9, 'G': 56.0}
for c in range(8, 12):  widths[CL(c)] = 11.4     # H-K: event-progress slots
for c in range(12, 23): widths[CL(c)] = 7.6      # L-V: 11 resource columns
for col, w in widths.items():
    ws.column_dimensions[col].width = w

wb.save(OUT)
print('written', OUT)
