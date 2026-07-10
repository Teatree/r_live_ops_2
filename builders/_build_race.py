# Builds Race_v1.xlsx — replacement for the workbook's BASE 'Race' config sheet ("All LB
# Challenges"): five blocks (Red's / Chunk's / Bomb's Challenge, Level Challenge, Flash Race).
#
# 2026-07-10 (user request): Flash Race + Level Challenge ladders updated from the REAL live
# server configs (previously a flagged modeling simplification — source_docs/README.md #7).
# The three bird-challenge blocks keep the existing generic ladder (their live configs not yet
# provided). Geometry is IDENTICAL to the current sheet (block pitch 18, ladders 10 rows,
# cols A..V) so engine LB_R_SPECS, PBP specs and V2Diff cell alignment stay valid.
#
# Item mapping from the live config JSON (user-approved 2026-07-10):
#   Coin / HardCoin -> Coins · SeasonPassToken -> SPT · Pre-Red/Chuck/Bomb -> Red/Chuck/Bomb
#   ShootingStar -> Comet · CherryBomb -> Comet · Hammer -> Slingshot (FLAGGED: not a real
#   Slingshot payout — Hammer isn't in the 13-resource universe; user chose the mapping)
#   adReward UnlimitedLives ignored (adEnabled false).
# Net effect vs the old generic ladder: Level Challenge values are UNCHANGED after mapping
# (only LBSize 10 -> 20); Flash Race is the real change (7 paid ranks, SPT 50/50/50/50/45/40/40,
# coins 100/50/25, ranks 8-10 zeroed, LBSize/numberOfPositions -> 7).
#
# ⚠ BASE-ONLY UPDATE (user decision): Race_v2 is NOT regenerated. Until Race_v2 is re-duplicated
# from this new base (keeping its EventDuration cells: 2/2/2/2/1), rewardR_ prices v2's OLD
# generic Flash ladder as a redesign — in particular R_SPT(Flash) = 0 (v2 pays no SPT), which
# zeroes Flash Race's simulated SPT. Re-duplicate Race_v2 in the live workbook to restore R = 1.
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

DISPLAY = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'display')

COLS = ['Position', 'Coins', 'SPT', 'SPT x2', 'Red', 'Chuck', 'Bomb', 'Slingshot', 'Shuffle',
        'Comet', 'Unlimited Lives', 'Unlimited Red', 'Unlimited Chuck', 'Unlimited Bomb',
        'COOP Token', 'Avatar', '1-star Dly', '2-star Dly', '3-star Dly', '4-star Dly',
        '5-star Dly', '6-star Dly']
NCOL = len(COLS)                                   # 22 (A..V)

# ladders: position -> {column header: amount}; 10 rows each, unlisted cells 0 (punch-card)
GENERIC = {1: {'Coins': 200, 'Comet': 1}, 2: {'Coins': 100, 'Shuffle': 1},
           3: {'Coins': 50, 'Shuffle': 1}, 4: {'Shuffle': 1}, 5: {'Shuffle': 1},
           6: {'Slingshot': 1}, 7: {'Slingshot': 1}, 8: {'Slingshot': 1},
           9: {'Slingshot': 1}, 10: {'Slingshot': 1}}
# LevelRace live config (HardCoin->Coins, CherryBomb->Comet, Hammer->Slingshot):
LEVEL = {1: {'Coins': 200, 'Comet': 1}, 2: {'Coins': 100, 'Shuffle': 1},
         3: {'Coins': 50, 'Shuffle': 1}, 4: {'Shuffle': 1}, 5: {'Shuffle': 1},
         6: {'Slingshot': 1}, 7: {'Slingshot': 1}, 8: {'Slingshot': 1},
         9: {'Slingshot': 1}, 10: {'Slingshot': 1}}   # == GENERIC after mapping (kept explicit)
# FlashRace live config (Coin->Coins, ShootingStar->Comet, SeasonPassToken->SPT, Pre-*->birds):
FLASH = {1: {'Coins': 100, 'Comet': 1, 'SPT': 50}, 2: {'Coins': 50, 'Shuffle': 1, 'SPT': 50},
         3: {'Coins': 25, 'Slingshot': 1, 'SPT': 50}, 4: {'Bomb': 1, 'SPT': 50},
         5: {'Chuck': 1, 'SPT': 45}, 6: {'Red': 1, 'SPT': 40}, 7: {'Red': 1, 'SPT': 40},
         8: {}, 9: {}, 10: {}}                     # league of 7 — ranks 8-10 don't exist

# (name, numberOfPositions, LBSize, EventDuration, ladder)
BLOCKS = [("Red's Challenge",   10, 10, 1, GENERIC),
          ("Chunk's Challenge", 10, 10, 1, GENERIC),   # [sic] — live sheet spelling, kept
          ("Bomb's Challenge",  10, 10, 1, GENERIC),
          ('Level Challenge',   10, 20, 1, LEVEL),     # LBSize = leagueGroupSize 20 (live cfg)
          ('Flash Race',         7,  7, 1, FLASH)]     # leagueGroupSize 7 (live cfg)

F_BAR, F_LBL, F_VAL, F_HDR = 'FF000000', 'FFFFD966', 'FFFFF2CC', 'FFF7CB4D'
fill = lambda rgb: PatternFill('solid', fgColor=rgb)
thin = Side(style='thin')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Race'
ws.sheet_view.showGridLines = False
ws.column_dimensions['A'].width = 21.0
ws.column_dimensions['B'].width = 4.5
ws.column_dimensions['W'].width = 8.6

ws['A1'] = 'All LB Challenges'
ws['A1'].font = Font(name='Arial', size=14, bold=True)

def bar(r, text):
    for c in range(1, NCOL + 1):
        ws.cell(r, c).fill = fill(F_BAR)
    cell = ws.cell(r, 1, text)
    cell.font = Font(name='Arial', size=11, bold=True, color='FFFFFFFF')

PITCH = 18
for b, (name, npos, lbsize, dur, ladder) in enumerate(BLOCKS):
    r0 = 3 + b * PITCH                              # panel title rows 3/21/39/57/75
    bar(r0, f'{name} - Config Panel')
    for i, (label, val) in enumerate([('numberOfPositions', npos), ('LBSize', lbsize),
                                      ('EventDuration', dur)]):
        lc = ws.cell(r0 + 1 + i, 1, label)
        lc.font = Font(name='Arial', size=11, bold=True)
        lc.fill = fill(F_LBL)
        vc = ws.cell(r0 + 1 + i, 2, val)
        vc.font = Font(name='Arial', size=11)
        vc.fill = fill(F_VAL)
        vc.alignment = Alignment(horizontal='center')
    bar(r0 + 5, f'{name} - Goals and Rewards')      # rewards title rows 8/26/44/62/80
    hr = r0 + 6                                     # header rows 9/27/45/63/81
    for c, h in enumerate(COLS, 1):
        cell = ws.cell(hr, c, h)
        cell.font = Font(name='Arial', size=11, bold=True)
        cell.fill = fill(F_HDR)
        cell.border = BORDER
        cell.alignment = Alignment(horizontal='center')
    for pos in range(1, 11):                        # ladder rows 10-19/28-37/46-55/64-73/82-91
        rew = ladder[pos]
        row = [pos] + [rew.get(h, 0) for h in COLS[1:]]
        for c, v in enumerate(row, 1):
            cell = ws.cell(hr + pos, c, v)
            cell.font = Font(name='Arial', size=11)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal='center')

# flagged assumptions (config-sheet convention: loud, below the blocks)
notes = [
    'NOTES (2026-07-10 live-config update — see builders/_build_race.py):',
    'Flash Race + Level Challenge ladders come from the REAL server configs; the three bird challenges keep the prior generic ladder (live configs not yet provided).',
    'Item mapping: Coin/HardCoin->Coins · SeasonPassToken->SPT · Pre-Red/Chuck/Bomb->Red/Chuck/Bomb · ShootingStar->Comet · CherryBomb->Comet · Hammer->Slingshot (FLAGGED: Hammer is not a tracked resource; mapped per user decision — Slingshot R pricing for Level ranks 6-10 rides on it).',
    'Level Challenge values are identical to the old generic ladder after mapping; only LBSize changed (10 -> 20 = live leagueGroupSize).',
    'Flash Race: league of 7 — ranks 8-10 zeroed; SPT 50/50/50/50/45/40/40 matches the "With SPT" calendar annotation.',
    '⚠ Race_v2 NOT regenerated (user decision): until Race_v2 is re-duplicated from this base (keep its EventDuration cells 2/2/2/2/1), R prices the OLD generic Flash ladder as a redesign — in particular R_SPT(Flash)=0 zeroes simulated Flash Race SPT.',
]
n0 = 3 + len(BLOCKS) * PITCH + 1                    # row 94
for i, txt in enumerate(notes):
    c = ws.cell(n0 + i, 1, txt)
    c.font = Font(name='Arial', size=9, bold=(i == 0), color='FF000000' if i == 0 else 'FF808080')

wb.save(os.path.join(DISPLAY, 'Race_v1.xlsx'))
print('written Race_v1.xlsx — 5 blocks, panel rows', [3 + b * PITCH for b in range(5)],
      ', ladder header rows', [9 + b * PITCH for b in range(5)])
print('changed vs live Race sheet: Flash Race ladder + LBSize/numberOfPositions=7; Level LBSize=20; all else identical')
