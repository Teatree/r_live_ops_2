# Builds PackConfig_v2.xlsx — replacement for the workbook's hidden 'PackConfig' sheet
# (card-collection pack/chest/reward configuration read by engine/CardOpenings.gs).
#
# 2026-08-03 (D19) — what changed vs the live PackConfig sheet:
#   1. PACK RARITY PROBABILITIES (the per-pack 1★..Gold probability grid) is GONE. It clashed with
#      the snap pool: CardOpenings' draw weighted `rarityWeight x poolCount`, so rarity was applied
#      twice. Rarity probability is now a PROPERTY OF THE POOL, identical for every pack type, and
#      it drifts as the pool is drawn down (draws are without replacement).
#   2. SNAP POOL stays and becomes the authoritative pool (CardPoolConfig is a test sheet and is no
#      longer read by the engine). It gains an 'Initial Probability' column — a FORMULA off the Qty
#      counts, never a typed number, so editing a quantity re-prices the pool automatically.
#   3. PACK DEFINITIONS keeps only Pack Type + Cards/Open. That plus PACK PITY CONFIG is now the
#      ONLY thing distinguishing a 6-star pack from a 1-star pack (user decision D19/9).
#   4. PACK PITY CONFIG is unchanged in content but is now actually READ (it never was).
#   5. NEW 'CHEST PURCHASING' panel — migrated from the deleted PlayerBehavior sheet and, likewise,
#      now actually read (CardOpenings used a hardcoded 0.85-of-season trigger and bought greedily).
#   6. 'Season Duration (days)' retired: the card sim runs the engine's 33-day calendar window.
#
# Every block is located by its LABEL at run time (findBlockRow_ in CardOpenings.gs), not by a
# fixed row number, so this layout can be re-ordered or grown without touching the engine.
#
# Style: Ph config-sheet rules — no merged cells, everything starts at column A, 0 (not blank) for
# empty numeric cells (punch-card: every in-game currency column present even when unused).
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter as CL

DISPLAY = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'display')

# The 21-column reward block shared by every config sheet in the workbook (Coins .. 6-star Dly).
REWARD_COLS = ['Coins', 'SPT', 'SPT x2', 'Red', 'Chuck', 'Bomb', 'Slingshot', 'Shuffle', 'Comet',
               'Unlimited Lives', 'Unlimited Red', 'Unlimited Chuck', 'Unlimited Bomb',
               'COOP Token', 'Avatar', '1-star Dly', '2-star Dly', '3-star Dly', '4-star Dly',
               '5-star Dly', '6-star Dly']
NCOL = 1 + len(REWARD_COLS)                      # 22 (A..V)

RARITIES = ['1★', '2★', '3★', '4★', '5★', 'Gold']
STARS_ON_DUPE = [1, 2, 3, 4, 5, 6]
SNAP_POOL = [281, 215, 143, 112, 66, 0]          # copies of each rarity in the season pool

PACKS = ['1-star Pack', '2-star Pack', '3-star Pack', '4-star Pack', '5-star Pack', '6-star Pack']
CARDS_PER_OPEN = [2, 3, 4, 5, 6, 7]
PITY_PROBS = ['[0]', '[0]', '[0]', '[0]', '[0, 0.33, 0.66, 1.0]', '[0, 0.8, 0.8, 1.0]']
PITY_FORCE = [False, False, False, False, False, True]

CHESTS = [('Bronze', 250, '1-star Pack'), ('Silver', 500, '4-star Pack'),
          ('Gold', 1000, '5-star Pack')]

# Set / Album rewards: carried over unchanged from the live sheet (authored inputs).
SET_REWARDS = {
    'Set 1': {'Coins': 300, 'Red': 2, 'Unlimited Lives': 30},
    'Set 2': {'Coins': 200, 'Chuck': 2},
    'Set 3': {'Coins': 100, 'Bomb': 2},
    'Set 4': {'Red': 1},
    'Set 5': {'Chuck': 1},
    'Set 6': {'Bomb': 1},
    'Set 7': {'Slingshot': 1},
    'Set 8': {'Shuffle': 1},
}
ALBUM_REWARDS = {
    'Album 1': {'Coins': 500, 'Red': 1},
    'Album 2': {'SPT': 20, 'Chuck': 1},
    'Album 3': {'Bomb': 1},
}

F_BAR, F_LBL, F_VAL, F_HDR, F_CALC = 'FF000000', 'FFFFD966', 'FFFFF2CC', 'FFF7CB4D', 'FFCFE2F3'
fill = lambda rgb: PatternFill('solid', fgColor=rgb)
thin = Side(style='thin')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
A = lambda **kw: Alignment(**kw)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'PackConfig'
ws.sheet_view.showGridLines = False

r = 1


def bar(text, width=NCOL):
    """Black section bar. The text in column A is the block's LOOKUP LABEL — CardOpenings.gs
    finds every block by scanning column A for it, so these strings are load-bearing."""
    global r
    for c in range(1, width + 1):
        ws.cell(r, c).fill = fill(F_BAR)
    ws.cell(r, 1, text).font = Font(name='Arial', size=11, bold=True, color='FFFFFFFF')
    r += 1


def header(cells):
    global r
    for c, h in enumerate(cells, 1):
        cell = ws.cell(r, c, h)
        cell.font = Font(name='Arial', size=11, bold=True)
        cell.fill = fill(F_HDR)
        cell.border = BORDER
        cell.alignment = A(horizontal='center', wrap_text=True)
    r += 1


def row(cells, calc_cols=()):
    """One data row. calc_cols (1-based) are formula/derived cells — blue 'data' fill, never typed."""
    global r
    for c, v in enumerate(cells, 1):
        cell = ws.cell(r, c, v)
        cell.font = Font(name='Arial', size=11)
        cell.border = BORDER
        cell.alignment = A(horizontal='left' if c == 1 else 'center')
        cell.fill = fill(F_CALC if c in calc_cols else F_VAL)
        if c in calc_cols:
            cell.number_format = '0.0000'
    r += 1


def note(text, bold=False):
    global r
    c = ws.cell(r, 1, text)
    c.font = Font(name='Arial', size=9, bold=bold, color='FF000000' if bold else 'FF808080')
    r += 1


def gap(n=1):
    global r
    r += n


ws['A1'] = 'Card Collection — Pack Configuration'
ws['A1'].font = Font(name='Arial', size=14, bold=True)
r = 3

# ---- 1. SEASON BASICS -------------------------------------------------------------------------
bar('SEASON BASICS', 3)
header(['Parameter', 'Value', 'Notes'])
row(['Total Cards per Season', 72, 'informational — AlbumConfig catalog is authoritative'])
row(['Number of Sets', 8, 'informational — derived from the AlbumConfig Set # column'])
row(['Cards per Set', 9, 'read by the sim (album grid width)'])
row(['Album Count (before loop)', 3, 'read by the sim; album N>count reuses the last reward row'])
note('Season Duration RETIRED 2026-08-03 (D19): the card sim runs the engine\'s 33-day calendar '
     'window on cal_new, so the season length is whatever the calendar says.')
gap()

# ---- 2. RARITY DEFINITIONS --------------------------------------------------------------------
bar('RARITY DEFINITIONS', 3)
header(['Rarity Tier', 'Stars on Duplicate', 'Notes'])
for rarity, stars in zip(RARITIES, STARS_ON_DUPE):
    row([rarity, stars, ''])
gap()

# ---- 3. SNAP POOL (authoritative pool + calculated initial probabilities) ----------------------
bar('SNAP POOL', 4)
pool_r0 = r + 1                                  # first data row (after the header)
pool_r1 = pool_r0 + len(RARITIES) - 1
header(['Rarity', 'Qty Count', 'Initial Probability', 'Notes'])
for i, (rarity, qty) in enumerate(zip(RARITIES, SNAP_POOL)):
    rr = pool_r0 + i
    row([rarity, qty,
         f'=IFERROR(B{rr}/SUM($B${pool_r0}:$B${pool_r1}),0)',
         'share of the pool at season start'], calc_cols=(3,))
tot_r = r
row(['TOTAL', f'=SUM(B{pool_r0}:B{pool_r1})', f'=SUM(C{pool_r0}:C{pool_r1})', ''], calc_cols=(2, 3))
ws.cell(tot_r, 1).font = Font(name='Arial', size=11, bold=True)
ws.cell(tot_r, 2).number_format = '0'
note('THE pool. Every pack — 1-star through 6-star — draws from this one distribution; the '
     'Initial Probability column IS the per-draw rarity chance at season start.', bold=True)
note('Calculated (blue) cells are formulas off Qty Count — never type into them. Edit a quantity '
     'and every probability re-prices.')
note('Draws are WITHOUT REPLACEMENT: each card drawn removes a copy, so the live probabilities '
     'drift away from these initial values as the season runs. The pool is rebuilt fresh only '
     'when an album is completed (D19/11).')
note('Copies are spread evenly across the AlbumConfig cards of each rarity (remainder to the '
     'lowest card indices), so a rarity with more distinct cards has fewer copies of each.')
gap()

# ---- 4. PACK DEFINITIONS ----------------------------------------------------------------------
bar('PACK DEFINITIONS', 3)
header(['Pack Type', 'Cards/Open', 'Notes'])
for p, n in zip(PACKS, CARDS_PER_OPEN):
    row([p, n, ''])
note('Cards/Open and the pity table below are the ONLY things that differentiate pack tiers '
     '(D19/9) — the rarity odds are identical for all of them and come from the SNAP POOL.',
     bold=True)
note('The old PACK RARITY PROBABILITIES grid was removed 2026-08-03: it multiplied the pool '
     'counts, applying rarity twice.')
gap()

# ---- 5. PACK PITY CONFIG ----------------------------------------------------------------------
bar('PACK PITY CONFIG', 4)
header(['Pack Type', 'PityProbabilities', 'PityForceHighestRarity', 'Notes'])
for p, pp, pf in zip(PACKS, PITY_PROBS, PITY_FORCE):
    row([p, pp, pf, ''])
note('SEMANTICS (implemented 2026-08-03; this table existed but was never read):', bold=True)
note('PityProbabilities is indexed by the number of CONSECUTIVE MISSES of the target rarity — NOT '
     'by card slot. p[0] applies to a pull with no misses behind it, p[1] after one miss, p[2] '
     'after two, and so on; entries past the end reuse the LAST value.')
note('So [0, 0.8, 0.8, 1.0] reads: no help at first; miss once and the next pull has an 80% chance '
     'of the target rarity; miss again, another 80%; miss a third time and the next pull is '
     'GUARANTEED.')
note('The counter RESETS to 0 the moment a pull lands on the target rarity — whether pity forced '
     'it or the player got there naturally — and starts at 0 on every pack open. It does NOT '
     'carry between packs.')
note('Target rarity = the highest rarity that STILL HAS COPIES in the snap pool when '
     'PityForceHighestRarity is TRUE. That is also the empty-tier fallback: Gold ships at Qty 0, '
     'so a 6-star pack\'s pity resolves to 5★ until Gold has stock, instead of chasing a card '
     'that cannot be drawn. When FALSE, the target is any rarity above the pool\'s most-stocked '
     'one.')
note('This is SEPARATE from the dry-streak pity in CardOpenings.gs, which chases a NEW card rather '
     'than a rare one (after 3 consecutive packs with no new card, the last card is forced to be '
     'an unowned type).')
gap()

# ---- 6. STAR CHEST COSTS & REWARDS ------------------------------------------------------------
bar('STAR CHEST COSTS & REWARDS', 4)
header(['Chest Tier', 'Cost (Stars)', 'Reward Pack Type', 'Notes'])
for tier, cost, pack in CHESTS:
    row([tier, cost, pack, ''])
gap()

# ---- 7. CHEST PURCHASING (migrated from PlayerBehavior) ---------------------------------------
bar('CHEST PURCHASING', 3)
header(['Parameter', 'Value', 'Description'])
row(['Min Stars to Consider Buying', 250, 'no purchase below this star balance'])
row(['Urgency Start Day', 14, 'before this day the buy probability is 0%'])
row(['End-of-Season Buy Probability', 0.95, 'buy probability on the final day; ramps linearly '
                                            'from 0% at Urgency Start Day'])
note('Migrated from the deleted PlayerBehavior sheet 2026-08-03 and now actually READ: the sim '
     'used to ignore all three and buy greedily after 85% of the season.', bold=True)
gap()

# ---- 8. SET REWARDS ---------------------------------------------------------------------------
bar('SET REWARDS')
header(['Set ID'] + REWARD_COLS)
for sid, rew in SET_REWARDS.items():
    row([sid] + [rew.get(h, 0) for h in REWARD_COLS])
gap()

# ---- 9. ALBUM REWARDS -------------------------------------------------------------------------
bar('ALBUM REWARDS')
header(['Album ID'] + REWARD_COLS)
for aid, rew in ALBUM_REWARDS.items():
    row([aid] + [rew.get(h, 0) for h in REWARD_COLS])
note('An album index beyond the last row defined here reuses that last row (albums loop).')
note('Set / Album reward payouts land in the SimOutput pack-log Note column. They are NOT fed '
     'back into EcoGainsSim — the card sim is a downstream consumer of the pack flow, not a '
     'source in the 25-category universe.')
gap()

# ---- widths -----------------------------------------------------------------------------------
ws.column_dimensions['A'].width = 30.0
ws.column_dimensions['B'].width = 20.0
ws.column_dimensions['C'].width = 24.0
ws.column_dimensions['D'].width = 52.0
for c in range(5, NCOL + 1):
    ws.column_dimensions[CL(c)].width = 13.0

out = os.path.join(DISPLAY, 'PackConfig_v2.xlsx')
wb.save(out)
print('written PackConfig_v2.xlsx')
print('  SNAP POOL data rows', pool_r0, '-', pool_r1, '(Initial Probability = formula off Qty Count)')
print('  blocks are located by their column-A label at run time — row numbers are not load-bearing')
print('  REMOVED: PACK RARITY PROBABILITIES, Season Duration   ADDED: CHEST PURCHASING')
