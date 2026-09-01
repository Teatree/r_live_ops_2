# Builds Col_Cards_Cloud_v1.xlsx + Col_Cards_Totals_v1.xlsx — the two output sheets of the
# STOCHASTIC card simulation (SimulateCardCloud in engine/CardOpenings.gs, D24 2026-09-01).
#
# ONE script, TWO files. That is a deliberate deviation from one-builder-per-xlsx: the two sheets
# share every geometry constant below, and splitting them would put those constants in two places —
# which is the exact drift this project keeps paying for.
#
# WHAT IS SHARED WITH THE ENGINE, and must be kept in step with engine/CardOpenings.gs:
#   * the BAR LABELS (BAR_MEANS / BAR_BANDS / TB_*) — the engine finds every block by scanning
#     column A for these, so row numbers here are NOT load-bearing and a block can be moved freely;
#   * BAND_STRIDE, SRC_ROWS, and the input/stamp cells.
# Everything else on these sheets — group labels, headers, permutation names, resource names, source
# names — is written by the ENGINE at run time. This file only reserves and styles the areas.
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter as CL

DISPLAY = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'display')

DAYS = 33
SEGMENTS = ['0-9', '10-19', '20-39', '40-99', '100+']      # == CLOUD_SEGMENTS ('A. 0' excluded)
PAYERS = ['NONPAYER', 'PAYER']                              # == CLOUD_PAYERS
PERMS = [f'{s} {p}' for s in SEGMENTS for p in PAYERS]      # 10, in engine column order
NPERM = len(PERMS)

METRICS = ['Packs Opened', 'Cards Drawn', 'Unique Cards',
           'Sets Completed', 'Album %', 'Star Balance']     # == CLOUD_METRICS labels
STATS = ['p10', 'p25', 'p50', 'p75', 'p90', 'MEAN']         # == CLOUD_STATS

BAR_MEANS = 'MEANS - ALL PERMUTATIONS'                      # == CLOUD_BAR_MEANS
BAR_BANDS = 'PER-PERMUTATION BANDS'                         # == CLOUD_BAR_BANDS
BAND_STRIDE = 37                                            # == CLOUD_BAND_STRIDE
SRC_ROWS = 30                                               # == CLOUD_SRC_ROWS

TB = [                                                       # == TB in the engine, in sheet order
    ('TOTALS (mean per player)', 10),
    ('TOTALS (p10-p90 across players)', 10),
    ('CADENCE (per calendar day, all 33)', 4),
    ('ECONOMY IMPACT - TOTAL (mean per player)', 21),
    ('ECONOMY IMPACT - FROM SET COMPLETIONS', 21),
    ('ECONOMY IMPACT - FROM ALBUM COMPLETIONS', 21),
    ('UNLIMITED BOOSTERS IN MINUTES', 4),
    ('PACKS PER SOURCE (mean)', SRC_ROWS),
    ('PACKS PER SOURCE (p10-p90)', SRC_ROWS),
    ('CARDS PER SOURCE (mean)', SRC_ROWS),
    ('CARDS PER SOURCE (p10-p90)', SRC_ROWS),
]
UL_BLOCK = 'UNLIMITED BOOSTERS IN MINUTES'                   # the one block with an input column

F_BAR, F_HDR, F_IN, F_OUT = 'FF000000', 'FFF7CB4D', 'FFFFF2CC', 'FFE2EFDA'
fill = lambda rgb: PatternFill('solid', fgColor=rgb)
thin = Side(style='thin')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
ARIAL = lambda **kw: Font(name='Arial', **kw)

# The single Gemini prompt (the user asked for one covering every chart).
CHART_PROMPT = (
    'Build one line chart per metric block on this sheet. Column A is the X axis (Day). '
    'In each PER-PERMUTATION BANDS block, plot MEAN as a bold line, p25 and p75 as thin lines, '
    'and shade the area between p10 and p90. In the MEANS block at the top, put all 10 permutation '
    'columns of a single metric on one chart, each permutation a distinct colour, titled with the '
    'metric name. No trendlines, no data labels.'
)


def bar(ws, r, text, c1):
    for c in range(1, c1 + 1):
        ws.cell(r, c).fill = fill(F_BAR)
    ws.cell(r, 1, text).font = ARIAL(size=11, bold=True, color='FFFFFFFF')


def hdr_row(ws, r, ncols):
    """Style only — the engine writes the header TEXT."""
    for c in range(1, ncols + 1):
        cell = ws.cell(r, c)
        cell.font = ARIAL(size=10, bold=True)
        cell.fill = fill(F_HDR)
        cell.border = BORDER
        cell.alignment = Alignment(horizontal='center', wrap_text=True)


def data_block(ws, r0, nrows, ncols, first_col_left=True):
    for i in range(nrows):
        for c in range(1, ncols + 1):
            cell = ws.cell(r0 + i, c, '-')
            cell.font = ARIAL(size=10)
            cell.fill = fill(F_OUT)
            cell.border = BORDER
            cell.alignment = Alignment(
                horizontal='left' if (c == 1 and first_col_left) else 'center')


def note(ws, r, text, c=1, bold=False):
    cell = ws.cell(r, c, text)
    cell.font = ARIAL(size=9, bold=bold, color='FF000000' if bold else 'FF808080')


# =============================== Col_Cards_Cloud ===============================================
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Col_Cards_Cloud'
ws.sheet_view.showGridLines = False

MEANS_COLS = 1 + len(METRICS) * NPERM          # 61
BAND_COLS = 1 + len(METRICS) * len(STATS)      # 37

ws['A1'] = 'CARD COLLECTION — STOCHASTIC SIMULATION (per-day distribution)'
ws['A1'].font = ARIAL(size=14, bold=True)
note(ws, 2, '(run stamp — written by EcoGainsSim ▸ Simulate card cloud)')

MEANS_BAR = 4
bar(ws, MEANS_BAR, BAR_MEANS, MEANS_COLS)
note(ws, MEANS_BAR + 1, '')                                  # group-label row (engine writes)
for c in range(1, MEANS_COLS + 1):
    ws.cell(MEANS_BAR + 1, c).font = ARIAL(size=10, bold=True, color='FF555555')
hdr_row(ws, MEANS_BAR + 2, MEANS_COLS)
data_block(ws, MEANS_BAR + 3, DAYS, MEANS_COLS, first_col_left=False)

BANDS_BAR = MEANS_BAR + 3 + DAYS + 1                         # 41
bar(ws, BANDS_BAR, BAR_BANDS, BAND_COLS)
note(ws, BANDS_BAR + 1,
     f'One block per permutation, {BAND_STRIDE} rows apart. Only the bar above is found by label — '
     f'the blocks sit at a fixed stride below it, so keep BAND_STRIDE in step with the engine.')

for j, label in enumerate(PERMS):
    r0 = BANDS_BAR + 2 + j * BAND_STRIDE                     # == engine's rBands + 2 + j*STRIDE
    ws.cell(r0, 1, label).font = ARIAL(size=11, bold=True)   # engine overwrites with the same text
    for c in range(1, BAND_COLS + 1):
        ws.cell(r0 + 1, c).font = ARIAL(size=10, bold=True, color='FF555555')
    hdr_row(ws, r0 + 2, BAND_COLS)
    data_block(ws, r0 + 3, DAYS, BAND_COLS, first_col_left=False)

PROMPT_R = BANDS_BAR + 2 + NPERM * BAND_STRIDE + 1
bar(ws, PROMPT_R, 'CHART PROMPT (paste into Google Sheets Gemini)', 8)
ws.cell(PROMPT_R + 1, 1, CHART_PROMPT).font = ARIAL(size=10)
ws.cell(PROMPT_R + 1, 1).alignment = Alignment(wrap_text=True, vertical='top')
ws.row_dimensions[PROMPT_R + 1].height = 60

ws.column_dimensions['A'].width = 10.0
for c in range(2, max(MEANS_COLS, BAND_COLS) + 1):
    ws.column_dimensions[CL(c)].width = 11.0

out_cloud = os.path.join(DISPLAY, 'Col_Cards_Cloud_v1.xlsx')
wb.save(out_cloud)

# =============================== Col_Cards_Totals ==============================================
wb2 = openpyxl.Workbook()
ws2 = wb2.active
ws2.title = 'Col_Cards_Totals'
ws2.sheet_view.showGridLines = False

TOT_COLS = 1 + NPERM                                          # A + 10 permutations
UL_COLS = 2 + NPERM                                           # A + minutes-per-unit input + 10

ws2['A1'] = 'CARD COLLECTION — STOCHASTIC SIMULATION (totals per permutation)'
ws2['A1'].font = ARIAL(size=14, bold=True)
for addr, label in (('A2', 'Players per permutation:'), ('C2', 'Seed:')):
    ws2[addr] = label
    ws2[addr].font = ARIAL(size=11, bold=True)
for addr, default in (('B2', 50), ('D2', '')):                # == CLOUD_PLAYERS_CELL / SEED_CELL
    ws2[addr] = default
    ws2[addr].font = ARIAL(size=11)
    ws2[addr].fill = fill(F_IN)
    ws2[addr].border = BORDER
    ws2[addr].alignment = Alignment(horizontal='center')
note(ws2, 2, '(run stamp)', c=6)

r = 4
positions = {}
for label, nrows in TB:
    ncols = UL_COLS if label == UL_BLOCK else TOT_COLS
    bar(ws2, r, label, ncols)
    hdr_row(ws2, r + 1, ncols)
    data_block(ws2, r + 2, nrows, ncols)
    if label == UL_BLOCK:
        # column B is an INPUT the engine reads and never overwrites: blank means "no conversion",
        # and the minutes cells then read '-' instead of a number nobody chose.
        for i in range(nrows):
            cell = ws2.cell(r + 2 + i, 2, None)
            cell.fill = fill(F_IN)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal='center')
        note(ws2, r + 2 + nrows, 'Type minutes-per-unit in the yellow column; blank leaves the '
                                 'minutes columns as "-". Raw counts are in ECONOMY IMPACT above.')
    positions[label] = r
    r += 2 + nrows + 1

bar(ws2, r, 'NOTES', 8)
note(ws2, r + 1, 'Ranges are p10-p90 ACROSS PLAYERS, not min-max: with 50 players the true '
                 'extremes are single outliers that move every run.')
note(ws2, r + 2, 'Cadence is per CALENDAR day (all 33), including days the player did not play — '
                 'so the permutations stay comparable.')
note(ws2, r + 3, 'ECONOMY IMPACT is what the collection feature PAYS OUT (SET + ALBUM REWARDS). '
                 'It is not the segment\'s whole economy.')
note(ws2, r + 4, 'A zero in a per-source row is a finding, not a gap — e.g. Kite Festival at its '
                 'assumed 0.35 opt-in, or a ladder with no pack authored on it.')
note(ws2, r + 5, '"A. 0" is absent by design: data_seg_beh has no row for it, so nothing can price '
                 'its reach.')

ws2.column_dimensions['A'].width = 34.0
for c in range(2, UL_COLS + 1):
    ws2.column_dimensions[CL(c)].width = 15.0

out_totals = os.path.join(DISPLAY, 'Col_Cards_Totals_v1.xlsx')
wb2.save(out_totals)

print('written Col_Cards_Cloud_v1.xlsx')
print(f'  {BAR_MEANS!r} bar at row {MEANS_BAR}, {DAYS} data rows, {MEANS_COLS} cols')
print(f'  {BAR_BANDS!r} bar at row {BANDS_BAR}, {NPERM} blocks x {BAND_STRIDE} rows, {BAND_COLS} cols')
print(f'  chart prompt at row {PROMPT_R}')
print('written Col_Cards_Totals_v1.xlsx')
for label, _ in TB:
    print(f'  {label!r} bar at row {positions[label]}')
print(f'  inputs B2 (players) / D2 (seed)')
