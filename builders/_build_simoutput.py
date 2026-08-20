# Builds SimOutput_v2.xlsx — replacement for the workbook's hidden 'SimOutput' sheet
# (the card-collection simulator's input panel + output surface, written by CardOpenings.gs).
#
# 2026-08-03 (D19) — what changed vs the live SimOutput sheet:
#   1. 'Selected Group' (Casual/Mid-Core/Hardcore x Free/Payer, from the deleted PlayerBehavior
#      sheet) becomes TWO real dropdowns: Segment (B2) and Payer (D2), the exact keys the rest of
#      the engine uses (data_seg_beh / data_gains).
#   2. The running-totals block runs 33 days, not 28 — the card sim now uses the engine's calendar
#      window on cal_new instead of a 29-entry attendance array.
#   3. Two new tally rows: 'Expected Packs (fractional)' — the unrounded simulated pack flow, so
#      the discrete-grant rounding is visible — and 'Segment / Payer' for provenance.
#   4. A 'Packs Opened' column joins the per-day running totals.
#   5. The album/set grid area moves to J55:Z260 (3 albums x 8 sets, stacked) and is scanned by
#      LABEL, so this layout can change freely as long as the 'Album #N' / 'Set #N' labels remain.
#
# Row geometry IS load-bearing here (CardOpenings.gs writes into fixed anchors):
#   B2 / D2 / G2 inputs · running totals rows 6-38 · tally values B42:B53 · pack log from row 57
# Keep those in sync with OUT_START_ROW / TOTALS_FIRST_ROW / TALLY_FIRST_ROW / GRID_SCAN_RANGE.
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter as CL

DISPLAY = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'display')

DAYS = 33
SEGMENTS = ['0-9', '10-19', '20-39', '40-99', '100+', 'A. 0']
PAYERS = ['NONPAYER', 'PAYER']

TOTALS_HDR_ROW = 5
TOTALS_R0 = 6                                    # == CardOpenings TOTALS_FIRST_ROW
TOTALS_R1 = TOTALS_R0 + DAYS - 1                 # 38
TALLY_BAR = TOTALS_R1 + 2                        # 40
TALLY_HDR = TALLY_BAR + 1                        # 41
TALLY_R0 = TALLY_HDR + 1                         # 42 == CardOpenings TALLY_FIRST_ROW
LOG_BAR = 55
LOG_HDR = 56
LOG_R0 = 57                                      # == CardOpenings OUT_START_ROW

TALLY_ROWS = [
    'Total Packs Opened', 'Total Cards Drawn', 'Unique Cards', 'Duplicate Cards',
    'Stars Earned', 'Stars Spent on Chests', 'Final Star Balance', 'Sets Completed',
    'Album Tier Reached', 'Day Album Completed', 'Expected Packs (fractional)', 'Segment / Payer',
]

TOTALS_HDRS = ['Day', 'Star Balance', 'Unique Cards', '% Complete', 'Sets Done', 'Album Tier',
               'Packs Opened']
LOG_HDRS = ['Day', 'Pack', 'Source', 'Earned From', 'Album', 'Cards Drawn', 'New', 'Dupes',
            'Stars Balance', 'Note']

# Album/set grids sit immediately RIGHT of the pack log, derived from the log width rather than
# hardcoded: the log gained an 'Earned From' column on 2026-08-18 and a fixed column J would have
# put 'Album #1' straight on top of 'Note'. One blank spacer column between the two blocks.
GRID_C0 = len(LOG_HDRS) + 2

N_ALBUMS, N_SETS = 3, 8

F_BAR, F_HDR, F_IN, F_OUT = 'FF000000', 'FFF7CB4D', 'FFFFF2CC', 'FFE2EFDA'
fill = lambda rgb: PatternFill('solid', fgColor=rgb)
thin = Side(style='thin')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
ARIAL = lambda **kw: Font(name='Arial', **kw)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'SimOutput'
ws.sheet_view.showGridLines = False


def bar(r, text, c0=1, c1=9):
    for c in range(c0, c1 + 1):
        ws.cell(r, c).fill = fill(F_BAR)
    ws.cell(r, c0, text).font = ARIAL(size=11, bold=True, color='FFFFFFFF')


def header(r, hdrs, c0=1):
    for i, h in enumerate(hdrs):
        cell = ws.cell(r, c0 + i, h)
        cell.font = ARIAL(size=11, bold=True)
        cell.fill = fill(F_HDR)
        cell.border = BORDER
        cell.alignment = Alignment(horizontal='center', wrap_text=True)


def note(r, text, bold=False):
    c = ws.cell(r, 1, text)
    c.font = ARIAL(size=9, bold=bold, color='FF000000' if bold else 'FF808080')


# ---- title + inputs ---------------------------------------------------------------------------
ws['A1'] = 'CARD COLLECTION — INDIVIDUAL SIMULATION LOG'
ws['A1'].font = ARIAL(size=14, bold=True)

for addr, label in (('A2', 'Segment:'), ('C2', 'Payer:'), ('F2', 'Sim Seed:')):
    ws[addr] = label
    ws[addr].font = ARIAL(size=11, bold=True)
for addr, default in (('B2', '10-19'), ('D2', 'NONPAYER'), ('G2', 1)):
    ws[addr] = default
    ws[addr].font = ARIAL(size=11)
    ws[addr].fill = fill(F_IN)
    ws[addr].border = BORDER
    ws[addr].alignment = Alignment(horizontal='center')

dv_seg = DataValidation(type='list', formula1='"' + ','.join(SEGMENTS) + '"', allow_blank=False)
dv_pay = DataValidation(type='list', formula1='"' + ','.join(PAYERS) + '"', allow_blank=False)
ws.add_data_validation(dv_seg)
ws.add_data_validation(dv_pay)
dv_seg.add(ws['B2'])
dv_pay.add(ws['D2'])

ws['I2'] = ('Segment x payer are the real engine keys (data_seg_beh). Blank seed -> generated and '
            'written back, so any run can be reproduced.')
ws['I2'].font = ARIAL(size=9, color='FF808080')

# ---- running totals ---------------------------------------------------------------------------
bar(TOTALS_HDR_ROW - 1, 'RUNNING TOTALS (per day, 33-day calendar window on cal_new)', 1, 7)
header(TOTALS_HDR_ROW, TOTALS_HDRS)
for i in range(DAYS):
    r = TOTALS_R0 + i
    for c in range(1, len(TOTALS_HDRS) + 1):
        cell = ws.cell(r, c, '-')
        cell.font = ARIAL(size=11)
        cell.fill = fill(F_OUT)
        cell.border = BORDER
        cell.alignment = Alignment(horizontal='center')
    ws.cell(r, 4).number_format = '0.0%'

# ---- tally ------------------------------------------------------------------------------------
bar(TALLY_BAR, 'SIMULATION TALLY', 1, 2)
header(TALLY_HDR, ['Metric', 'Value'])
for i, label in enumerate(TALLY_ROWS):
    r = TALLY_R0 + i
    lc = ws.cell(r, 1, label)
    lc.font = ARIAL(size=11)
    lc.border = BORDER
    vc = ws.cell(r, 2, '-')
    vc.font = ARIAL(size=11)
    vc.fill = fill(F_OUT)
    vc.border = BORDER
    vc.alignment = Alignment(horizontal='center')
ws.cell(TALLY_R0 + TALLY_ROWS.index('Expected Packs (fractional)'), 3,
        'unrounded simulated pack flow — the gap vs Total Packs Opened is the discrete-grant '
        'rounding (seeded Bernoulli on each trailing fraction)').font = ARIAL(size=9, color='FF808080')

# ---- pack log ---------------------------------------------------------------------------------
bar(LOG_BAR, 'DAY-BY-DAY PACK LOG (OUTPUT)', 1, len(LOG_HDRS))
header(LOG_HDR, LOG_HDRS)
for r in range(LOG_R0, LOG_R0 + 220):
    for c in range(1, len(LOG_HDRS) + 1):
        ws.cell(r, c).font = ARIAL(size=10)

# ---- album / set grids (scanned by label — 'Album #N' / 'Set #N') -------------------------------
bar(LOG_BAR, 'PACKS RESULT', GRID_C0, GRID_C0 + 3)
gr = LOG_BAR + 1
grid_first, grid_last = gr, gr
for a in range(1, N_ALBUMS + 1):
    ac = ws.cell(gr, GRID_C0, f'Album #{a}')
    ac.font = ARIAL(size=11, bold=True)
    gr += 1
    for s in range(1, N_SETS + 1):
        sc = ws.cell(gr, GRID_C0, f'Set #{s}')
        sc.font = ARIAL(size=10, bold=True)
        gr += 1
        for i in range(3):                       # the 3x3 grid CardOpenings paints
            for j in range(3):
                cell = ws.cell(gr + i, GRID_C0 + j)
                cell.font = ARIAL(size=10)
                cell.border = BORDER
                cell.alignment = Alignment(horizontal='center')
        gr += 3
    gr += 1
grid_last = gr

# ---- notes ------------------------------------------------------------------------------------
n = gr + 1
note(n, 'NOTES (2026-08-03, D19 — see engine/CardOpenings.gs):', bold=True)
note(n + 1, 'Packs come from the SAME engine path as EcoGainsSim_Daily: per-day expected packs '
            'per tier per source, priced off the _v2 reward ladders against cal_new. The old '
            'EcoPackGains rate table and its hardcoded schedule strings are gone.')
note(n + 2, 'Attendance is the deterministic expected p_day from data_seg_beh (weekday/weekend '
            'active rates). Randomness enters only through the card DRAWS and the trailing-'
            'fraction pack grant, both driven by the seed in G2.')
note(n + 3, 'Rarity odds are a property of the SNAP POOL and are identical for every pack tier; '
            'they drift as the pool is drawn down. Pack tier differs only by Cards/Open and the '
            'PACK PITY CONFIG table.')
note(n + 4, 'Set / Album rewards are reported in the Note column only — the card sim consumes the '
            'pack flow, it is not a source in the 25-category gains universe.')

# ---- widths -----------------------------------------------------------------------------------
for col, w in {'A': 22.0, 'B': 13.0, 'C': 30.0, 'D': 11.0, 'E': 46.0, 'F': 30.0, 'G': 30.0,
               'H': 12.0, 'I': 60.0}.items():
    ws.column_dimensions[col].width = w
for c in range(GRID_C0, GRID_C0 + 4):
    ws.column_dimensions[CL(c)].width = 9.0

out = os.path.join(DISPLAY, 'SimOutput_v2.xlsx')
wb.save(out)
print('written SimOutput_v2.xlsx')
print(f'  inputs B2 (segment) / D2 (payer) / G2 (seed)')
print(f'  running totals rows {TOTALS_R0}-{TOTALS_R1} ({DAYS} days) · tally values B{TALLY_R0}-'
      f'B{TALLY_R0 + len(TALLY_ROWS) - 1} · pack log from row {LOG_R0}')
print(f'  album/set grids rows {grid_first}-{grid_last} in column {CL(GRID_C0)} '
      f'(GRID_SCAN_RANGE must cover them)')
