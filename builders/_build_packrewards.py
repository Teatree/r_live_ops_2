# Emits display/PackRewards_v1.xlsx — the pack-reward config sheets, ready to import into the
# Google workbook (D21).
#
# Strategy: start from the workbook of record and DELETE everything except the 13 config sheets
# that carry pack ladders, then write the pack columns. Starting from the live workbook (rather
# than building sheets from scratch) preserves every formula, fill, number format and column width
# exactly as they are today — these config sheets carry 400+ formulas between them (RM_1st alone
# has 97), and re-authoring them would silently drop those. Verified before writing: none of the
# target sheets' formulas reference an external sheet, and no pack cell currently holds a formula,
# so trimming the workbook cannot break a reference and the pack writes only touch literals.
#
# The ladder layout lives in _pack_spec.json (shared with harness/_solve_packs.js) — this script
# never invents a row index. Every authored cell is 1: the "one pack per star per ladder row" rule.
#
# Run:  python builders/_build_packrewards.py
import os, sys, glob, json, math
import openpyxl

ROOT = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..')
SPEC_PATH = os.path.join(ROOT, 'builders', '_pack_spec.json')
OUT = os.path.join(ROOT, 'display', 'PackRewards_v1.xlsx')

N_PACK_TIERS = 6

# The user's source x tier table (2026-08-04), as a hard whitelist: a tier may ONLY appear on
# these blocks. Team Race is absent (it has no config sheet, so the engine prices it at 0 anyway);
# Level Race is absent by the table even though it shares Race_v2 with the bird challenges, so its
# rank block must stay pack-free. 6-star/Gold has no source at all.
WHITELIST = {
    1: {'Season Pass Free', 'Photoshoot', 'Jigsaw', 'Hatchling Hideaway', "Bomb's Ballet",
        'Rainbow Maker (1st)', 'Rainbow Maker (2nd)'},
    2: {'Season Pass Free', 'Photoshoot', 'Jigsaw', 'Hatchling Hideaway', "Bomb's Ballet",
        'Rainbow Maker (1st)', 'Rainbow Maker (2nd)'},
    3: {'Red Challenge', 'Chuck Challenge', 'Bomb Challenge', 'Flash Race', 'Flock Flurry',
        'Kite Festival (rank)', 'Kite Festival (score)', 'Target Day',
        'Rainbow Maker (1st)', 'Rainbow Maker (2nd)',
        'Night Sky 0-9', 'Night Sky 10-19', 'Night Sky 20-39', 'Night Sky 40-99', 'Night Sky 100+'},
    4: {'Red Challenge', 'Chuck Challenge', 'Bomb Challenge', 'Flash Race', 'Flock Flurry',
        'Kite Festival (rank)', 'Kite Festival (score)', 'Target Day',
        'Rainbow Maker (1st)', 'Rainbow Maker (2nd)',
        'Night Sky 0-9', 'Night Sky 10-19', 'Night Sky 20-39', 'Night Sky 40-99', 'Night Sky 100+',
        'Team Event (leaderboard)', 'Team Event (contribution)'},
    5: {'Season Pass Paid', 'Team Event (leaderboard)', 'Team Event (contribution)',
        'Rainbow Maker (1st)', 'Rainbow Maker (2nd)'},
    6: set(),
}

# Blocks the engine reads that must stay pack-free (not in the user's table).
PACK_FREE_BLOCKS = [
    {'id': 'Level Race', 'sheet': 'Race_v2', 'r0': 63, 'r1': 72, 'colBase': 16},
]


def newest_workbook():
    pat = os.path.join(ROOT, 'workbooks', 'NEW_LIVEOPS_CALENDAR_ECO (*).xlsx')
    files = glob.glob(pat)
    if not files:
        sys.exit('no NEW_LIVEOPS_CALENDAR_ECO workbook found in workbooks/')
    return max(files, key=lambda p: int(p.split('(')[1].split(')')[0]))


def tier_rows(block, cfg, scale):
    """Rows (0-based, sheet coords) that carry one pack of this tier. Mirrors tierRows() in
    harness/_solve_packs.js — keep the two in step."""
    if cfg.get('rows'):          # explicit authored ramp; scale does not apply (see spec _doc)
        return [r for r in cfg['rows'] if block['r0'] <= r <= block['r1']]
    n = block['r1'] - block['r0'] + 1
    frac = cfg.get('frac', 1.0)
    every = cfg.get('every', 1)
    need = frac * (1.0 if scale is None else scale)
    if need <= 1:
        frac = need
    else:
        frac, every = 1.0, max(1, round(every / need))
    k = max(0, min(n, math.ceil(frac * n)))
    # Default walks from the GOOD end (best rank / deepest milestone); from='low' walks from the
    # ENTRY end, so 1-star reaches players who never get deep into a ladder.
    low_first = cfg.get('from') == 'low'
    rows = []
    i = 0
    while i < k:
        if low_first:
            rows.append(block['r1'] - i if block['dir'] == 'rank' else block['r0'] + i)
        else:
            rows.append(block['r0'] + i if block['dir'] == 'rank' else block['r1'] - i)
        i += every
    return rows


def main():
    spec = json.load(open(SPEC_PATH, encoding='utf8'))
    scale = spec.get('scale', 1.0)
    src = newest_workbook()
    print('source workbook : %s' % os.path.basename(src))
    print('spec scale      : %s' % scale)

    targets = []
    for b in spec['blocks']:
        if b['sheet'] not in targets:
            targets.append(b['sheet'])

    wb = openpyxl.load_workbook(src, data_only=False)
    missing = [t for t in targets if t not in wb.sheetnames]
    if missing:
        sys.exit('workbook is missing config sheets: %s' % ', '.join(missing))

    for name in list(wb.sheetnames):
        if name not in targets:
            del wb[name]

    # 1) zero every pack cell inside every spec'd block, so a re-run is idempotent and unused
    #    tiers read 0 rather than blank (config-sheet punch-card rule: 0, not blank).
    for b in spec['blocks']:
        ws = wb[b['sheet']]
        for r in range(b['r0'], b['r1'] + 1):
            for t in range(N_PACK_TIERS):
                ws.cell(row=r + 1, column=b['colBase'] + t + 1).value = 0

    # 2) author the ladders
    written = {}
    for b in spec['blocks']:
        ws = wb[b['sheet']]
        for star, cfg in b['tiers'].items():
            star = int(star)
            if star == 6:
                sys.exit('spec authors a 6-star pack; the table has no 6-star source')
            if b['id'] not in WHITELIST[star]:
                sys.exit('WHITELIST violation: %d-star on "%s" is not in the source table'
                         % (star, b['id']))
            rows = tier_rows(b, cfg, scale)
            for r in rows:
                if not (b['r0'] <= r <= b['r1']):
                    sys.exit('spec row %d outside block %s %s' % (r, b['sheet'], b['id']))
                ws.cell(row=r + 1, column=b['colBase'] + star).value = 1
            written.setdefault(b['id'], []).append('%d*x%d' % (star, len(rows)))

    # 3) invariant sweep: nothing above 1 anywhere in a pack column, nothing in the 6-star column
    bad = []
    for b in spec['blocks']:
        ws = wb[b['sheet']]
        for r in range(b['r0'], b['r1'] + 1):
            for t in range(N_PACK_TIERS):
                v = ws.cell(row=r + 1, column=b['colBase'] + t + 1).value
                if v not in (0, 1):
                    bad.append('%s r%d %d-star = %r' % (b['sheet'], r + 1, t + 1, v))
                if t == 5 and v:
                    bad.append('%s r%d 6-star nonzero' % (b['sheet'], r + 1))
    # 4) blocks that must stay pack-free (in the sheet, but not in the user's table)
    for pf in PACK_FREE_BLOCKS:
        ws = wb[pf['sheet']]
        for r in range(pf['r0'], pf['r1'] + 1):
            for t in range(N_PACK_TIERS):
                v = ws.cell(row=r + 1, column=pf['colBase'] + t + 1).value
                if v:
                    bad.append('%s (%s) r%d %d-star = %r must be pack-free'
                               % (pf['id'], pf['sheet'], r + 1, t + 1, v))

    if bad:
        sys.exit('INVARIANT FAILED:\n  ' + '\n  '.join(bad[:20]))

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    wb.save(OUT)

    print('\nsheets written  : %s' % ', '.join(targets))
    for k in sorted(written):
        print('  %-28s %s' % (k, ' '.join(written[k])))
    total = sum(int(x.split('x')[1]) for v in written.values() for x in v)
    print('\ntotal authored cells: %d (all = 1)' % total)
    print('wrote %s' % os.path.relpath(OUT, ROOT))


if __name__ == '__main__':
    main()
