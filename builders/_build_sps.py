# Builds Sim_per_Segment_v3.xlsx — the 'Sim per Segment' rollup, one table per resource.
# GAINS block (per earner, from the engine) + NET block (per EARNER too since v3 — same
# resource_earners denominator as the gains block — spend held constant).
# Each payer block ends with an 'overall' row: gains = unique_players-weighted average of the 5
# segs; NET columns = resource_earners-weighted (written by the menu script).
#
# Prefilled with the offline gains from _sps_values.json and unique_players from the live workbook.
# The NET columns (current_spend/current_net/new_net/net_diff) need the per-earner columns on the
# 'data_econ' sheet (gain_per_earner / spend_per_earner / resource_earners — see
# sqls/data_econ_PROMPT.md v2) and are populated by the menu action (SimPerSegmentFill.gs) —
# left blank here until that data exists.
import glob
import json, os, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

HERE = os.path.dirname(os.path.abspath(__file__))
VALS = json.load(open(os.path.join(HERE, '_sps_values.json')))
RES = ['HC', 'Slingshot', 'Shuffle', 'Comet', 'Red', 'Chuck', 'Bomb',
       'UL Bomb', 'UL Chuck', 'UL Red', 'Unlimited Lives']
GROUPS = ['PAID', 'ADS', 'CORE', 'META']
SEGS = ['0-9', '10-19', '20-39', '40-99', '100+']
PAYERS = ['NONPAYER', 'PAYER']

# unique_players weights (for the prefilled 'overall' rows) from the workbook of record
# (highest-numbered NEW_LIVEOPS_CALENDAR_ECO (N).xlsx — same pick rule as harness/_dump_mockdata.py)
import re as _re
_books = glob.glob(os.path.join(HERE, '..', 'workbooks', 'NEW_LIVEOPS_CALENDAR_ECO*.xlsx'))
_books.sort(key=lambda n: int((_re.search(r'\((\d+)\)', n) or [0, 0])[1]))
SRC = _books[-1]
_sb = openpyxl.load_workbook(SRC, data_only=True)['data_seg_beh']
UP = {}
for r in range(2, _sb.max_row + 1):
    seg, payer, up = _sb.cell(r, 1).value, _sb.cell(r, 2).value, _sb.cell(r, 4).value
    if seg: UP[(str(seg), str(payer))] = float(up or 0)
def wavg(payer, vals_by_seg):
    num = sum(vals_by_seg[s] * UP.get((s, payer), 0) for s in SEGS)
    den = sum(UP.get((s, payer), 0) for s in SEGS)
    return num / den if den else 0

F_MARK, F_CUR, F_SIM, F_DELTA, F_NET = 'FF1F4E78', 'FF2E75B6', 'FF548235', 'FF666666', 'FF7030A0'
F_HDRSEG, F_HDRCUR, F_HDRSIM, F_HDRDL, F_HDRNET = 'FFD9D9D9', 'FFCFE2F3', 'FFE2EFDA', 'FFEFEFEF', 'FFE9DEF3'
F_PAYERBAND, F_OVERALL = 'FF305496', 'FFDDEBF7'
fill = lambda rgb: PatternFill('solid', fgColor=rgb)
thin = Side(style='thin')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
FMT_VAL, FMT_PCT = '#,##0.0', '+0%;-0%;0%'
FMT_NET = '#,##0.0;[Red]-#,##0.0'          # net can be negative (spend > gain)
CL = lambda i: openpyxl.utils.get_column_letter(i)

# NET block: cols U:X (21..24), gap at T(20)
NET_C0 = 21
NET_HDRS = ['cur spend', 'cur net', 'new net', 'net Δ']

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Sim per Segment'
ws.sheet_view.showGridLines = False
for col, w in {'A': 2.0, 'B': 11.0, 'H': 2.88, 'N': 2.88, 'T': 2.88}.items():
    ws.column_dimensions[col].width = w
for c in list(range(3, 8)) + list(range(9, 14)) + list(range(15, 20)) + list(range(NET_C0, NET_C0 + 4)):
    ws.column_dimensions[CL(c)].width = 8.5

ws['A1'] = 'RESOURCE GAINS + NET (33 day period)'
ws['A1'].font = Font(name='Arial', size=13, bold=True)

PITCH = 19                                  # 3 header rows + 2 blocks x (band+5 seg+overall=7) + 2 blank
def styled(r, c, val=None, *, font=None, rgb=None, fmt=None, border=False, center=True):
    cell = ws.cell(r, c, val)
    cell.font = font or Font(name='Arial', size=11)
    if rgb: cell.fill = fill(rgb)
    if fmt: cell.number_format = fmt
    if border: cell.border = BORDER
    if center: cell.alignment = Alignment(horizontal='center')
    return cell

for t, res in enumerate(RES):
    m = 3 + t * PITCH
    mk = styled(m, 2, f'◆ {res}', font=Font(name='Arial', size=11, bold=True, color='FFFFFFFF'),
                rgb=F_MARK, center=False)
    # band row
    bands = [(3, 7, '◄ CURRENT (actuals: data + calc) ►', F_CUR),
             (9, 13, '◄ SIMULATED (EcoGainsSim v4) ►', F_SIM),
             (15, 19, 'Δ', F_DELTA),
             (NET_C0, NET_C0 + 3, '◄ NET / earner (spend held constant) ►', F_NET)]
    bf = Font(name='Arial', size=11, bold=True, color='FFFFFFFF')
    for c1, c2, txt, rgb in bands:
        for c in range(c1, c2 + 1):
            styled(m + 1, c, rgb=rgb, font=bf)
        ws.cell(m + 1, c1).value = txt
    # header row
    hr = m + 2
    hf = Font(name='Arial', size=9, bold=True)
    styled(hr, 2, 'Segment', font=hf, rgb=F_HDRSEG, border=True)
    for base, rgb in ((3, F_HDRCUR), (9, F_HDRSIM), (15, F_HDRDL)):
        for j, g in enumerate(GROUPS + ['Total']):
            styled(hr, base + j, g, font=hf, rgb=rgb, border=True)
    for j, g in enumerate(NET_HDRS):
        styled(hr, NET_C0 + j, g, font=hf, rgb=F_HDRNET, border=True)

    for pb, payer in enumerate(PAYERS):
        band_r = m + 3 + pb * 7
        styled(band_r, 2, payer, font=Font(name='Arial', size=10, bold=True, color='FFFFFFFF'),
               rgb=F_PAYERBAND, center=False)
        # 5 segment rows + 1 overall row
        rowspec = [(band_r + 1 + i, s, False) for i, s in enumerate(SEGS)] + [(band_r + 6, 'overall', True)]
        for r, label, is_overall in rowspec:
            lf = Font(name='Arial', size=11, bold=is_overall)
            lab = styled(r, 2, label, font=lf)
            if is_overall: lab.fill = fill(F_OVERALL)
            for j in range(4):                                    # gain groups (cur/sim)
                for base, key in ((3, 'cur'), (9, 'sim')):
                    if is_overall:
                        v = wavg(payer, {s: VALS[res][payer][s][key][GROUPS[j]] for s in SEGS})
                    else:
                        v = VALS[res][payer][label][key][GROUPS[j]]
                    c = styled(r, base + j, round(v, 4), fmt=FMT_VAL)
                    if is_overall: c.fill = fill(F_OVERALL)
            for base in (3, 9):                                   # Total = SUM
                c = styled(r, base + 4, f'=SUM({CL(base)}{r}:{CL(base+3)}{r})', fmt=FMT_VAL)
                if is_overall: c.fill = fill(F_OVERALL)
            for j in range(5):                                    # Δ = sim/cur-1
                c = styled(r, 15 + j, f'=IFERROR({CL(9+j)}{r}/{CL(3+j)}{r}-1,"")', fmt=FMT_PCT)
                if is_overall: c.fill = fill(F_OVERALL)
            for j in range(4):                                    # NET block — filled by the menu script
                c = styled(r, NET_C0 + j, None, fmt=FMT_NET)
                if is_overall: c.fill = fill(F_OVERALL)
            # net Δ = new_net - cur_net (live)
            ws.cell(r, NET_C0 + 3).value = f'=IFERROR({CL(NET_C0+2)}{r}-{CL(NET_C0+1)}{r},"")'

wb.save(os.path.join(HERE, '..', 'display', 'Sim_per_Segment_v3.xlsx'))
print('written Sim_per_Segment_v3.xlsx (from', os.path.basename(SRC), ') —', len(RES),
      'tables, PITCH', PITCH, ', markers at', [3 + t * PITCH for t in range(len(RES))])
