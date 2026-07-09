# Restores the ECOGAINS_* custom-function calls that Google sometimes clears from the display
# sheets, WITHOUT disturbing the styling. It starts from the live workbook (so the styling is
# exactly what is in NEW_LIVEOPS_CALENDAR_ECO (N).xlsx, including any hand-edits) and only writes
# the formula cells. Each target sheet is exported as a standalone single-sheet xlsx you can
# re-import into the Google workbook when its formulas go missing.
#
# Run:  python builders/_restore_formulas.py            # uses the highest NEW_LIVEOPS_CALENDAR_ECO
#       python builders/_restore_formulas.py "<path>"   # or point at a specific workbook
#
# Anchors were verified against workbook (7) + the v2 Daily layout (NET blocks added 2026-07-09);
# they follow the builders (_build_hc_v4 / _build_daily / _build_pbp_v6). If a sheet is ever
# restructured, update ANCHORS below. NOTE: the AN9/AZ9/BL9 NET anchors only exist in the Google
# workbook after the EcoGainsSim_Daily_v2 display sheet has been imported.
import os, sys, glob, re
import openpyxl

ROOT = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..')
OUTDIR = os.path.join(ROOT, 'display', 'restored')

def newest_workbook():
    cands = glob.glob(os.path.join(ROOT, 'workbooks', 'NEW_LIVEOPS_CALENDAR_ECO (*).xlsx'))
    def n(p):
        m = re.search(r'\((\d+)\)', os.path.basename(p))
        return int(m.group(1)) if m else -1
    return max(cands, key=n) if cands else None

SRC = sys.argv[1] if len(sys.argv) > 1 else newest_workbook()

# The trailing sim_refresh!$A$1 on HC/Daily/cal_new anchors is the engine's refresh NONCE
# (ignored by the functions; changing it is what re-runs the sims — since the nonce refactor the
# engine never clears formulas, so disappearances should be history; this script stays as the
# manual fallback). PBP keeps bare args: its sheet isn't on the engine's refresh list.
NONCE = 'sim_refresh!$A$1'

# --- HC: 6 segment blocks; SIM in col C, DIFF in col O, at data-row-0 = header row + 2 ---
def hc_formulas():
    cells = {}
    for hdr in (6, 35, 64, 93, 122, 151):
        d0 = hdr + 2
        cells[f'C{d0}'] = f'=LET(payer, $C$3, segment, $B${hdr}, ECOGAINS_SIM(payer, segment, {NONCE}))'
        cells[f'O{d0}'] = f'=LET(payer, $C$3, segment, $B${hdr}, ECOGAINS_DIFF(payer, segment, {NONCE}))'
    return cells

DAILY = {
    'D9':  f'=LET(payer,$C$3, segment,$C$4, source,$C$5, ECOGAINS_DAILY(payer, segment, source, "CURRENT", {NONCE}))',
    'P9':  f'=LET(payer,$C$3, segment,$C$4, source,$C$5, ECOGAINS_DAILY(payer, segment, source, "NEW", {NONCE}))',
    'AB9': f'=LET(payer,$C$3, segment,$C$4, source,$C$5, ECOGAINS_DAILY(payer, segment, source, "DIFF", {NONCE}))',
    # v2 NET blocks (net Δ at BX is plain sheet formulas, not a spill — nothing to restore there)
    'AN9': f'=LET(payer,$C$3, segment,$C$4, source,$C$5, ECOGAINS_DAILY(payer, segment, source, "SPEND", {NONCE}))',
    'AZ9': f'=LET(payer,$C$3, segment,$C$4, source,$C$5, ECOGAINS_DAILY(payer, segment, source, "CURNET", {NONCE}))',
    'BL9': f'=LET(payer,$C$3, segment,$C$4, source,$C$5, ECOGAINS_DAILY(payer, segment, source, "NEWNET", {NONCE}))',
}
PBP = {
    'A14': '=ECOGAINS_PBP_PROFILE($B$5,$B$6)',
    'A23': '=ECOGAINS_PBP_EVENTS($B$3,$B$4,$B$5,$B$6,$B$8)',
    'A52': '=ECOGAINS_PBP($B$3,$B$4,$B$5,$B$6,$B$7,$B$8,$B$9,$B$10,$B$11)',
}
CALNEW = {'E38': f'=ECOGAINS_CAL_COUNTS({NONCE})'}   # CalStats.gs — the instances/event-days summary

# sheet name in the workbook -> {cell: formula}
ANCHORS = {
    'EcoGainsSim_HC':       hc_formulas(),
    'EcoGainsSim_Daily':    DAILY,
    'EcoGainsSim_PlybyPly':  PBP,
    'cal_new':              CALNEW,
}

def export_sheet(sheet_name, formulas):
    wb = openpyxl.load_workbook(SRC, data_only=False)
    if sheet_name not in wb.sheetnames:
        print(f'  !! {sheet_name} not in workbook — skipped'); return
    # keep only the target sheet so the export is a clean, single-sheet, import-ready file
    for s in list(wb.sheetnames):
        if s != sheet_name:
            del wb[s]
    try:
        wb.defined_names = type(wb.defined_names)()   # drop workbook-scoped names (may ref dropped sheets)
    except Exception:
        pass
    ws = wb[sheet_name]
    for cell, f in formulas.items():
        ws[cell] = f
    os.makedirs(OUTDIR, exist_ok=True)
    out = os.path.join(OUTDIR, sheet_name + '.xlsx')
    wb.save(out)
    print(f'  {sheet_name:22s} -> {os.path.relpath(out, ROOT)}   ({len(formulas)} formula cells: {", ".join(sorted(formulas))})')

if __name__ == '__main__':
    if not SRC or not os.path.exists(SRC):
        sys.exit(f'workbook not found: {SRC}')
    print('source workbook:', os.path.relpath(SRC, ROOT))
    for name, cells in ANCHORS.items():
        export_sheet(name, cells)
    print('done. Re-import any of these sheets when its function-call cells get cleared.')
