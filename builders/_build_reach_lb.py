# Builds EventReach_LB_v1.xlsx — replicas of TaD_v2 / F_v2 / Race_v2 (values + styles) from
# NEW_LIVEOPS_CALENDAR_ECO (5).xlsx, each extended with a "Player Rank Simulation" table at
# column AH driven purely by Google Sheets formulas: measured rank percentile (data_event_inst
# position_p25/p50/p75) -> reward row of the sheet's own leaderboard ladder. No population sim.
# Leaderboard counterpart of _build_reach.py (collections); same anchor, inputs, and styling.
from copy import copy
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter as CL

HERE = os.path.dirname(os.path.abspath(__file__))
SRC = os.path.join(HERE, '..', 'workbooks', 'NEW_LIVEOPS_CALENDAR_ECO (5).xlsx')
OUT = os.path.join(HERE, '..', 'display', 'EventReach_LB_v1.xlsx')
SEGS = ['0-9', '10-19', '20-39', '40-99', '100+']
SEG_HDR = ['0-9', '10-19', '20-39', '40-99', '100']       # header labels per user spec

F_YELLOW, F_GRAY, F_HDRGRAY = 'FFFFF2CC', 'FFEFEFEF', 'FFD9D9D9'
fill = lambda rgb: PatternFill('solid', fgColor=rgb)
thin = Side(style='thin')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

# percentile dropdown value -> data_event_inst column (K=p25, L=p50, M=p75)
PCT_COLS = [('p25', 'K'), ('p50', 'L'), ('p75', 'M')]
NCOLS = 21   # every ladder reward block spans 21 columns (Coins .. 6-star Dly)

# events: (display label, data_event_inst event_name, last paying-config position,
#          ladder reward range, ladder header range)
CFG = {
    'TaD_v2': dict(events=[
        ('Target Day (LB)', 'Target Day', 20, '$C$36:$W$55', '$C$35:$W$35')],
        notes=[
            'leaderboard track only — milestone ladder NOT simulated (all milestone reward '
            'cells are 0 in TaD/TaD_v2 config; deliberate scope cut)',
            'LB size 50 per config; measured avg_bots ≈ 0–1; positions 11–20 configured but pay nothing']),
    'F_v2': dict(events=[
        ('Flock Flurry', 'Flock Flurry', 5, '$B$11:$V$15', '$B$10:$V$10')],
        notes=[
            'winner-take-all: only position 1 pays (20 SPT + 4 boosters); SPT is outside the 11-resource universe',
            'the 1h Unlimited Lives on opt-in is a JOIN grant, not a rank reward — not shown here (see flock-flurry.md)']),
    'Race_v2': dict(events=[
        ('Red Challenge',   'Red',        10, '$B$10:$V$19', '$B$9:$V$9'),
        ('Chuck Challenge', 'Chuck',      10, '$B$28:$V$37', '$B$27:$V$27'),
        ('Bomb Challenge',  'Bomb',       10, '$B$46:$V$55', '$B$45:$V$45'),
        ('Level Race',      'Level Race', 10, '$B$64:$V$73', '$B$63:$V$63'),
        ('Flash Race',      'Flash Race', 10, '$B$82:$V$91', '$B$81:$V$81')],
        notes=[
            'Level Race: observed positions exceed configured LBSize=10 (p75 up to 15) — live '
            'board size understated; ranks 11+ shown as below ladder',
            'Flash Race: telemetry shows ≈0 of the 11 tracked resources (event pays SPT) — '
            'ladder HC here is config-nominal, uncorroborated',
            'all five ladders are identical (family template) — authenticity vs live config unverified']),
}

src = openpyxl.load_workbook(SRC)          # these sheets are values-only
out = openpyxl.Workbook()
out.remove(out.active)

def copy_sheet(name):
    s, t = src[name], out.create_sheet(name)
    t.sheet_view.showGridLines = s.sheet_view.showGridLines
    for row in s.iter_rows():
        for c in row:
            if c.value is None and not c.has_style:
                continue
            n = t.cell(c.row, c.column, c.value)
            if c.has_style:
                n.font = copy(c.font); n.fill = copy(c.fill); n.border = copy(c.border)
                n.alignment = copy(c.alignment); n.number_format = c.number_format
    for k, dim in s.column_dimensions.items():
        if dim.width: t.column_dimensions[k].width = dim.width
        if dim.hidden: t.column_dimensions[k].hidden = True
    for k, dim in s.row_dimensions.items():
        if dim.height: t.row_dimensions[k].height = dim.height
    for m in s.merged_cells.ranges:
        t.merge_cells(str(m))
    return t

def pos_lookup(col, event, seg):
    return (f"SUMIFS(data_event_inst!${col}:${col},"
            f"data_event_inst!$A:$A,\"{event}\","
            f"data_event_inst!$B:$B,$AI$3,"
            f"data_event_inst!$C:$C,\"{seg}\")")

def build_block(ws, cfg):
    ws['AH1'] = 'Player Rank Simulation (measured rank percentile → ladder reward) — SIMULATED'
    ws['AH1'].font = Font(name='Arial', size=11, bold=True)
    for r, (label, default, dv_list) in enumerate([
            ('Percentile', 'p50', '"' + ','.join(p for p, _ in PCT_COLS) + '"'),
            ('Payer', 'NONPAYER', '"NONPAYER,PAYER"')]):
        lab = ws.cell(2 + r, 34, label)
        lab.font = Font(name='Arial', size=9, bold=True)
        v = ws.cell(2 + r, 35, default)
        v.font = Font(name='Arial', size=9, bold=True)
        v.fill = fill(F_YELLOW)
        v.alignment = Alignment(horizontal='center')
        v.border = BORDER
        dv = DataValidation(type='list', formula1=dv_list, allow_blank=False, showDropDown=False)
        ws.add_data_validation(dv)
        dv.add(v)
    # header row 5
    hdr_cells = ['Event']
    for s in SEG_HDR:
        hdr_cells += [f'{s}_pos', f'{s}_reward']
    for j, txt in enumerate(hdr_cells):
        c = ws.cell(5, 34 + j, txt)
        c.font = Font(name='Arial', size=9, bold=True)
        c.fill = fill(F_HDRGRAY)
        c.alignment = Alignment(horizontal='center')
        c.border = BORDER
    # one row per event
    for i, (label, event, nmax, rew, hdr) in enumerate(cfg['events']):
        r = 6 + i
        lc = ws.cell(r, 34, label)
        lc.fill = fill(F_GRAY); lc.border = BORDER
        lc.font = Font(name='Arial', size=9)
        for si, seg in enumerate(SEGS):
            pos_col, rw_col = 35 + si * 2, 36 + si * 2
            pos_ref = f'{CL(pos_col)}{r}'
            lookups = ','.join(pos_lookup(col, event, seg) for _, col in PCT_COLS)
            pl = ','.join(f'"{p}"' for p, _ in PCT_COLS)
            pos_f = (f'=LET(pos,CHOOSE(MATCH($AI$2,{{{pl}}},0),{lookups}),'
                     f'IF(pos=0,"n/a",pos))')
            pc = ws.cell(r, pos_col, pos_f)
            pc.fill = fill(F_GRAY); pc.border = BORDER; pc.number_format = '0'
            pc.alignment = Alignment(horizontal='center')
            rw_f = (f'=LET(p,{pos_ref},IF(p="n/a","n/a",IF(p>{nmax},"— (below ladder)",'
                    f'LET(s,INDEX({rew},p,0),'
                    f'b,TEXTJOIN(", ",TRUE,ARRAYFORMULA(IF(s>0,{hdr}&": "&ROUND(s,2),""))),'
                    f'IF(b="","{{}} (rank pays nothing)","{{"&b&"}}")))))')
            rc = ws.cell(r, rw_col, rw_f)
            rc.fill = fill(F_GRAY); rc.border = BORDER
            rc.alignment = Alignment(horizontal='left')
    # notes
    nr = 6 + len(cfg['events']) + 1
    notes = ['p25 = top-quartile finisher (lower position = better); source: data_event_inst '
             'position percentiles, measured on the CURRENT calendar'] + cfg['notes']
    for k, txt in enumerate(notes):
        c = ws.cell(nr + k, 34, '⚠ ' + txt if k else txt)
        c.font = Font(name='Arial', size=8, color='FF808080')
    # widths
    ws.column_dimensions['AG'].width = 2.88
    ws.column_dimensions['AH'].width = 16.0
    for si in range(5):
        ws.column_dimensions[CL(35 + si * 2)].width = 7.0
        ws.column_dimensions[CL(36 + si * 2)].width = 34.0

for name, cfg in CFG.items():
    ws = copy_sheet(name)
    build_block(ws, cfg)

out.save(OUT)
print('written', OUT, '— sheets:', out.sheetnames)
