# Rewrites the on-sheet "Player Reach Simulation" on Ph_v2 to model a GREEDY-SMART buyer instead
# of the current sequential one (AT36 note: "items bought sequentially").
#
# Greedy model (confirmed with the user):
#   - The resource bonus sits on item 3 of each set; the player always buys item 3 to grab it.
#   - To unlock set k they must own >= its threshold (col E: 0,2,4,6,8,10,13,16,20,25), so they buy
#     the CHEAPEST filler items needed to hit each gate - nothing more ("smart").
#   - Sets therefore complete in order; the token cost to COMPLETE each set (min-cost path) is:
GREEDY_SET_COST = [155, 395, 640, 985, 1415, 1845, 2555, 3415, 4675, 6735]     # sets 1..10
#   - "Last-day leftover -> buy any item": leftover tokens go to non-reward items, so they never
#     complete an extra set -> no effect on the reward. (Captured: reward = sets whose cost<=tokens.)
#
# What changes on the sheet (nothing else is touched):
#   - New helper AW/AX (rows 4-14): the 10 greedy set-completion costs. The engine's R-term axis in
#     column AU is LEFT ALONE (Photoshoot R=1 regardless, so it doesn't matter, but we don't disturb it).
#   - _ms columns (AI/AK/AM/AO/AQ, rows 6-12): count greedy SET costs affordable, i.e. sets completed,
#     by pointing the COUNTIF at $AX$5:$AX$14 instead of the sequential $AU$5:$AU$34.
#   - _reward columns (AJ/AL/AN/AP/AR): sum the first m SET bonuses (item-3 rows) instead of the first
#     m item rows -> the MMULT matrix becomes the 10 item-3 rows, selected robustly by set count.
#   - Title (AH1) + a note explaining the greedy model.
import os, openpyxl
from openpyxl.worksheet.formula import ArrayFormula

ROOT = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..')
SRC  = os.path.join(ROOT, 'workbooks', 'NEW_LIVEOPS_CALENDAR_ECO (7).xlsx')
OUT  = os.path.join(ROOT, 'display', 'Ph_v2_greedy_localsim.xlsx')

# item-3 (set-completion) reward rows on the ladder, in set order
ITEM3_ROWS = [27, 30, 33, 36, 39, 42, 45, 48, 51, 54]
VSTACK_SETBONUS = 'VSTACK(' + ','.join(f'$H${r}:$AB${r}' for r in ITEM3_ROWS) + ')'
# robust "sum first m rows of the 10-row set-bonus matrix": 1x10 selector (first m ones) x 10x21
SEL_SUM = f'MMULT(TRANSPOSE((SEQUENCE(10,1,1,1)<=m)*1),{VSTACK_SETBONUS})'

MS_COLS  = {'0-9':'AI', '10-19':'AK', '20-39':'AM', '40-99':'AO', '100+':'AQ'}
REW_COLS = {'0-9':'AJ', '10-19':'AL', '20-39':'AN', '40-99':'AP', '100+':'AR'}
ROWS = range(6, 13)   # day rows 1..7 (only <= EventDuration render, rest resolve to "")

def ftext(cell):
    v = cell.value
    return v.text if isinstance(v, ArrayFormula) else v

def main():
    wb = openpyxl.load_workbook(SRC, data_only=False)
    # keep only Ph_v2 -> clean single-sheet, import-ready file (same as the restore exports)
    for s in list(wb.sheetnames):
        if s != 'Ph_v2':
            del wb[s]
    try: wb.defined_names = type(wb.defined_names)()
    except Exception: pass
    ws = wb['Ph_v2']

    n_ms = n_rew = 0
    for seg, mscol in MS_COLS.items():
        rewcol = REW_COLS[seg]
        for r in ROWS:
            # _ms: point the affordability COUNTIF at the greedy set costs (=> "sets completed")
            mcell = ws[f'{mscol}{r}']
            t = ftext(mcell)
            if isinstance(t, str) and '$AU$5:$AU$34' in t:
                ws[f'{mscol}{r}'] = ArrayFormula(f'{mscol}{r}', t.replace('$AU$5:$AU$34', '$AX$5:$AX$14'))
                n_ms += 1
            # _reward: sum the first m SET bonuses instead of the first m item rows
            rcell = ws[f'{rewcol}{r}']
            t2 = ftext(rcell)
            if isinstance(t2, str) and '$H$25:$AB$54' in t2:
                t2 = t2.replace('MMULT(SEQUENCE(1,m,1,0),$H$25:$AB$54)', SEL_SUM)
                ws[f'{rewcol}{r}'] = ArrayFormula(f'{rewcol}{r}', t2)
                n_rew += 1

    # greedy helper table (fresh columns; engine's AU axis untouched)
    ws['AW4'] = 'Set #'
    ws['AX4'] = 'Greedy cum. cost'
    for i in range(10):
        ws[f'AW{5+i}'] = i + 1
        ws[f'AX{5+i}'] = GREEDY_SET_COST[i]

    ws['AH1'] = 'Player Reach Simulation (per event day) - SIMULATED - GREEDY (reward-item-first)'
    ws['AW16'] = ('GREEDY model: buy each set\'s reward item (item 3) first + the cheapest fillers '
                  'needed to unlock the set (col E thresholds). AX = cumulative tokens to COMPLETE '
                  'the set. _ms = # sets (reward milestones) completed; _reward sums those sets\' '
                  'bonuses. Last-day leftover tokens buy non-reward items (no extra reward).')

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    wb.save(OUT)
    print(f'patched _ms cells: {n_ms}   _reward cells: {n_rew}')
    print(f'greedy set costs (AX5:AX14): {GREEDY_SET_COST}')
    print(f'written: {os.path.relpath(OUT, ROOT)}')

if __name__ == '__main__':
    main()
