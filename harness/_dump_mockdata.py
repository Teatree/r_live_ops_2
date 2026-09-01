# Regenerates _mockdata.json from the current workbook of record (highest NEW_LIVEOPS_CALENDAR_ECO).
# Dumps values (data_only=True — cached results of live formulas) + merges for every sheet the
# engines read: EcoGainsSim_v4.gs, EcoGainsSim_Daily.gs, EcoGainsSim_PBP.gs. Run after every
# workbook re-export, before offline harness runs (_mock_run.js / _mock_daily.js / _mock_pbp.js).
#
# --workbook / --out (2026-09-01, D24). The default glob only ever matched NEW_LIVEOPS_CALENDAR_ECO*,
# so the CARD-COLLECTION workbooks (COLLECTIONS_UNDER_NEW_CALENDAR*) could not be dumped at all and
# _mock_cards.js had no data to run against: the workbook of record reverted PackConfig pre-D19, so
# the card harness died on 'PACK DEFINITIONS block is empty' rather than reporting a single gate.
# The two lineages need two dumps, not one:
#   python harness/_dump_mockdata.py
#   python harness/_dump_mockdata.py --workbook "workbooks/COLLECTIONS_UNDER_NEW_CALENDAR (3).xlsx" \
#          --out harness/_mockdata_collections.json
import argparse
import datetime
import glob
import json
import os
import re
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))

ap = argparse.ArgumentParser(description="Dump a workbook to a harness mockdata JSON.")
ap.add_argument('--workbook', help='path to the .xlsx to dump (default: highest-numbered '
                                   'workbooks/NEW_LIVEOPS_CALENDAR_ECO*.xlsx)')
ap.add_argument('--out', help='output json path (default: harness/_mockdata.json)')
args = ap.parse_args()

if args.workbook:
    SRC = (args.workbook if os.path.isabs(args.workbook)
           else os.path.join(HERE, '..', args.workbook))
    if not os.path.exists(SRC):
        # also accept a bare filename living in workbooks/
        alt = os.path.join(HERE, '..', 'workbooks', os.path.basename(args.workbook))
        if not os.path.exists(alt):
            raise SystemExit('No such workbook: ' + args.workbook)
        SRC = alt
else:
    books = glob.glob(os.path.join(HERE, '..', 'workbooks', 'NEW_LIVEOPS_CALENDAR_ECO*.xlsx'))
    if not books:
        raise SystemExit('No NEW_LIVEOPS_CALENDAR_ECO*.xlsx in workbooks/ — pass --workbook.')
    books.sort(key=lambda n: int((re.search(r'\((\d+)\)', n) or [0, 0])[1]))
    SRC = books[-1]

OUT = ((args.out if os.path.isabs(args.out)
        else os.path.join(HERE, '..', args.out))
       if args.out else os.path.join(HERE, '_mockdata.json'))

SHEETS = [
    # v4 engine
    'data_gains', 'data_seg_beh', 'data_event_accrual', 'data_event_kite_accrual', 'data_RM',
    'cal_curr', 'cal_new', 'cal_parsed',
    'c_saga', 'c_saga_v2', 'c_day', 'c_day_v2', 'RM', 'NS', 'NS_v2', 'Sim per Segment',
    # The display sheet itself, so a gate can check CATEGORY_ORDER against the row LABELS the
    # spill lands next to. Those labels are static text nothing validated, so a row added to
    # the sheet without a matching CATEGORY_ORDER entry shifted every row below it onto the
    # wrong source in silence (found 2026-08-21).
    'EcoGainsSim', 'EcoGainsSim_HC',
    # PBP engine additions
    'data_streaks', 'data_event_inst',
    'J_v2', 'HH_v2', 'BB_v2', 'Ph_v2', 'Ki_v2', 'TaD_v2', 'Race_v2', 'F_v2',
    # base config sheets (R-term pairs: reward-config ratio v2/base, added 2026-07-06)
    'J', 'HH', 'BB', 'Ph', 'Ki', 'TaD', 'Race', 'F',
    # NET inputs (SimPerSegmentFill per-earner NET / ECOGAINS_DAILY net blocks) — expected MISSING
    # until the per-earner data_econ re-pull and the new data_econ_daily sheet land in the workbook
    'data_econ', 'data_econ_daily',
    # Season Pass (D16): track + challenge configs. The _v2 pair is expected MISSING until the
    # user duplicates SP -> SP_v2 / SP_lb -> SP_lb_v2 in the live workbook (the engine then falls
    # back to the base sheets — that fallback path is exactly what the offline gates exercise).
    'SP', 'SP_lb', 'SP_v2', 'SP_lb_v2',
    # Rainbow Maker split configs (2026-07-10, hardcoded RM_1st x3 / RM_2nd x2 — see CLAUDE.md):
    # expected MISSING until the next workbook export; the engine falls back to 'RM'.
    'RM_1st', 'RM_2nd', 'RM_1st_v2', 'RM_2nd_v2',
    # Card collection (D19, 2026-08-03). TE feeds the Team Event pack overlay (PACK_ONLY_SPECS);
    # the rest feed CardOpenings.gs / _mock_cards.js. EcoPackGains and PlayerBehavior are GONE.
    'TE', 'PackConfig', 'AlbumConfig', 'CardPoolConfig', 'Col_Cards_Daily',
]

# Sheets that have been REBUILT by a builder but not yet imported into the live workbook. The
# generated display xlsx overlays whatever the workbook still has, so the offline harness tests
# the layout the engine actually expects. Drop an entry once the sheet is imported (the overlay
# then just reproduces the workbook).  {sheet name: display xlsx filename}
# PackConfig / SimOutput were imported as of workbook (14) — the workbook is now the truth.
# Sheets whose LOCAL build supersedes the workbook copy until it is re-imported. SimOutput is
# listed again (2026-08-18): the pack log gained an 'Earned From' column and the album grids moved
# J -> L, so the workbook's copy is a layout the engine no longer writes. Drop the entry once the
# sheet has been re-imported into the live workbook.
# Col_Cards_Daily DROPPED 2026-09-01: the collections workbook now ships the current layout
# (Source_Detail in D, album grids at L), so the overlay was replacing the real sheet - and
# with it the user's own hand-added formulas - by a builder artefact. The ECO lineage has no
# Col_Cards_Daily at all, which is honest: that workbook has no card sim. Run the card
# harnesses against the collections dump instead (--data collections).
PENDING_IMPORT = {}


def dump_sheet(ws):
    vals = []
    for row in ws.iter_rows(values_only=True):
        vals.append(['' if v is None else (v if isinstance(v, (int, float, bool)) else str(v))
                     for v in row])
    while vals and all(v == '' for v in vals[-1]):
        vals.pop()
    merges = [{'r': m.min_row, 'c': m.min_col,
               'nr': m.max_row - m.min_row + 1, 'nc': m.max_col - m.min_col + 1}
              for m in ws.merged_cells.ranges]
    return {'values': vals, 'merges': merges}


wb = openpyxl.load_workbook(SRC, data_only=True)
out = {}
for name in SHEETS:
    if name not in wb.sheetnames:
        print('MISSING sheet:', name)
        continue
    out[name] = dump_sheet(wb[name])

for name, fname in PENDING_IMPORT.items():
    p = os.path.join(HERE, '..', 'display', fname)
    if not os.path.exists(p):
        print('PENDING_IMPORT source not built yet:', fname)
        continue
    pw = openpyxl.load_workbook(p, data_only=True)
    src_ws = pw[name] if name in pw.sheetnames else pw.worksheets[0]
    out[name] = dump_sheet(src_ws)
    print(f'OVERLAY {name} <- display/{fname} (rebuilt, not yet imported into the workbook)')

# Provenance. A dump is indistinguishable from any other once written, and this repo now keeps two
# (the ECO lineage and the collections lineage) — a harness reading the wrong one reports confusing
# failures rather than an obvious 'wrong file'. Not a sheet: no harness iterates the top-level keys.
out['_meta'] = {'source': os.path.basename(SRC),
                'dumped': datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'sheets': sorted(k for k in out if k != '_meta')}

with open(OUT, 'w', encoding='utf-8') as f:
    json.dump(out, f)
print('written', os.path.relpath(OUT, os.path.join(HERE, '..')), 'from',
      os.path.basename(SRC), '—', len(out) - 1, 'sheets')
