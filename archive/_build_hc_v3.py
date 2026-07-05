# Builds EcoGainsSim_HC_v3.xlsx — the EcoGainsSim_HC display sheet with the new 'Saga' row
# (25 categories), replicating the style of the sheet in NEW_LIVEOPS_CALENDAR_ECO (4).xlsx.
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

CATS = ['Ads', 'Bomb Challenge', "Bomb's Ballet", 'Chuck Challenge', 'Core', 'Daily Gift',
        'Daily Night Sky Prize', 'Flock Flurry', 'Hatchling Hideaway', 'Jigsaw', 'Kite Festival',
        'Level Race', 'Other', 'Photoshoot', 'Red Challenge', 'River Rush', 'Saga',
        'Season Pass (Free)', 'Target Day', 'Team Event', 'Team Race', 'Flash Race',
        'FlowerCoop', 'Rainbow Maker', 'IAPs']
RES = ['HC', 'Slingshot', 'Shuffle', 'Comet', 'Red', 'Chuck', 'Bomb',
       'UL Bomb', 'UL Chuck', 'UL Red', 'Unlimited Lives']
ALWAYS_ON = {'Daily Gift', 'Saga'}                      # yellow label + gray sim data
EVENT_SIM = {'Bomb Challenge', "Bomb's Ballet", 'Chuck Challenge', 'Hatchling Hideaway', 'Jigsaw',
             'Kite Festival', 'Level Race', 'Photoshoot', 'Red Challenge', 'River Rush',
             'Target Day', 'Flash Race', 'Rainbow Maker'}  # pink label + gray sim data
# everything else (incl. Core, now carried) -> plain label + blue data fill

SEGMENTS = [('0-9', '0-9 Segment - Over Time Period  (uses 1-9 gains data)'),
            ('10-19', '10-19 Segment - Over Time Period'),
            ('20-39', '20-39 Segment - Over Time Period'),
            ('40-99', '40-99 Segment - Over Time Period'),
            ('100+', '100+ Segment - Over Time Period')]

F_BLUE, F_GRAY = 'FFDDEBF7', 'FFEFEFEF'
F_YELLOW, F_PINK, F_ORANGE, F_HDR = 'FFFFF2CC', 'FFF4CCCC', 'FFFCE5CD', 'FFD9D9D9'
fill = lambda rgb: PatternFill('solid', fgColor=rgb)
thin = Side(style='thin')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
FMT_SIM, FMT_DIFF = '#,##0.00;[Red]-#,##0.00', '#,##0.00;-#,##0.00'

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'EcoGainsSim_HC'
ws.sheet_view.showGridLines = False
for col, w in {'A': 1.75, 'B': 19.25, 'C': 8.25, 'Z': 7.63}.items():
    ws.column_dimensions[col].width = w

ws['B2'] = 'Per-earner gains over the 33-day'
ws['B2'].font = Font(name='Arial', size=12, bold=True)
ws['B3'] = 'Payer'
ws['B3'].font = Font(name='Arial', size=10, bold=True)
ws['C3'] = 'PAYER'
ws['C3'].font = Font(name='Arial', size=10, bold=True)
ws['C3'].fill = fill(F_YELLOW)
ws['D3'] = '◀ input (global)'
ws['D3'].font = Font(name='Arial', size=9, color='FF808080')

N = len(CATS)                       # 25
first_hdr = 6
sim_rows = []                       # first-table row numbers of simulated categories (for J3)
for t, (seg, title) in enumerate(SEGMENTS):
    hdr = first_hdr + t * (N + 4)   # 6, 35, 64, 93, 122
    cols_row, d0 = hdr + 1, hdr + 2
    tag = ws.cell(hdr, 2, seg)
    tag.font = Font(name='Arial', size=10, bold=True)
    tag.fill = fill(F_ORANGE)
    tag.alignment = Alignment(horizontal='center')
    for c, txt in [(3, title), (15, title.split('  (')[0] + ' - Difference')]:
        ws.cell(hdr, c, txt).font = Font(name='Arial', size=11, bold=True)
    hf = Font(name='Arial', size=9, bold=True)
    ws.cell(cols_row, 2, 'Source').font = hf
    ws.cell(cols_row, 2).fill = fill(F_HDR)
    ws.cell(cols_row, 2).border = BORDER
    for j, r in enumerate(RES):
        for base in (3, 15):        # C.. and O..
            cell = ws.cell(cols_row, base + j, r)
            cell.font, cell.fill, cell.border = hf, fill(F_HDR), BORDER
            cell.alignment = Alignment(horizontal='center')
    for i, cat in enumerate(CATS):
        row = d0 + i
        lab = ws.cell(row, 2, cat)
        lab.font = Font(name='Arial', size=9, bold=True)
        lab.border = BORDER
        if cat in ALWAYS_ON:
            lab.fill = fill(F_YELLOW)
        elif cat in EVENT_SIM:
            lab.fill = fill(F_PINK)
        data_fill = fill(F_GRAY) if (cat in ALWAYS_ON or cat in EVENT_SIM) else fill(F_BLUE)
        for j in range(len(RES)):
            c1 = ws.cell(row, 3 + j)
            c1.fill, c1.border, c1.number_format = data_fill, BORDER, FMT_SIM
            c2 = ws.cell(row, 15 + j)
            c2.border, c2.number_format = BORDER, FMT_DIFF
        if t == 0 and (cat in ALWAYS_ON or cat in EVENT_SIM):
            sim_rows.append(row)
    ws.cell(d0, 3).value = f'=LET(payer, $C$3, segment, $B${hdr}, ECOGAINS_SIM(payer, segment))'
    ws.cell(d0, 15).value = f'=LET(payer, $C$3, segment, $B${hdr}, ECOGAINS_DIFF(payer, segment))'
    diff_rng = f'O{d0}:Y{d0 + N - 1}'
    ws.conditional_formatting.add(diff_rng, CellIsRule(
        operator='lessThan', formula=['0'],
        font=Font(color='FF990000', bold=False), fill=fill(F_PINK)))
    ws.conditional_formatting.add(diff_rng, CellIsRule(
        operator='greaterThan', formula=['0'],
        font=Font(color='FF006100', bold=False), fill=fill('FFD9EAD3')))
    for r in range(hdr, d0 + N):
        ws.row_dimensions[r].height = 15.75

ws['J3'] = '=JOIN(",", {' + ','.join(f'B{r}' for r in sim_rows) + '})'

wb.save('EcoGainsSim_HC_v3.xlsx')
print('written EcoGainsSim_HC_v3.xlsx')
print('tables at hdr rows:', [first_hdr + t * (N + 4) for t in range(5)])
print('J3 sim rows:', sim_rows)
