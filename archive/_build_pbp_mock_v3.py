# Builds EcoGainsSim_PlybyPly_v3.xlsx — real-data mock of the play-by-play session sim,
# superseding v2 (_build_pbp_mock.py). All scenario numbers are hand-derived from REAL
# workbook (6) data: data_streaks (win rate, p_continue_after_win, streak percentiles),
# data_event_inst v2 (Kite rows, final_balance_p25/50/75), accrual curves, config ladders.
# Scenario: Day 5 (Sunday) / 10-19 / NONPAYER / cal_new / Luck p50 / Sampled seed 42.
# Engine (ECOGAINS_PBP) still not wired — this is the build spec.
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter as CL

OUT = 'EcoGainsSim_PlybyPly_v3.xlsx'
F_INPUT, F_DATA, F_SIM, F_HDR = 'FFFFF2CC', 'FFCFE2F3', 'FFE2EFDA', 'FFD9D9D9'
RED = 'FFCC0000'
fill = lambda rgb: PatternFill('solid', fgColor=rgb)
thin = Side(style='thin')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
A9 = Font(name='Arial', size=9)
A9B = Font(name='Arial', size=9, bold=True)
A8G = Font(name='Arial', size=8, color='FF808080')

INV = ['Coins', 'SPT', 'SPT x2', 'Red', 'Chuck', 'Bomb', 'Slingshot', 'Shuffle', 'Comet',
       'Unlimited Lives', 'Unlimited Red', 'Unlimited Chuck', 'Unlimited Bomb']
CATS = ['Ads', 'Bomb Challenge', "Bomb's Ballet", 'Chuck Challenge', 'Core', 'Daily Gift',
        'Daily Night Sky Prize', 'Flock Flurry', 'Hatchling Hideaway', 'Jigsaw', 'Kite Festival',
        'Level Race', 'Other', 'Photoshoot', 'Red Challenge', 'River Rush', 'Saga',
        'Season Pass (Free)', 'Target Day', 'Team Event', 'Team Race', 'Flash Race',
        'FlowerCoop', 'Rainbow Maker', 'IAPs']

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'EcoGainsSim_PlybyPly'
ws.sheet_view.showGridLines = False

def put(r, c, v, font=A9, fl=None, border=False, align=None, numfmt=None):
    cell = ws.cell(r, c, v)
    cell.font = font
    if fl: cell.fill = fill(fl)
    if border: cell.border = BORDER
    if align: cell.alignment = Alignment(horizontal=align)
    if numfmt: cell.number_format = numfmt
    return cell

def dv_cell(r, c, v, options):
    cell = put(r, c, v, A9B, F_INPUT, True, 'center')
    d = DataValidation(type='list', formula1='"' + ','.join(options) + '"',
                       allow_blank=False, showDropDown=False)
    ws.add_data_validation(d)
    d.add(cell)

# ---------- title + banner ----------
put(1, 2, 'EcoGainsSim_PlybyPly — one player, one day, play by play', Font(name='Arial', size=12, bold=True))
put(2, 2, 'MOCK v3 — values hand-derived from REAL workbook (6) data for Day 5 / 10-19 / NONPAYER / '
          'cal_new / Luck p50 / seed 42. Engine (ECOGAINS_PBP) not wired yet — this sheet is its build spec.',
    Font(name='Arial', size=9, bold=True, color=RED))

# ---------- Config Panel ----------
put(4, 2, 'Config Panel (input)', A9B)
cfg = [('Calendar', 'cal_new', ['cal_curr', 'cal_new']),
       ('Day to Simulate (1-33)', 5, None),
       ('Player Segment', '10-19', ['0-9', '10-19', '20-39', '40-99', '100+']),
       ('Payer', 'NONPAYER', ['NONPAYER', 'PAYER']),
       ('Mode', 'Sampled', ['Expected', 'Sampled']),
       ('Luck (progress + LB placement)', 'p50', ['p25', 'p50', 'p75']),
       ('Seed (Sampled mode)', 42, None),
       ('Levels Played (blank = auto)', '', None)]
for i, (label, val, opts) in enumerate(cfg):
    put(5 + i, 2, label, A9)
    if opts: dv_cell(5 + i, 3, val, opts)
    else: put(5 + i, 3, val, A9B, F_INPUT, True, 'center')

# ---------- Player Profile (data) ----------
put(4, 5, 'Player Profile (data — data_streaks + data_seg_beh)', A9B)
prof = [('Attempts / active day', 26.0, 'data_streaks.attempts_per_day_mean → N = 26'),
        ('Win rate', '52.6%', 'data_streaks.win_rate_mean (× seg_beh cross-check ✓)'),
        ('P(win | prev win)', '78.6%', 'data_streaks.p_continue_after_win — streak persistence, incl. OOM saves'),
        ('Max win streak / day (p50)', 7, 'data_streaks.max_streak_per_day_p50 → today\'s showcase streak'),
        ('Mean streak length', 4.1, 'data_streaks.mean_streak_len'),
        ('Login streak (p50)', 3, 'data_seg_beh.login_streak_p50 → Daily Gift cycle day 3'),
        ('NS max streak (p50)', 7, 'data_seg_beh.daily_max_streak_p50 → Night Sky night 7'),
        ('Daily Gift claim rate', '89.5%', 'data_seg_beh.daily_gift_claim_rate_pct'),
        ('P(active on this day)', '33.6%', 'sun_active_rate — sim CONDITIONS on the player playing')]
for i, (label, val, srcnote) in enumerate(prof):
    put(5 + i, 5, label, A9)
    put(5 + i, 7, val, A9B, F_DATA, True, 'center')
    put(5 + i, 8, srcnote, A8G)

# ---------- Day Context (sim) ----------
put(4, 13, 'Day Context (sim)', A9B)
day = [('Day of week', 'Sunday (weekend)'), ('Weekend?', 'YES'),
       ('Active config set', 'cal_new → _v2 configs'),
       ('Events running today', '6 (see Active Events)')]
for i, (label, val) in enumerate(day):
    put(5 + i, 13, label, A9)
    put(5 + i, 15, val, A9B, F_SIM, True, 'center')

# ---------- Active Events ----------
r0 = 16
put(r0 - 1, 2, 'Active Events on this day (sim — cal_parsed × data_event_inst × accrual curves × config ladders; progress at Luck=p50)', A9B)
hdrs = ['Event', 'Family', 'Inst (days)', 'Event day', 'Ends today?', 'Progress @ session start',
        'Accrual model today', 'Next milestone', 'End-of-day payout basis']
for j, h in enumerate(hdrs):
    put(r0, 2 + j, h, A9B, F_HDR, True, 'center')
ev = [
 ('Jigsaw', 'collection', '3–5', '3 / 3', 'YES',
  '213 tok = final_balance_p50 316.8 × curve d2/d3 (.502/.747)',
  '+7.4 tok / win (103.7 left ÷ 13.7 exp. wins)',
  'm4 @ 250 → {UL Lives: 30 min}; m5 @ 370 out of reach at p50',
  'milestones claimed during play (no LB)'),
 ('Kite Festival', 'score + LB (60-board)', '3–5', '3 / 3', 'YES',
  '5,187 = score_p50 7,264 × kite curve d2/d3 (.332/.466)',
  'streak-ladder score / win: 1/10/100/200/500/1000 (Ki_v2)',
  'm1 @ 100 banked day 1 (not in today\'s ledger)',
  'rank p50 = 27 of 60 → BELOW paying range (ladder pays ~top 25) → {}'),
 ('Target Day', 'score + LB (50-board)', '5', '1 / 1', 'YES',
  '0 (instance starts today)',
  'streak-mult score / win: ×1/5/10/20/100 (TaD_v2), calibrated to balance_p50 ≈ 357/day',
  'milestones all-zero (not simmed, per decision)',
  'LB rank p50 = 6 → {Slingshot: 1}'),
 ('Flash Race', 'leaderboard', '5', '1 / 1', 'YES', '0 (starts today)', 'rank race (not token)',
  '—', 'LB rank p50 = 3 → {Coins: 50, Shuffle: 1}'),
 ('Night Sky', 'daily streak', '5', '1 / 1', 'YES', 'streak night 7', '+1 night if active',
  'round 2 @ 13 nights → {Coins: 30, Red: 1}', 'not reached today (7 < 13) — no payout'),
 ('Daily Gift', 'always-on', '—', 'cycle day 3', '—', '—', '—', '—',
  'claimed at session start: {Slingshot: 1} (c_day_v2 V1 d3, claim p 89%)'),
]
for i, row in enumerate(ev):
    for j, v in enumerate(row):
        put(r0 + 1 + i, 2 + j, v, A9, F_SIM, True)
put(r0 + 7, 2, 'excluded (not play-driven / carried): Team Race, Season Pass, offer rows, Ads, Other, IAPs', A8G)

# ---------- Ledger ----------
L = r0 + 10
put(L - 1, 2, 'Play-by-play ledger (sim — S = session-start claims + opening inventory; E = day-end LB payouts; Sampled draws, seed 42)', A9B)
c_play, c_ev0 = 3, 8
c_inv0 = 11
c_spend = c_inv0 + 13 + 1
c_src0 = c_spend + 2
put(L, c_ev0, 'Event Progress (cum)', A9B)
put(L, c_inv0, 'Inventory (running)', A9B)
put(L, c_src0, 'Source Gains (bundle on the play it lands)', A9B)
cols = (['Play #', 'Level', 'Win?', 'Streak', 'Mult'] +
        ['Jigsaw tok', 'Kite score', 'TaD score'] + INV + ['', 'Spend', ''] + CATS)
for j, h in enumerate(cols):
    if h: put(L + 1, c_play + j, h, A9B, F_HDR, True, 'center')
put(L + 1, 2, 'Row', A9B, F_HDR, True, 'center')

inv = dict(zip(INV, [250, 120, 0, 2, 1, 0, 3, 2, 1, 0, 0, 0, 0]))

# (tag, level, win, {grants}); Kite/TaD score per win follow their streak ladders
KITE_STEP = {1: 1, 2: 10, 3: 100, 4: 200, 5: 500}     # 6+ -> 1000
TAD_STEP = {1: 1, 2: 5, 3: 10, 4: 20}                 # 5+ -> 100
pattern = [
    ('S', None, None, {'Daily Gift': '{Slingshot: 1}'}),
    (1, 100, True, {}), (2, 101, True, {}), (3, 102, True, {}), (4, 103, True, {}),
    (5, 104, True, {'Jigsaw': '{Unlimited Lives: 30} (m4 @ 250 crossed)'}),
    (6, 105, True, {}), (7, 106, True, {}),
    (8, 107, False, {'__spend': '{Coins: 100} (OOM — 7-streak save w/ TaD ×100 + Kite ×1000 live; failed)'}),
    (9, 107, True, {'Core': '{Coins: 15} (chapter complete)'}),
    (10, 108, False, {}),
    ('⋮', '… plays 11–24: 14 plays, 4 wins (short streaks: +27 TaD, +22 Kite, +29.6 Jigsaw) …', None, {}),
    (25, 112, True, {}), (26, 113, True, {}),
    ('E', None, None, {'Flash Race': '{Coins: 50, Shuffle: 1} (rank 3)',
                       'Target Day': '{Slingshot: 1} (rank 6)',
                       'Kite Festival': '{} (rank 27 of 60 — below paying range)'}),
]
jig, kite, tad, streak = 213.1, 5187, 0, 0
r = L + 2
for row in pattern:
    tag = row[0]
    put(r, 2, {'S': 'S — start', 'E': 'E — day end'}.get(tag, tag), A9B, F_HDR, True, 'center')
    if tag == '⋮':
        put(r, c_play, row[1], A8G)
        jig += 29.6; kite += 22; tad += 27; streak = 1   # ends mid short streak
        r += 1; continue
    grants = row[3]
    if tag == 'S':
        if 'Daily Gift' in grants: inv['Slingshot'] += 1
        put(r, c_play, 0, A9, F_SIM, True, 'center')
        put(r, c_play + 1, '—', A9, F_SIM, True, 'center')
        put(r, c_play + 2, 'claims', A9, F_SIM, True, 'center')
    elif tag == 'E':
        inv['Coins'] += 50; inv['Shuffle'] += 1; inv['Slingshot'] += 1
        put(r, c_play, '—', A9, F_SIM, True, 'center')
        put(r, c_play + 1, '—', A9, F_SIM, True, 'center')
        put(r, c_play + 2, 'LB end', A9, F_SIM, True, 'center')
    else:
        n, lvl, win = row[0], row[1], row[2]
        streak = streak + 1 if win else 0
        if win:
            jig += 7.4
            kite += KITE_STEP.get(streak, 1000)
            tad += TAD_STEP.get(streak, 100)
        if '__spend' in grants: inv['Coins'] -= 100
        if 'Core' in grants: inv['Coins'] += 15
        if 'Jigsaw' in grants: inv['Unlimited Lives'] += 30
        put(r, c_play, n, A9, F_SIM, True, 'center')
        put(r, c_play + 1, lvl, A9, F_SIM, True, 'center')
        put(r, c_play + 2, 'WIN' if win else 'LOSS', A9, F_SIM, True, 'center')
        put(r, c_play + 3, streak, A9, F_SIM, True, 'center')
        put(r, c_play + 4, (f'K×{KITE_STEP.get(streak,1000)} T×{TAD_STEP.get(streak,100)}') if win else '—',
            A9, F_SIM, True, 'center')
    if tag != 'S':
        put(r, c_ev0, round(jig, 1), A9, F_SIM, True, 'center')
        put(r, c_ev0 + 1, int(kite), A9, F_SIM, True, 'center')
        put(r, c_ev0 + 2, tad, A9, F_SIM, True, 'center')
    for j, k in enumerate(INV):
        put(r, c_inv0 + j, round(inv[k], 1), A9, F_INPUT if tag == 'S' else F_SIM, True, 'center')
    if tag == 'S':
        put(r, c_spend, 'opening inventory (input) →', A8G)
    else:
        put(r, c_spend, grants.get('__spend', '{}'), A9, F_HDR, True)
    for j, cat in enumerate(CATS):
        b = grants.get(cat, '')
        put(r, c_src0 + j, b if b else 0, A9, F_SIM, True)
    r += 1

# ---------- Session Summary ----------
r += 1
put(r, 2, 'Session Summary (sim — day totals per source × 11 sim resources; UL in minutes)', A9B)
r += 1
RES11 = ['HC', 'Slingshot', 'Shuffle', 'Comet', 'Red', 'Chuck', 'Bomb', 'UL Bomb', 'UL Chuck', 'UL Red', 'Unlimited Lives']
put(r, 2, 'Source', A9B, F_HDR, True)
for j, h in enumerate(RES11):
    put(r, 3 + j, h, A9B, F_HDR, True, 'center')
summ = [('Daily Gift', {'Slingshot': 1}), ('Core', {'HC': 15}), ('Jigsaw', {'Unlimited Lives': 30}),
        ('Kite Festival', {}), ('Target Day', {'Slingshot': 1}), ('Flash Race', {'HC': 50, 'Shuffle': 1}),
        ('Daily Night Sky Prize', {}),
        ('TOTAL', {'HC': 65, 'Slingshot': 2, 'Shuffle': 1, 'Unlimited Lives': 30})]
for i, (src, vals) in enumerate(summ):
    put(r + 1 + i, 2, src, A9B if src == 'TOTAL' else A9, F_HDR if src == 'TOTAL' else F_SIM, True)
    for j, res in enumerate(RES11):
        put(r + 1 + i, 3 + j, vals.get(res, 0), A9B if src == 'TOTAL' else A9,
            F_HDR if src == 'TOTAL' else F_SIM, True, 'center')
r += len(summ) + 1
put(r, 2, 'memo: Spend −100 HC (manual column — data_spend task was skipped; no sink telemetry)', A8G)
r += 2

# ---------- assumptions / flags ----------
notes = [
 'ASSUMPTIONS & FLAGS (the engine must carry these):',
 '1. N plays / win rate / streak persistence from data_streaks (attempts_per_day, win_rate_mean, p_continue_after_win). Sim CONDITIONS on the player being active.',
 '2. Progress states use final_balance_p25/p50/p75 selected by Luck (avg kept for reference). Milestones banked before today are excluded from the ledger.',
 '3. BOTH score events are win-streak-driven and share one streak engine: Kite score/win = 1/10/100/200/500/1000 by streak step (Ki_v2), Target Day = 1×mult 1/5/10/20/100 (TaD_v2).',
 '4. TaD calibration: persistence 0.786 → E[mult] ≈ 43 → ≈580 score/day modelled vs ≈357 measured (balance_p50 388 on ~2d × d1 share .92). Engine scales streak-model output to the measured percentile — model gives the SHAPE (when in the session score lands), measurement pins the LEVEL.',
 '5. Kite pays only ~top 25 of 60: rank p50 27 (10-19) → nothing. Only 40-99 (p50 6 → {SPT 140, Slingshot 1}) and 100+ (p50 2 → {Coins 200, SPT 300, Shuffle 1}) cash Kite at p50.',
 '6. LB placement at instance end = position percentile by Luck; Sampled mode jitters between percentiles with the Seed (deterministic per seed — no volatile RAND()).',
 '7. Kite m1 (@100 → 30min UL) banks on event day 1 — appears in a Day-3 ledger only as prior state.',
 '8. Spend column is MANUAL (data_spend skipped — sink attribution unreliable). OOM row is narrative: streak saves are WHY p_continue is 0.786.',
 '9. Source columns = engine CATEGORY_ORDER (25). Night Sky per NS ladder (night 7 of 13 → no payout) — NS bottom-up sim remains ON HOLD in the main engine.',
]
for i, n in enumerate(notes):
    put(r + i, 2, n, A9B if i == 0 else A8G)

# widths
ws.column_dimensions['A'].width = 2
ws.column_dimensions['B'].width = 24
for c in range(3, 8): ws.column_dimensions[CL(c)].width = 10
ws.column_dimensions['E'].width = 26
for c in ['H', 'I', 'J']: ws.column_dimensions[c].width = 10
ws.column_dimensions['G'].width = 12
for c in range(c_inv0, c_inv0 + 13): ws.column_dimensions[CL(c)].width = 8
ws.column_dimensions[CL(c_spend)].width = 34
for c in range(c_src0, c_src0 + 25): ws.column_dimensions[CL(c)].width = 18

wb.save(OUT)
print('written', OUT)
