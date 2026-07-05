# Builds EcoGainsSim_PlybyPly_v2.xlsx — LAYOUT PROPOSAL (mock) for the play-by-play session sim.
# All sim/data values are ILLUSTRATIVE, hand-computed from real workbook (5) data for the
# showcase scenario Day 5 (Sunday) / segment 10-19 / NONPAYER / cal_new. No engine is wired;
# this shows the organization: Config -> Player Profile -> Day Context -> Active Events ->
# play-by-play ledger (session start / plays / day-end LB payouts) -> Session Summary.
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter as CL

OUT = 'EcoGainsSim_PlybyPly_v2.xlsx'
F_INPUT, F_DATA, F_SIM, F_HDR = 'FFFFF2CC', 'FFCFE2F3', 'FFE2EFDA', 'FFD9D9D9'
RED = 'FFCC0000'
fill = lambda rgb: PatternFill('solid', fgColor=rgb)
thin = Side(style='thin')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
A9 = Font(name='Arial', size=9)
A9B = Font(name='Arial', size=9, bold=True)
A8G = Font(name='Arial', size=8, color='FF808080')

INV = ['Coins', 'SPT', 'SPT x2', 'Red', 'Chuck', 'Bomb', 'Slingshot', 'Shuffle', 'Comet',
       'Unlimited Lives', 'Unlimited Red', 'Unlimited Chuck', 'Unlimited Bomb']
# engine CATEGORY_ORDER (25, incl. Saga + Level Race — the user's sketch lacked both; flagged)
CATS = ['Ads', 'Bomb Challenge', "Bomb's Ballet", 'Chuck Challenge', 'Core', 'Daily Gift',
        'Daily Night Sky Prize', 'Flock Flurry', 'Hatchling Hideaway', 'Jigsaw', 'Kite Festival',
        'Level Race', 'Other', 'Photoshoot', 'Red Challenge', 'River Rush', 'Saga',
        'Season Pass (Free)', 'Target Day', 'Team Event', 'Team Race', 'Flash Race',
        'FlowerCoop', 'Rainbow Maker', 'IAPs']

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'EcoGainsSim_PlybyPly'
ws.sheet_view.showGridLines = False

def put(r, c, v, font=A9, fl=None, border=False, align=None, numfmt=None):
    cell = ws.cell(r, c, v)
    cell.font = font
    if fl: cell.fill = fill(fl)
    if border: cell.border = BORDER
    if align: cell.alignment = Alignment(horizontal=align)
    if numfmt: cell.number_format = numfmt
    return cell

def dv_cell(r, c, v, options):
    cell = put(r, c, v, A9B, F_INPUT, True, 'center')
    d = DataValidation(type='list', formula1='"' + ','.join(options) + '"',
                       allow_blank=False, showDropDown=False)
    ws.add_data_validation(d)
    d.add(cell)

# ---------- title + banner ----------
put(1, 2, 'EcoGainsSim_PlybyPly — one player, one day, play by play', Font(name='Arial', size=12, bold=True))
put(2, 2, 'MOCK v2 — LAYOUT PROPOSAL. Every green/blue value below is ILLUSTRATIVE (hand-derived '
          'from real data for Day 5 / 10-19 / NONPAYER / cal_new). Engine (ECOGAINS_PBP) not wired yet.',
    Font(name='Arial', size=9, bold=True, color=RED))

# ---------- Config Panel ----------
put(4, 2, 'Config Panel (input)', A9B)
cfg = [('Calendar', 'cal_new', ['cal_curr', 'cal_new']),
       ('Day to Simulate (1-33)', 5, None),
       ('Player Segment', '10-19', ['0-9', '10-19', '20-39', '40-99', '100+']),
       ('Payer', 'NONPAYER', ['NONPAYER', 'PAYER']),
       ('Mode', 'Sampled', ['Expected', 'Sampled']),
       ('Luck (LB placement / draws)', 'p50', ['p25', 'p50', 'p75']),
       ('Seed (Sampled mode)', 42, None),
       ('Levels Played (blank = auto)', '', None)]
for i, (label, val, opts) in enumerate(cfg):
    put(5 + i, 2, label, A9)
    if opts: dv_cell(5 + i, 3, val, opts)
    else: put(5 + i, 3, val, A9B, F_INPUT, True, 'center')

# ---------- Player Profile (data) ----------
put(4, 5, 'Player Profile (data — data_seg_beh, segment × payer)', A9B)
prof = [('Levels played / active day', 26.7, 'levels_played_per_active_day → N plays = 27'),
        ('Win rate', '53.4%', 'levels_completed / levels_played'),
        ('Saga completes / day', 13.7, 'saga_completes_per_active_day'),
        ('Minutes / active day', 55.5, 'minutes_per_active_day (context only)'),
        ('Login streak (p50)', 3, 'login_streak_p50 → Daily Gift cycle day 3'),
        ('NS max streak (p50)', 7, 'daily_max_streak_p50 → Night Sky streak night 7'),
        ('Daily Gift claim rate', '89.5%', 'daily_gift_claim_rate_pct'),
        ('P(active on this day)', '33.6%', 'sun_active_rate — context: this sim CONDITIONS on the player playing')]
for i, (label, val, srcnote) in enumerate(prof):
    put(5 + i, 5, label, A9)
    put(5 + i, 7, val, A9B, F_DATA, True, 'center')
    put(5 + i, 8, srcnote, A8G)

# ---------- Day Context (sim) ----------
put(4, 13, 'Day Context (sim)', A9B)
day = [('Day of week', 'Sunday (weekend)'), ('Weekend?', 'YES'),
       ('Active config set', 'cal_new → _v2 configs'),
       ('Events running today', '6 (see Active Events)')]
for i, (label, val) in enumerate(day):
    put(5 + i, 13, label, A9)
    put(5 + i, 15, val, A9B, F_SIM, True, 'center')

# ---------- Active Events ----------
r0 = 15
put(r0 - 1, 2, 'Active Events on this day (sim — cal_parsed × data_event_inst × accrual curves × config ladders)', A9B)
hdrs = ['Event', 'Family', 'Inst (days)', 'Event day', 'Ends today?', 'Opt-in', 'Progress @ session start',
        'Accrual today', 'Next milestone', 'End-of-day payout basis']
for j, h in enumerate(hdrs):
    put(r0, 2 + j, h, A9B, F_HDR, True, 'center')
ev = [
 ('Jigsaw', 'collection', '3–5', '3 / 3', 'YES', '95%', '249 tok (curve share d2/d3 × 369 measured final)',
  '+8.6 tok / win', 'm5 @ 370 → {Coins: 25}', 'milestones claimed during play (no LB)'),
 ('Kite Festival', 'score event', '3–5', '3 / 3', 'YES', 'n/a ⚠', '≈71% of final score (kite curve d2/d3)',
  '+score / win', 'score ladder (Ki_v2)', '⚠ Kite has NO data_event_inst rows — payout basis needs query ext'),
 ('Target Day', 'score + LB', '5', '1 / 1', 'YES', '12%', '0 (instance starts today)',
  '+1 × streak-mult / win (1/5/10/20/100)', 'milestones all-zero (not simmed)', 'LB rank p50 = 6 → {Slingshot: 1}'),
 ('Flash Race', 'leaderboard', '5', '1 / 1', 'YES', '69%', '0 (instance starts today)',
  'rank race (not token)', '—', 'LB rank p50 = 3 → {Coins: 50, Shuffle: 1}'),
 ('Night Sky', 'daily streak', '5', '1 / 1', 'YES', '19%', 'streak night 7',
  '+1 night if active', 'round 2 @ 13 nights → {Coins: 30, Red: 1}', 'NOT reached today (7 < 13) — no payout'),
 ('Daily Gift', 'always-on', '—', 'cycle day 3', '—', '89%', '—', '—', '—',
  'claimed at session start: {Slingshot: 1} (c_day_v2 V1 d3)'),
]
for i, row in enumerate(ev):
    for j, v in enumerate(row):
        put(r0 + 1 + i, 2 + j, v, A9, F_SIM, True)
put(r0 + 7, 2, 'excluded from the ledger (not play-driven / carried): Team Race, Season Pass, offer rows, Ads, Other, IAPs', A8G)

# ---------- Ledger ----------
L = r0 + 10   # zone-title row
put(L - 1, 2, 'Play-by-play ledger (sim — row S = session-start claims & opening inventory; row E = day-end LB payouts)', A9B)
c_play, c_ev0 = 3, 8            # C..G core, H..J event progress
c_inv0 = 11                     # K.. inventory (13)
c_spend = c_inv0 + 13 + 1       # skip one col
c_src0 = c_spend + 2            # source gains (25)
put(L, c_ev0, 'Event Progress (cum)', A9B)
put(L, c_inv0, 'Inventory (running)', A9B)
put(L, c_src0, 'Source Gains (bundle on the play it lands)', A9B)
cols = (['Play #', 'Level', 'Win?', 'Streak', 'TaD ×'] +
        ['Jigsaw tok', 'Kite score', 'TaD score'] + INV + ['', 'Spend', ''] + CATS)
for j, h in enumerate(cols):
    if h: put(L + 1, c_play + j, h, A9B, F_HDR, True, 'center')
put(L + 1, 2, 'Row', A9B, F_HDR, True, 'center')

inv = dict(zip(INV, [250, 120, 0, 2, 1, 0, 3, 2, 1, 0, 0, 0, 0]))
def inv_vals(): return [round(inv[k], 1) for k in INV]

# (win, spend-bundle, {source: bundle})
pattern = [
    ('S', None, None, {'Daily Gift': '{Slingshot: 1}'}),                       # session start
    (1, 100, True, {}), (2, 101, True, {}), (3, 102, False, {}),
    (4, 102, True, {}), (5, 103, True, {}), (6, 104, True, {}),
    (7, 105, False, {'__spend': '{Coins: 100} (OOM — streak save, failed)'}),
    (8, 105, False, {}),
    ('⋮', '… plays 9–25: 17 plays, 7 wins (best streak 4) …', None, {}),
    (26, 111, True, {'Core': '{Coins: 15} (chapter complete)'}),
    (27, 112, True, {'Jigsaw': '{Coins: 25} (m5 @ 370 crossed)'}),
    ('E', None, None, {'Flash Race': '{Coins: 50, Shuffle: 1} (rank 3)',
                       'Target Day': '{Slingshot: 1} (rank 6)',
                       'Kite Festival': '⚠ n/a — no rank/score-total data'}),  # day end
]
jig, kite, tad, streak, tadmult = 249.4, 1240, 0, 0, 0
MULT = {0: 0, 1: 1, 2: 5, 3: 10, 4: 20}
r = L + 2
for row in pattern:
    tag = row[0]
    put(r, 2, {'S': 'S — start', 'E': 'E — day end'}.get(tag, tag), A9B, F_HDR, True, 'center')
    if tag == '⋮':
        put(r, c_play, row[1], A8G)
        jig += 7 * 8.65; kite += 7 * 35; tad += 42; streak = 0
        r += 1; continue
    grants = row[3]
    if tag == 'S':
        if 'Daily Gift' in grants: inv['Slingshot'] += 1
        put(r, c_play, 0, A9, F_SIM, True, 'center')
        put(r, c_play + 1, '—', A9, F_SIM, True, 'center')
        put(r, c_play + 2, 'claims', A9, F_SIM, True, 'center')
    elif tag == 'E':
        inv['Coins'] += 50; inv['Shuffle'] += 1; inv['Slingshot'] += 1
        put(r, c_play, '—', A9, F_SIM, True, 'center')
        put(r, c_play + 1, '—', A9, F_SIM, True, 'center')
        put(r, c_play + 2, 'LB end', A9, F_SIM, True, 'center')
    else:
        n, lvl, win = row[0], row[1], row[2]
        streak = streak + 1 if win else 0
        m = MULT.get(min(streak, 4), 100)
        if win:
            jig += 8.65; kite += 35; tad += m
        if '__spend' in grants: inv['Coins'] -= 100
        if 'Core' in grants: inv['Coins'] += 15
        if 'Jigsaw' in grants: inv['Coins'] += 25
        put(r, c_play, n, A9, F_SIM, True, 'center')
        put(r, c_play + 1, lvl, A9, F_SIM, True, 'center')
        put(r, c_play + 2, 'WIN' if win else 'LOSS', A9, F_SIM, True, 'center')
        put(r, c_play + 3, streak, A9, F_SIM, True, 'center')
        put(r, c_play + 4, f'×{m}' if win else '—', A9, F_SIM, True, 'center')
    if tag != 'S':
        put(r, c_ev0, round(jig, 1), A9, F_SIM, True, 'center')
        put(r, c_ev0 + 1, int(kite), A9, F_SIM, True, 'center')
        put(r, c_ev0 + 2, tad, A9, F_SIM, True, 'center')
    for j, v in enumerate(inv_vals()):
        put(r, c_inv0 + j, v, A9, F_INPUT if tag == 'S' else F_SIM, True, 'center')
    if tag == 'S':
        put(r, c_spend, 'opening inventory (input) →', A8G)
    else:
        put(r, c_spend, grants.get('__spend', '{}'), A9, F_HDR, True)
    for j, cat in enumerate(CATS):
        b = grants.get(cat, '')
        put(r, c_src0 + j, b if b else 0, A9, F_SIM, True)
    r += 1

# ---------- Session Summary ----------
r += 1
put(r, 2, 'Session Summary (sim — day totals per source × 11 sim resources; SPT shown for info only)', A9B)
r += 1
RES11 = ['HC', 'Slingshot', 'Shuffle', 'Comet', 'Red', 'Chuck', 'Bomb', 'UL Bomb', 'UL Chuck', 'UL Red', 'Unlimited Lives']
put(r, 2, 'Source', A9B, F_HDR, True)
for j, h in enumerate(RES11):
    put(r, 3 + j, h, A9B, F_HDR, True, 'center')
summ = [('Daily Gift', {'Slingshot': 1}), ('Core', {'HC': 15}), ('Jigsaw', {'HC': 25}),
        ('Target Day', {'Slingshot': 1}), ('Flash Race', {'HC': 50, 'Shuffle': 1}),
        ('Kite Festival', None), ('Daily Night Sky Prize', {}), ('TOTAL', {'HC': 90, 'Slingshot': 2, 'Shuffle': 1})]
for i, (src, vals) in enumerate(summ):
    put(r + 1 + i, 2, src, A9B if src == 'TOTAL' else A9, F_HDR if src == 'TOTAL' else F_SIM, True)
    for j, res in enumerate(RES11):
        v = '⚠ n/a' if vals is None else vals.get(res, 0)
        put(r + 1 + i, 3 + j, v, A9B if src == 'TOTAL' else A9, F_HDR if src == 'TOTAL' else F_SIM, True, 'center')
r += len(summ) + 2

# ---------- assumptions / flags ----------
notes = [
 'ASSUMPTIONS & FLAGS (engine must carry these):',
 '1. N plays / win rate from data_seg_beh (levels_played, levels_completed per active day). Sim CONDITIONS on the player being active.',
 '2. Token accrual per play: measured final balance × accrual-curve day-share delta, spread over today\'s expected wins (uniform within day).',
 '3. Target Day score is simulated DIRECTLY from the win-streak multiplier ladder (1/5/10/20/100) — no accrual curve. ⚠ calibrate: naive model gives ~70/day vs measured ≈373/day (avg_final_token_balance 746 on ~2d) — real players sustain long streaks (OOM saves); tune streak persistence or multiplier mapping until day-score matches measured.',
 '4. LB placement at instance end = data_event_inst position percentile chosen by Luck; Sampled mode jitters between percentiles with the Seed (deterministic per seed — no volatile RAND).',
 '5. Milestones already banked before today are NOT in this ledger (they were paid on earlier days); only crossings during today\'s plays pay.',
 '6. ⚠ Kite Festival missing from data_event_inst (no score totals / placements) — needs a query extension before its payout can be computed.',
 '7. Spend column is MANUAL (no spend telemetry in this workbook) — illustrative OOM example only.',
 '8. Source columns = engine CATEGORY_ORDER (25; the original sketch lacked Saga and Level Race — added for consistency).',
 '9. Night Sky per NS config ladder (streak-night reach); NS bottom-up sim is ON HOLD in the main engine — same caution applies here.',
]
for i, n in enumerate(notes):
    put(r + i, 2, n, A9B if i == 0 else A8G)

# widths
ws.column_dimensions['A'].width = 2
ws.column_dimensions['B'].width = 24
for c in range(3, 8): ws.column_dimensions[CL(c)].width = 10
ws.column_dimensions['E'].width = 26
ws.column_dimensions['H'].width = 10
ws.column_dimensions['I'].width = 10
ws.column_dimensions['J'].width = 10
for c in range(c_inv0, c_inv0 + 13): ws.column_dimensions[CL(c)].width = 8
ws.column_dimensions[CL(c_spend)].width = 30
for c in range(c_src0, c_src0 + 25): ws.column_dimensions[CL(c)].width = 16

wb.save(OUT)
print('written', OUT)
