# Builds EcoGainsSim_PlybyPly_v4.xlsx — the PRODUCTION play-by-play sheet, wired to
# EcoGainsSim_PBP.gs (supersedes the value-mock _build_pbp_mock_v3.py). Three spill anchors:
#   ECOGAINS_PBP_PROFILE  (player profile block, 9 rows x 3 cols)
#   ECOGAINS_PBP_EVENTS   (active events table, header + variable rows x 9 cols)
#   ECOGAINS_PBP          (the ledger — brings its own 46-col header; up to ~160 play rows)
# Spill regions are left EMPTY (any content there = #REF! spill blockage in Sheets).
# Import into the workbook replacing the EcoGainsSim_PlybyPly sheet; requires EcoGainsSim_v4.gs
# + EcoGainsSim_PBP.gs in the Apps Script project. Formulas show #NAME? in Excel — expected.
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter as CL

OUT = 'EcoGainsSim_PlybyPly_v4.xlsx'
F_INPUT, F_DATA, F_SIM, F_HDR = 'FFFFF2CC', 'FFCFE2F3', 'FFE2EFDA', 'FFD9D9D9'
fill = lambda rgb: PatternFill('solid', fgColor=rgb)
thin = Side(style='thin')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
A9 = Font(name='Arial', size=9)
A9B = Font(name='Arial', size=9, bold=True)
A8G = Font(name='Arial', size=8, color='FF808080')

RES = ['HC', 'Slingshot', 'Shuffle', 'Comet', 'Red', 'Chuck', 'Bomb',
       'UL Bomb', 'UL Chuck', 'UL Red', 'Unlimited Lives']

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'EcoGainsSim_PlybyPly'
ws.sheet_view.showGridLines = False

def put(r, c, v, font=A9, fl=None, border=False, align=None):
    cell = ws.cell(r, c, v)
    cell.font = font
    if fl: cell.fill = fill(fl)
    if border: cell.border = BORDER
    if align: cell.alignment = Alignment(horizontal=align)
    return cell

def dv_cell(r, c, v, options):
    cell = put(r, c, v, A9B, F_INPUT, True, 'center')
    d = DataValidation(type='list', formula1='"' + ','.join(options) + '"',
                       allow_blank=True, showDropDown=False)
    ws.add_data_validation(d)
    d.add(cell)

# ---------- title ----------
put(1, 2, 'EcoGainsSim_PlybyPly — one player, one day, play by play',
    Font(name='Arial', size=12, bold=True))
put(2, 2, 'Live sheet — requires EcoGainsSim_v4.gs + EcoGainsSim_PBP.gs in the Apps Script '
          'project (formulas show #NAME? until installed). Model spec: SIMULATION_METHODOLOGY.md §14.',
    A8G)

# ---------- Config Panel (inputs; the three anchors read these cells) ----------
put(4, 2, 'Config Panel (input)', A9B)
cfg = [('Calendar',                      'cal_new',  ['cal_curr', 'cal_new']),
       ('Day to Simulate (1-33)',        5,          None),
       ('Player Segment',                '10-19',    ['0-9', '10-19', '20-39', '40-99', '100+']),
       ('Payer',                         'NONPAYER', ['NONPAYER', 'PAYER']),
       ('Mode',                          'Sampled',  ['Expected', 'Sampled']),
       ('Luck (progress + placement)',   'p50',      ['p25', 'p50', 'p75']),
       ('Seed (Sampled mode)',           42,         None),
       ('Levels Played (blank = auto)',  None,       None)]
for i, (label, val, opts) in enumerate(cfg):
    put(5 + i, 2, label, A9)
    if opts: dv_cell(5 + i, 3, val, opts)
    else: put(5 + i, 3, val, A9B, F_INPUT, True, 'center')

# ---------- Player Profile (spill anchor; data) ----------
put(4, 5, 'Player Profile (data — data_streaks + data_seg_beh; spills 9×3)', A9B)
put(5, 5, '=ECOGAINS_PBP_PROFILE($C$7,$C$8)', A9, F_DATA)
# rows 5-13 x cols E-G reserved for the spill — leave empty

# ---------- Opening inventory (input, optional; read by the ledger) ----------
put(15, 2, 'Opening inventory (input — optional; resource order below)', A9B)
for j, r in enumerate(RES):
    put(16, 3 + j, r, A9B, F_HDR, True, 'center')
    put(17, 3 + j, 0, A9B, F_INPUT, True, 'center')

# ---------- Active Events (spill anchor; sim) ----------
put(19, 2, 'Active Events on this day (sim — spills header + one row per running event '
           'and per session-start claim)', A9B)
put(20, 2, '=ECOGAINS_PBP_EVENTS($C$5,$C$6,$C$7,$C$8,$C$10)', A9, F_SIM)
# rows 20-32 x cols B-J reserved — leave empty

# ---------- Assumptions / flags (static — kept ABOVE the ledger: its spill length varies) ----
notes = [
 'ASSUMPTIONS & FLAGS (model spec: SIMULATION_METHODOLOGY.md §14):',
 '1. N plays / win rate / streak persistence from data_streaks; the sim CONDITIONS on the player being active and participating in every running event.',
 '2. Progress @ session start = final_balance_p25/50/75 (Luck) × accrual-curve share(k−1); milestones banked on earlier days never appear in this ledger.',
 '3. Score events are streak-driven off their config ladders (Ki_v2 steps / TaD_v2 multipliers), then scaled so the day total hits the measured target — model = shape, measurement = level.',
 '4. LB payouts land on row E only for instances ending today, at the Luck percentile of position_pXX (Sampled: seeded ±0.25-quantile jitter; deterministic per Seed).',
 '5. Daily Gift = login-streak-p50 cycle day, equal-weight average of config variant ladders; Night Sky pays only on the exact milestone night (daily_max_streak_p50).',
 "6. Saga pays at NODE boundaries (config ladder: 10 levels per node, alternating 10/0 HC, node 10 = 25) walked from node 1 — the player's true cycle position is unknown. Core chapter chests NOT simulated. Flock Flurry opt-in = 60 min UL (design-PDF constant).",
 '7. HH/Ph milestone requirements read from the EventReach helper columns on HH_v2 (AV5:AV9) / Ph_v2 (AU5:AU34) — do not remove them.',
 '8. SPT / COOP / Avatar / Dly rewards are outside the 11-resource universe and are not tracked.',
]
for i, n in enumerate(notes):
    put(34 + i, 2, n, A9B if i == 0 else A8G)

# ---------- Ledger (spill anchor; sim) ----------
put(45, 2, 'Play-by-play ledger (sim — spills its own header; S = session-start claims, '
           'E = day-end LB payouts, then Session Summary per source × 11 resources)', A9B)
put(46, 2, '=ECOGAINS_PBP($C$5,$C$6,$C$7,$C$8,$C$9,$C$10,$C$11,$C$12,$C$17:$M$17)', A9, F_SIM)
# rows 46-250 x 46 cols reserved for the spill — MUST stay empty

# ---------- widths (ledger spills B..W: Row|Play|Level|Win?|Streak|Mult|Claims|4 slots|11 res) --
ws.column_dimensions['A'].width = 2
ws.column_dimensions['B'].width = 26
ws.column_dimensions['C'].width = 10   # Play # / config values / inventory inputs
ws.column_dimensions['D'].width = 8    # Level
ws.column_dimensions['E'].width = 26   # Win? / profile stat labels
ws.column_dimensions['F'].width = 8    # Streak / profile values
ws.column_dimensions['G'].width = 34   # Mult / profile source notes
ws.column_dimensions['H'].width = 64   # Claims — the play-by-play reward narration
for c in range(9, 13):                 # I-L: event-progress slots
    ws.column_dimensions[CL(c)].width = 13
for c in range(13, 24):                # M-W: 11 resource columns (running inventory)
    ws.column_dimensions[CL(c)].width = 9

wb.save(OUT)
print('written', OUT)
