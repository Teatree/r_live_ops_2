# Builds EcoGainsSim_PlybyPly_v5.xlsx — the production play-by-play sheet in TaD_v2 STYLE
# (black section bars w/ white bold text, gold FFD966 config labels + FFF2CC values, F7CB4D
# table headers, thin borders, Arial 11, everything anchored at column A, no merges).
# Supersedes _build_pbp_v4.py. Wired to EcoGainsSim_PBP.gs (v2 signature with startLevel):
#   A18: ECOGAINS_PBP_PROFILE   A29: ECOGAINS_PBP_EVENTS   A55: ECOGAINS_PBP
# Spill regions are left EMPTY. Formulas show #NAME? in Excel — expected (Sheets + script only).
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter as CL

OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'display',
                   'EcoGainsSim_PlybyPly_v5.xlsx')
F_SECTION, F_LABEL, F_VALUE, F_TBLHDR = 'FF000000', 'FFFFD966', 'FFFFF2CC', 'FFF7CB4D'
fill = lambda rgb: PatternFill('solid', fgColor=rgb)
thin = Side(style='thin')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
A11 = Font(name='Arial', size=11)
A11B = Font(name='Arial', size=11, bold=True)
A11W = Font(name='Arial', size=11, bold=True, color='FFFFFFFF')
A9G = Font(name='Arial', size=9, color='FF808080')

RES = ['HC', 'Slingshot', 'Shuffle', 'Comet', 'Red', 'Chuck', 'Bomb',
       'UL Bomb', 'UL Chuck', 'UL Red', 'Unlimited Lives']

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'EcoGainsSim_PlybyPly'
ws.sheet_view.showGridLines = False

def put(r, c, v, font=A11, fl=None, border=False, align=None):
    cell = ws.cell(r, c, v)
    cell.font = font
    if fl: cell.fill = fill(fl)
    if border: cell.border = BORDER
    if align: cell.alignment = Alignment(horizontal=align)
    return cell

def section(r, text):
    put(r, 1, text, A11W, F_SECTION, False, 'left')

def dv_cell(r, c, v, options):
    cell = put(r, c, v, A11, F_VALUE, False, 'center')
    d = DataValidation(type='list', formula1='"' + ','.join(options) + '"',
                       allow_blank=True, showDropDown=False)
    ws.add_data_validation(d)
    d.add(cell)

# ---------- title + config panel (TaD_v2 pattern: A2 black bar, gold labels, yellow values) ----
put(1, 1, 'EcoGainsSim PlybyPly', Font(name='Arial', size=14, bold=True))
section(2, 'Config Panel')
cfg = [('Calendar',                                 'cal_new',  ['cal_curr', 'cal_new']),
       ('Day to Simulate (1-33)',                   5,          None),
       ('Player Segment',                           '10-19',    ['0-9', '10-19', '20-39', '40-99', '100+']),
       ('Payer',                                    'NONPAYER', ['NONPAYER', 'PAYER']),
       ('Mode',                                     'Sampled',  ['Expected', 'Sampled']),
       ('Luck (progress + placement)',              'p50',      ['p25', 'p50', 'p75']),
       ('Seed (Sampled mode)',                      42,         None),
       ('Levels Played (blank = auto)',             None,       None),
       ('Starting Saga Level (blank = rnd 100-400)', None,      None)]
for i, (label, val, opts) in enumerate(cfg):
    put(3 + i, 1, label, A11B, F_LABEL, False, 'left')
    if opts: dv_cell(3 + i, 2, val, opts)
    else: put(3 + i, 2, val, A11, F_VALUE, False, 'center')

# ---------- opening inventory ----------
section(13, 'Opening Inventory (optional — read by the ledger; 0 = start empty)')
for j, r in enumerate(RES):
    put(14, 1 + j, r, A11B, F_TBLHDR, True, 'center')
    put(15, 1 + j, 0, A11, F_VALUE, True, 'center')

# ---------- player profile (spill) ----------
section(17, 'Player Profile (data — data_streaks + data_seg_beh; spills 9 rows × 3 cols)')
put(18, 1, '=ECOGAINS_PBP_PROFILE($B$5,$B$6)', A11)
# rows 18-26 × cols A-C reserved

# ---------- active events (spill) ----------
section(28, 'Active Events on This Day (sim — spills one row per running event + per session-start claim)')
put(29, 1, '=ECOGAINS_PBP_EVENTS($B$3,$B$4,$B$5,$B$6,$B$8)', A11)
# rows 29-41 × cols A-I reserved

# ---------- assumptions ----------
section(43, 'Assumptions & Flags (model spec: SIMULATION_METHODOLOGY.md §14)')
notes = [
 '1. N plays / win rate / streak persistence from data_streaks; the sim CONDITIONS on the player being active and participating in every running event.',
 '2. The player is NOT fresh: they walk in at Starting Saga Level (input, or seeded random 100-400); the Level column is the absolute saga level.',
 '3. Event progress @ session start = final_balance_p25/50/75 (Luck) × accrual-curve share(k−1); milestones banked on earlier days never appear in this ledger.',
 '4. Score events are streak-driven off their config ladders (Ki_v2 steps / TaD_v2 multipliers), scaled so the day total hits the measured target — model = shape, measurement = level.',
 '5. LB payouts land on row E only for instances ending today, at the Luck percentile of position_pXX (Sampled: seeded ±0.25-quantile jitter; deterministic per Seed).',
 '6. Saga pays at config node boundaries anchored to the ABSOLUTE level (10-level nodes cycling every 100: level 240 = node 4). Core chapter chests NOT simulated.',
 '7. Daily Gift = login-streak-p50 cycle day, equal-weight average of config variant ladders; Night Sky pays only on the exact milestone night (daily_max_streak_p50).',
 '8. Flock Flurry opt-in = 60 min UL (design-PDF constant). HH/Ph milestone reqs read from the EventReach helper columns on HH_v2 / Ph_v2 — do not remove them.',
 '9. SPT / COOP / Avatar / Dly rewards are outside the 11-resource universe and are not tracked.',
]
for i, n in enumerate(notes):
    put(44 + i, 1, n, A9G)

# ---------- ledger (spill) ----------
section(54, 'Play-by-Play Ledger (sim — spills its own header; S = session start & claims, '
            'E = day-end LB payouts, then Session Summary per source × 11 resources)')
put(55, 1, '=ECOGAINS_PBP($B$3,$B$4,$B$5,$B$6,$B$7,$B$8,$B$9,$B$10,$B$11,$A$15:$K$15)', A11)
# rows 55-260 × cols A-V reserved

# ---------- widths (ledger spills A..V: Row|Play|Level|Win?|Streak|Mult|Claims|4 slots|11 res) --
ws.column_dimensions['A'].width = 30   # labels / Row tags / event names
ws.column_dimensions['B'].width = 14   # config values / Play # / profile values
ws.column_dimensions['C'].width = 12   # Level / profile source (overflows right, cells empty)
ws.column_dimensions['D'].width = 14   # Win?
ws.column_dimensions['E'].width = 10   # Streak
ws.column_dimensions['F'].width = 14   # Mult
ws.column_dimensions['G'].width = 64   # Claims — the play-by-play reward narration
for c in range(8, 12):                 # H-K: event-progress slots
    ws.column_dimensions[CL(c)].width = 13
for c in range(12, 23):                # L-V: 11 resource columns (running inventory)
    ws.column_dimensions[CL(c)].width = 9

wb.save(OUT)
print('written', OUT)
