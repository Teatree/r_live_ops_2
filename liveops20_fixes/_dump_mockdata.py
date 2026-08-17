# Dump the generated variant-basis workbook to JSON for the offline harness.
#
# Same shape as harness/_mockdata.json in the main stack: {sheet: {values: [[...]], merges: [...]}}
# so the node harnesses can mock SpreadsheetApp against it. Merges matter — the calendar reader
# treats one merged range as one instance and its width as the duration.
#
# The generated workbook ships with EMPTY data_* sheets (the analytics LLM pushes those). For a
# structural smoke test that is not trivially all-zero, --borrow-data copies the data_* sheets from
# the newest NEW_LIVEOPS_CALENDAR_ECO. Those numbers are the pre-test snapshot, NOT variant data —
# they are only there to prove the plumbing computes.
#
# Usage: python liveops20_fixes/_dump_mockdata.py [--borrow-data]
import argparse
import glob
import json
import os
import re

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.join(HERE, '..')
BUILT = os.path.join(ROOT, 'display', 'LiveOps20_variant_basis.xlsx')
OUT = os.path.join(HERE, 'harness', '_mockdata_variant.json')
DATA_PREFIX = 'data_'


def newest_source():
    best, bn = None, -1
    for p in glob.glob(os.path.join(ROOT, 'workbooks', 'NEW_LIVEOPS_CALENDAR_ECO*.xlsx')):
        m = re.search(r'\((\d+)\)', os.path.basename(p))
        n = int(m.group(1)) if m else 0
        if n > bn:
            best, bn = p, n
    return best


def cell(v):
    """JSON-safe: dates become ISO strings (the calendar header rows carry real dates, and the
    sim_refresh nonce is a timestamp). The engine only ever reads these as text."""
    if v is None:
        return ''
    if hasattr(v, 'isoformat'):
        return v.isoformat()
    return v


def dump_sheet(sh):
    values = []
    for row in sh.iter_rows(values_only=True):
        values.append([cell(v) for v in row])
    while values and all(v == '' for v in values[-1]):
        values.pop()
    merges = [{'r': r.min_row, 'c': r.min_col,
               'nr': r.max_row - r.min_row + 1, 'nc': r.max_col - r.min_col + 1}
              for r in sh.merged_cells.ranges]
    return {'values': values, 'merges': merges}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--borrow-data', action='store_true',
                    help='fill data_* sheets from the newest calendar workbook (smoke test only)')
    ap.add_argument('--workbook', default=BUILT,
                    help='source workbook (default: the generated one; use the _2 copy when the '
                         'original is locked by Excel)')
    args = ap.parse_args()

    if not os.path.exists(args.workbook):
        raise SystemExit(f'{args.workbook} not found — run liveops20_fixes/_build_workbook.py first')
    wb = openpyxl.load_workbook(args.workbook, data_only=True)
    out = {}
    for name in wb.sheetnames:
        out[name] = dump_sheet(wb[name])

    borrowed = []
    if args.borrow_data:
        src = newest_source()
        wbs = openpyxl.load_workbook(src, data_only=True)
        for name in wbs.sheetnames:
            if name.startswith(DATA_PREFIX) and name in out:
                out[name] = dump_sheet(wbs[name])
                borrowed.append(name)

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, 'w', encoding='utf-8') as f:
        json.dump(out, f)
    empties = [n for n, s in out.items()
               if n.startswith(DATA_PREFIX) and len(s['values']) <= 1]
    print(f'written {OUT}')
    print(f'  sheets: {len(out)}')
    print(f'  _v2 proposal sheets present: {len([n for n in out if n.endswith("_v2")])}')
    print(f'  data_* still empty: {sorted(empties) if empties else "none"}')
    if borrowed:
        print(f'  BORROWED (pre-test numbers, not variant): {", ".join(sorted(borrowed))}')


if __name__ == '__main__':
    main()
