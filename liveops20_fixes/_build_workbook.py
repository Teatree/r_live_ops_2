# Build the variant-basis workbook for the LiveOps 2.0 fixes stack.
#
# STRATEGY: clone the highest-numbered NEW_LIVEOPS_CALENDAR_ECO and transform it, rather than
# authoring 60 sheets from scratch. Every config sheet in that workbook is hand-authored (ladders,
# panels, conditional formatting, the Ph reach helper columns); regenerating them from code would
# throw that away and introduce a second source of truth. So the builder only makes the changes the
# re-base actually needs, and reports exactly what it touched.
#
# Transformations
#   1. reset the PROPOSAL layer: every *_v2 config sheet is overwritten with a clone of its BASE
#      sheet, so a fresh workbook is neutral — R = 1 => diff 0 until you edit a _v2 sheet. (Until
#      2026-08-18 this added a third *_v3 layer instead; that was removed because it made editing
#      a _v2 sheet silently do nothing. base = the config the variant ran, _v2 = your proposal.)
#   2. trim both calendars to the 21-day window and rewrite the day/date header rows
#   3. write cal_curr as the AS-RUN variant schedule (table below — edit it, it is the one place
#      the as-run schedule is asserted) and keep cal_new as the proposal calendar
#   4. clear every data_* sheet to headers only — the analytics LLM pushes the variant data in
#   5. re-point the display sheets at 21 days (EcoGainsSim_Daily's day rows; the ECOGAINS_* formulas
#      are untouched because they read the engine, not a range)
#
# Output: display/LiveOps20_variant_basis.xlsx  (import as a NEW Google workbook, then paste the
# four .gs files from liveops20_fixes/engine/ into its Apps Script project)
#
# Usage: python liveops20_fixes/_build_workbook.py [--out <path>]
import argparse
import glob
import os
import re
import shutil
from copy import copy

import openpyxl
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.join(HERE, '..')
WB_DIR = os.path.join(ROOT, 'workbooks')
OUT_DIR = os.path.join(ROOT, 'display')

SIM_DAYS = 21                      # must match SIM_DAYS in liveops20_fixes/engine/EcoGainsSim_v4.gs
DAY_ONE = '2026-07-27'             # Monday; day 21 = 2026-08-16
CAL_FIRST_ROW, CAL_LAST_ROW = 5, 25
CAL_FIRST_COL = 2                  # day = column - 1

# config sheets that come in base/_v2 pairs. base = the anchor (what the variant ran), _v2 = the
# proposal you author. Must match CONFIG_PAIRED in liveops20_fixes/engine/EcoGainsSim_v4.gs.
PAIRED = ['c_saga', 'c_day', 'NS', 'SP', 'SP_lb', 'Race', 'TaD', 'Ki', 'J', 'HH', 'BB', 'Ph',
          'RR', 'F']
# bottom-up sheets (one ladder, no base/_v2 pair) — nothing to reset
BOTTOM_UP = ['RM_1st', 'RM_2nd', 'RM']
# pack/collection sheets that exist in the workbook but not in the 13-resource model
LEAVE_ALONE = ['PackConfig', 'AlbumConfig', 'CardPoolConfig', 'EcoPackGains', 'SimOutput',
               'SimSummary', 'PlayerBehavior', 'oldEcoGainsSim_PlybyPly', 'LOCal NS_v2', 'TE']

DATA_SHEETS = ['data_gains', 'data_seg_beh', 'data_event_inst', 'data_event_accrual',
               'data_event_kite_accrual', 'data_RM', 'data_streaks', 'data_econ',
               'data_econ_daily']
# the three new sheets the prompt asks for (created empty with their agreed headers)
NEW_DATA_SHEETS = {
    'data_core_spt': ['segment', 'payer_flag', 'difficulty_tier', 'player_days',
                      'levels_completed_total', 'levels_completed_per_active_day',
                      'spt_from_level_completes_total', 'spt_per_level_completed',
                      'spt_per_active_player_day'],
    'data_spend_action': ['segment', 'payer_flag', 'resource', 'action', 'spend_context',
                          'spend_events', 'spender_days', 'amount_spent', 'amount_spent_free',
                          'amount_spent_paid', 'spend_per_event', 'amount_per_active_player_day',
                          'pct_of_resource_spent'],
    'data_ns_rounds': ['segment', 'payer_flag', 'round', 'cum_streak_req', 'claims_total',
                       'player_days', 'claim_rate_per_player_day', 'players_finished_pct',
                       'hc_granted_total', 'hc_per_claim'],
}

# ---------------------------------------------------------------------------------------------
# AS-RUN VARIANT SCHEDULE for cal_curr — the anchor calendar.
# day 1 = 2026-07-27 (Monday). Each entry: (calendar label, [(start_day, duration), ...]).
# Seeded from the v2 plan mapped onto this window, with the two divergences the sim-vs-actual work
# established. EDIT THIS — it is the single place the as-run schedule is asserted, and the T term
# for every event depends on it.
AS_RUN_SCHEDULE = [
    # always-on / daily
    ('Night Sky',              [(d, 1) for d in range(1, SIM_DAYS + 1)]),
    ('Season Pass',            [(1, SIM_DAYS)]),
    # weekly cadence lanes (Mon-start window: day 1, 8, 15 are Mondays)
    ('Team Event',             [(1, 1), (8, 1), (15, 1)]),
    ('Top Teams Leaderboard',  [(1, 1), (8, 1), (15, 1)]),
    ('Top Players Leaderboard', [(1, 1), (8, 1), (15, 1)]),
    ('Team Race',              [(3, 1), (10, 1), (17, 1)]),
    ('Rainbow Maker',          [(1, 4), (8, 4), (15, 4)]),
    ('Hatchling Hideaway',     [(1, 4), (8, 4), (15, 4)]),
    ('Flock Flurry',           [(2, 2), (6, 4), (13, 4), (20, 2)]),
    ('Chuck\'s Flash Race',    [(3, 3), (10, 3), (17, 3)]),
    ('Target Day',             [(3, 3), (10, 3), (17, 3)]),
    ('Kite Festival',          [(3, 8), (17, 5)]),
    # leaderboard events, as run
    ('Chuck\'s Challenge',     [(6, 3)]),
    ('Bomb\'s Challenge',      [(13, 3)]),
    ('Red\'s Challenge',       [(1, 1), (20, 2)]),
    ('Level Race',             [(20, 2)]),
    # collections — the two known divergences from the plan. They share ONE grid row, as they do in
    # the live workbook: they never overlap, and the grid is only 21 rows deep.
    ('Jigsaw Puzzle',          [(12, 4)], 'collections'),   # ran ~Aug 7-10 (plan: days 17-19)
    ('Bomb\'s Ballet Show',    [(4, 3)], 'collections'),    # ended ~Aug 2 (plan: days 10-12)
    ('Photoshoot',             [(17, 3)], 'collections'),
    # offers (not simulated, kept for parity with the current workbook)
    ('Season Offers',          [(1, 1), (8, 1), (15, 1)]),
    ('Progress Pack',          [(3, 1), (10, 1), (17, 1)]),
    ('Special Offer',          [(4, 1), (11, 1), (18, 1)]),
]

DOW = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']


def newest_workbook():
    best, bn = None, -1
    for p in glob.glob(os.path.join(WB_DIR, 'NEW_LIVEOPS_CALENDAR_ECO*.xlsx')):
        m = re.search(r'\((\d+)\)', os.path.basename(p))
        n = int(m.group(1)) if m else 0
        if n > bn:
            best, bn = p, n
    return best, bn


def clone_sheet(wb, src_name, dst_name):
    """openpyxl's copy_worksheet keeps values, formulas, styles and merges within one workbook."""
    if dst_name in wb.sheetnames:
        del wb[dst_name]
    dst = wb.copy_worksheet(wb[src_name])
    dst.title = dst_name
    return dst


def day_date(day):
    import datetime
    return (datetime.date.fromisoformat(DAY_ONE) + datetime.timedelta(days=day - 1))


def write_calendar(sh, schedule, label):
    """Clear the grid, write the day/date header rows, then lay the schedule out one event per row.
    Merged range = one instance, width = duration (the engine's calendar rule)."""
    last_col = CAL_FIRST_COL + SIM_DAYS - 1
    # unmerge everything inside the grid, then clear it
    for rng in list(sh.merged_cells.ranges):
        if rng.min_row >= CAL_FIRST_ROW - 3 and rng.max_row <= CAL_LAST_ROW:
            sh.unmerge_cells(str(rng))
    for r in range(CAL_FIRST_ROW - 3, CAL_LAST_ROW + 1):
        for c in range(1, max(sh.max_column, 40) + 1):
            sh.cell(row=r, column=c).value = None
    # header rows: day index, weekday, date (rows 2,3,4 above the grid)
    sh.cell(row=1, column=1).value = f'{label} — {SIM_DAYS}-day A/B window, day 1 = {DAY_ONE}'
    for day in range(1, SIM_DAYS + 1):
        c = day + 1
        d = day_date(day)
        sh.cell(row=CAL_FIRST_ROW - 3, column=c).value = day
        sh.cell(row=CAL_FIRST_ROW - 2, column=c).value = DOW[d.weekday()]
        sh.cell(row=CAL_FIRST_ROW - 1, column=c).value = d.isoformat()
    # events by grid row. An entry may carry a third element naming a row GROUP: events in the same
    # group share one row (the live workbook does this for the collections). Overlap inside a group
    # would put two events in one cell, so it fails loudly instead.
    groups, order = {}, []
    for entry in schedule:
        name, instances = entry[0], entry[1]
        g = entry[2] if len(entry) > 2 else name
        if g not in groups:
            groups[g] = []
            order.append(g)
        groups[g].append((name, instances))
    if len(order) > CAL_LAST_ROW - CAL_FIRST_ROW + 1:
        raise SystemExit(f'{label}: {len(order)} lanes into '
                         f'{CAL_LAST_ROW - CAL_FIRST_ROW + 1} grid rows — group some, or widen '
                         f'CAL_LAST_ROW here AND in the engine')
    written = []
    for i, g in enumerate(order):
        row = CAL_FIRST_ROW + i
        occupied = {}
        for name, instances in groups[g]:
            for (start, dur) in instances:
                if start < 1 or start + dur - 1 > SIM_DAYS:
                    continue                                # clipped out of the window
                for d in range(start, start + dur):
                    if d in occupied:
                        raise SystemExit(f'{label}: {name} overlaps {occupied[d]} on day {d} in '
                                         f'shared row group "{g}" — give one of them its own row')
                    occupied[d] = name
                c0 = start + 1
                sh.cell(row=row, column=c0).value = name
                if dur > 1:
                    sh.merge_cells(start_row=row, start_column=c0, end_row=row,
                                   end_column=c0 + dur - 1)
            written.append(name)
    return written, last_col


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--out', default=os.path.join(OUT_DIR, 'LiveOps20_variant_basis.xlsx'))
    args = ap.parse_args()

    src, n = newest_workbook()
    print(f'source workbook: {os.path.basename(src)}')
    os.makedirs(OUT_DIR, exist_ok=True)
    tmp = args.out + '.tmp.xlsx'
    shutil.copyfile(src, tmp)
    # keep_vba off, formulas preserved (data_only=False is the default)
    wb = openpyxl.load_workbook(tmp)
    report = {'v2_reset': [], 'data_cleared': [], 'data_created': [], 'calendars': {}}

    # ---- 1. proposal layer ------------------------------------------------------------------
    # Reset each _v2 to its base sheet. The source workbook's _v2 sheets carry the OLD v2-vs-Control
    # redesign; carrying those over would mean a brand-new workbook already reports a diff nobody
    # authored in it. Bottom-up sheets have no pair and are left as they are.
    for logical in PAIRED:
        prop = f'{logical}_v2'
        if logical not in wb.sheetnames:
            print(f'  ! {logical} missing — cannot reset {prop}')
            continue
        clone_sheet(wb, logical, prop)
        report['v2_reset'].append(prop)

    # ---- 2 + 3. calendars -------------------------------------------------------------------
    if 'cal_curr' in wb.sheetnames:
        names, last_col = write_calendar(wb['cal_curr'], AS_RUN_SCHEDULE, 'cal_curr (as-run variant)')
        report['calendars']['cal_curr'] = f'{len(names)} lanes, cols B..{get_column_letter(last_col)}'
    if 'cal_new' in wb.sheetnames:
        # the proposal calendar starts as a copy of the as-run one: a fresh workbook must be neutral
        names, last_col = write_calendar(wb['cal_new'], AS_RUN_SCHEDULE, 'cal_new (proposal)')
        report['calendars']['cal_new'] = f'{len(names)} lanes, cols B..{get_column_letter(last_col)}'
    if 'cal_parsed' in wb.sheetnames:
        del wb['cal_parsed']            # stale precompute; re-run the menu item in the new workbook

    # ---- 4. data sheets ---------------------------------------------------------------------
    for name in DATA_SHEETS:
        if name not in wb.sheetnames:
            continue
        sh = wb[name]
        hdr = [c.value for c in sh[1]]
        while hdr and hdr[-1] is None:
            hdr.pop()
        max_row, max_col = sh.max_row, sh.max_column
        for r in range(2, max_row + 1):
            for c in range(1, max_col + 1):
                sh.cell(row=r, column=c).value = None
        report['data_cleared'].append(f'{name} ({len(hdr)} cols, {max_row - 1} rows cleared)')
    for name, hdr in NEW_DATA_SHEETS.items():
        if name in wb.sheetnames:
            del wb[name]
        sh = wb.create_sheet(name)
        for i, h in enumerate(hdr, start=1):
            sh.cell(row=1, column=i).value = h
        report['data_created'].append(f'{name} ({len(hdr)} cols)')

    # ---- 5. display sheets ------------------------------------------------------------------
    # EcoGainsSim_Daily lists one row per day; drop the rows past the window so the sheet does not
    # show 12 empty days. The ECOGAINS_* anchors themselves are untouched.
    if 'EcoGainsSim_Daily' in wb.sheetnames:
        sh = wb['EcoGainsSim_Daily']
        day_col, first_day_row = None, None
        for r in range(1, min(sh.max_row, 20) + 1):
            for c in range(1, min(sh.max_column, 6) + 1):
                if str(sh.cell(row=r, column=c).value or '').strip() == 'Day':
                    day_col, first_day_row = c, r + 1
                    break
            if day_col:
                break
        if day_col:
            trimmed = 0
            for r in range(first_day_row, sh.max_row + 1):
                v = sh.cell(row=r, column=day_col).value
                if isinstance(v, (int, float)) and v > SIM_DAYS:
                    for c in range(1, sh.max_column + 1):
                        sh.cell(row=r, column=c).value = None
                    trimmed += 1
            report['daily_rows_trimmed'] = trimmed

    # a one-page note so nobody wonders which workbook this is
    if 'contents' in wb.sheetnames:
        sh = wb['contents']
        sh.cell(row=1, column=1).value = 'LiveOps 2.0 fixes — variant-basis simulation'
        sh.cell(row=2, column=1).value = (
            f'Anchor = the as-run variant (base sheets). Proposals go in *_v2. Window = {SIM_DAYS} '
            f'days from {DAY_ONE}. Paste the four .gs files from liveops20_fixes/engine/ into this '
            f'workbook\'s Apps Script project, then run EcoGainsSim > Precompute calendars. '
            f'data_* sheets are empty on purpose — they are pushed by the variant-only export '
            f'(see liveops20_fixes/PROMPT_variant_data_request.md).')

    wb.save(args.out)
    os.remove(tmp)

    # ---- verification ------------------------------------------------------------------------
    chk = openpyxl.load_workbook(args.out)
    problems = []
    for logical in PAIRED:
        if logical in chk.sheetnames and f'{logical}_v2' not in chk.sheetnames:
            problems.append(f'{logical}_v2 missing')
    for name in DATA_SHEETS:
        if name in chk.sheetnames:
            sh = chk[name]
            filled = sum(1 for r in sh.iter_rows(min_row=2, values_only=True)
                         if any(v is not None and str(v).strip() != '' for v in r))
            if filled:
                problems.append(f'{name} still has {filled} data rows')
    for cal in ('cal_curr', 'cal_new'):
        if cal in chk.sheetnames:
            sh = chk[cal]
            beyond = 0
            for r in range(CAL_FIRST_ROW, CAL_LAST_ROW + 1):
                for c in range(CAL_FIRST_COL + SIM_DAYS, min(sh.max_column, 40) + 1):
                    if sh.cell(row=r, column=c).value not in (None, ''):
                        beyond += 1
            if beyond:
                problems.append(f'{cal} has {beyond} cells past day {SIM_DAYS}')

    print(f'\nwritten {args.out}')
    print(f'  sheets: {len(chk.sheetnames)}')
    print(f'  proposal layer reset to base: {", ".join(report["v2_reset"])}')
    print(f'  calendars: {report["calendars"]}')
    print(f'  data sheets cleared to headers: {len(report["data_cleared"])}')
    for d in report['data_cleared']:
        print(f'    - {d}')
    print(f'  new data sheets: {", ".join(report["data_created"])}')
    if 'daily_rows_trimmed' in report:
        print(f'  EcoGainsSim_Daily rows past day {SIM_DAYS} cleared: {report["daily_rows_trimmed"]}')
    print('\nGATES')
    if problems:
        for p in problems:
            print(f'  FAIL {p}')
        raise SystemExit(1)
    print(f'  PASS every base config sheet has a *_v2 proposal twin')
    print(f'  PASS every data_* sheet is headers-only')
    print(f'  PASS neither calendar has content past day {SIM_DAYS}')
    print('\nNEXT: import into a new Google workbook, paste liveops20_fixes/engine/*.gs, run '
          'EcoGainsSim > Precompute calendars, then give the workbook ID to the analytics LLM '
          'along with PROMPT_variant_data_request.md.')


if __name__ == '__main__':
    main()
