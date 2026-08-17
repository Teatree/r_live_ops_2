# Load the analytics LLM's CSV drop into the variant-basis workbook.
#
# The delivered files match every sheet's schema exactly (column names, order, segment-label
# conventions), so loading is mechanical. What is NOT mechanical is the basis: this script measures
# what the data actually is — which arm, which window — and records the verdict in a `data_basis`
# sheet inside the workbook, so nobody reads a number out of it without knowing what it rests on.
#
# Usage: python liveops20_fixes/_load_data.py [--zip <path>] [--workbook <path>]
import argparse
import csv
import glob
import io
import os
import zipfile
from collections import defaultdict

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.join(HERE, '..')
DATA_IN = os.path.join(HERE, 'data_in')
WORKBOOK = os.path.join(ROOT, 'display', 'LiveOps20_variant_basis.xlsx')

SIM_DAYS = 21                      # what the workbook's engine assumes
WINDOW = ('2026-07-27', '2026-08-16')

# csv stem -> workbook sheet. The two *_ab_summary files have no sheet today; they are the only
# genuinely arm-split data in the drop, so they get created rather than dropped on the floor.
NEW_ASKS = ['data_core_spt', 'data_spend_action', 'data_ns_rounds']
SHEETS = ['data_gains', 'data_seg_beh', 'data_event_inst', 'data_event_accrual',
          'data_event_kite_accrual', 'data_RM', 'data_streaks', 'data_econ',
          'data_econ_daily'] + NEW_ASKS          # the new asks load exactly like the rest
EXTRA_SHEETS = ['data_gains_ab_summary', 'data_gains_ab_summary_wide']


def num(x):
    if x is None or x == '':
        return None
    try:
        f = float(x)
        return int(f) if f.is_integer() and abs(f) < 1e15 and '.' not in str(x) else f
    except ValueError:
        return x


def read_csv(path):
    with io.open(path, encoding='utf-8-sig') as f:
        r = csv.reader(f)
        rows = list(r)
    return rows[0], rows[1:]


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--zip', default=None, help='extract this zip into data_in/ first')
    ap.add_argument('--workbook', default=WORKBOOK)
    args = ap.parse_args()

    if args.zip:
        with zipfile.ZipFile(args.zip) as z:
            z.extractall(DATA_IN)
        print(f'extracted {args.zip} -> {DATA_IN}')

    available = {os.path.basename(p)[:-4]: p for p in glob.glob(os.path.join(DATA_IN, '*.csv'))}
    wb = openpyxl.load_workbook(args.workbook)
    report, problems, basis = [], [], []

    # ---------------------------------------------------------------- basis checks
    # PER SHEET, not globally: a drop can be half re-pulled, and mixing a variant anchor with
    # whole-population distributions is exactly the kind of thing that silently biases R and D.
    # Every test below is self-contained — no dependence on a previous drop.
    def dicts(name):
        hdr, rows = read_csv(available[name])
        return hdr, [dict(zip(hdr, r)) for r in rows]

    def fnum(x):
        try:
            return float(x)
        except (TypeError, ValueError):
            return 0.0

    verdicts = {}          # sheet -> (verdict, evidence)

    # reference cohort size, from the arm-split summary (the only file that carries ab_group)
    ref_pd = ref_players = None
    if 'data_gains_ab_summary' in available:
        _, srows = dicts('data_gains_ab_summary')
        arms = sorted({r['ab_group'] for r in srows})
        wins = sorted({(r['start_date'], r['end_date']) for r in srows})
        basis.append(('arm label upstream', ', '.join(arms),
                      f'{len(wins)} window(s): ' + '; '.join(f'{a}..{b}' for a, b in wins)))
        var = [a for a in arms if a.lower().startswith('variant')]
        if var:
            # segment_player_days repeats across source/resource rows, so take the max within each
            # cohort cell and sum the cells. The cell is (segment, payer_flag) once the summary
            # carries payer_flag — keying on segment alone would silently drop one payer's days.
            has_payer = 'payer_flag' in srows[0]
            per_seg_pd, per_seg_pl = {}, {}
            for r in srows:
                if r['ab_group'] != var[0]:
                    continue
                k = (r['segment'], r['payer_flag']) if has_payer else (r['segment'],)
                per_seg_pd[k] = max(per_seg_pd.get(k, 0), fnum(r.get('segment_player_days')))
                per_seg_pl[k] = max(per_seg_pl.get(k, 0), fnum(r.get('segment_players')))
            ref_pd, ref_players = sum(per_seg_pd.values()), sum(per_seg_pl.values())
            basis.append(('variant cohort (from summary)',
                          f'{ref_pd:,.0f} player-days, {ref_players:,.0f} players',
                          f'arm "{var[0]}" over {wins[0][0]}..{wins[0][1]}'))
        if wins and wins[0] != WINDOW:
            problems.append(f'the arm-split file covers {wins[0][0]}..{wins[0][1]}, '
                            f'not {WINDOW[0]}..{WINDOW[1]}')

    # data_gains — decisive: its HC totals must tie to the summary's variant column
    if 'data_gains' in available and 'data_gains_ab_summary' in available:
        _, grows = dicts('data_gains')
        _, srows = dicts('data_gains_ab_summary')
        segmap = {'B. 1-9': '0-9', 'C. 10-19': '10-19', 'D. 20-39': '20-39',
                  'E. 40-99': '40-99', 'F. 100+': '100+'}
        gsum, ssum = defaultdict(float), defaultdict(float)
        # This gate asks ONE question: is data_gains built on the variant cohort? It answers it by
        # tying HC totals to the summary. Since the 2026-08-17 drop, data_gains ships a CORRECTED
        # amount plus the raw one in *_uncorrected, while the summary stays raw — so comparing the
        # corrected column would measure the size of the correction (41.6% at 100+, all Rainbow
        # Maker) and report it as a wrong cohort, which is a different failure entirely. Tie on the
        # uncorrected column when it exists so the gate keeps testing the cohort; the correction is
        # reported separately below.
        amt = ('category_amount_uncorrected'
               if grows and 'category_amount_uncorrected' in grows[0] else 'category_amount')
        for r in grows:
            if r['resource'] == 'HC' and r['engagement_segment'] in segmap:
                gsum[segmap[r['engagement_segment']]] += fnum(r[amt])
        var = next((a for a in {r['ab_group'] for r in srows} if a.lower().startswith('variant')), None)
        for r in srows:
            if r['resource'] == 'HC' and r['ab_group'] == var:
                ssum[r['segment']] += fnum(r['total_amount'])
        worst = max((abs(gsum[s] - ssum[s]) / max(ssum[s], 1) for s in ssum), default=1)
        raw = ' (on the raw column; the correction is measured separately)' \
              if amt.endswith('_uncorrected') else ''
        verdicts['data_gains'] = (('VARIANT' if worst < 0.01 else 'NOT THE VARIANT ARM'),
                                  f'HC totals tie to summary "{var}" within {worst:.2%}{raw}')

        # The correction itself: how much of each segment's HC faucet the export withdrew. Reported,
        # never gated — it is the analytics side fixing a known over-count, not a data defect.
        if amt.endswith('_uncorrected'):
            corr = defaultdict(float)
            for r in grows:
                if r['resource'] == 'HC' and r['engagement_segment'] in segmap:
                    corr[segmap[r['engagement_segment']]] += fnum(r['category_amount'])
            hits = sorted(((s, 1 - corr[s] / ssum[s]) for s in ssum if ssum[s]),
                          key=lambda x: -x[1])
            verdicts['data_gains (correction)'] = (
                'APPLIED',
                'HC withdrawn by segment: ' + ', '.join(f'{s} {p:.1%}' for s, p in hits))

    # data_seg_beh — cohort size against the summary reference
    if 'data_seg_beh' in available and ref_pd:
        _, brows = dicts('data_seg_beh')
        pd_tot = sum(fnum(r['player_days']) for r in brows)
        ratio = pd_tot / ref_pd
        verdicts['data_seg_beh'] = (('VARIANT' if ratio < 1.6 else 'WHOLE POPULATION'),
                                    f'{pd_tot:,.0f} player-days = {ratio:.2f}x the variant cohort')

    # everything else — compare each sheet's own cohort proxy against the two sheets now known good
    beh_players, beh_pd = {}, 0.0
    if 'data_seg_beh' in available:
        _, brows = dicts('data_seg_beh')
        for r in brows:
            beh_players[(r['segment'], r['payer_flag'])] = fnum(r['unique_players'])
        beh_pd = sum(fnum(r['player_days']) for r in brows)

    if 'data_streaks' in available and beh_pd:
        _, rows = dicts('data_streaks')
        pd_tot = sum(fnum(r['player_days']) for r in rows)
        ratio = pd_tot / beh_pd
        verdicts['data_streaks'] = (('VARIANT' if ratio < 1.6 else 'WHOLE POPULATION'),
                                    f'{pd_tot:,.0f} player-days = {ratio:.1f}x data_seg_beh')

    if 'data_gains' in available:
        _, grows = dicts('data_gains')
        gearn = {}
        for r in grows:
            if r['resource'] == 'HC':
                gearn[(r['engagement_segment'], r['payer_flag'])] = fnum(r['resource_earners'])
        segmap2 = {'0-9': 'B. 1-9', '10-19': 'C. 10-19', '20-39': 'D. 20-39',
                   '40-99': 'E. 40-99', '100+': 'F. 100+'}
        for nm in ('data_econ', 'data_econ_daily'):
            if nm not in available:
                continue
            _, rows = dicts(nm)
            seen, tot = set(), 0.0
            for r in rows:
                if r['currency'] != 'HC':
                    continue
                k = (r['segment'], r['payer_flag'])
                if k in seen:
                    continue
                seen.add(k)
                tot += fnum(r['resource_earners'])
            ref = sum(gearn.get((segmap2.get(s, s), p), 0) for (s, p) in seen)
            ratio = tot / ref if ref else None
            if ratio:
                verdicts[nm] = (('VARIANT' if ratio < 1.6 else 'WHOLE POPULATION'),
                                f'HC earners {tot:,.0f} = {ratio:.1f}x data_gains')

    # Event sheets — participation cannot exceed 100% of the cohort. Careful with the denominator:
    # `avg_participants_per_instance` is already per instance, but `n_participants` on the accrual
    # sheets is POOLED across every instance in the window, so it legitimately exceeds the player
    # count for a multi-instance event (Night Sky runs ~90 instances here). Divide by n_instances
    # for those, or the check condemns correct data.
    for nm, col, per_instance in (('data_event_inst', 'avg_participants_per_instance', True),
                                  ('data_event_accrual', 'n_participants', False),
                                  ('data_event_kite_accrual', 'n_participants', False)):
        if nm not in available or not beh_players:
            continue
        _, rows = dicts(nm)
        worst_ratio, worst_note = 0, None
        for r in rows:
            pl = beh_players.get((r['segment'], r['payer_flag']))
            if not pl:
                continue
            denom = pl if per_instance else pl * max(fnum(r.get('n_instances')) or 1, 1)
            ratio = fnum(r[col]) / denom
            if ratio > worst_ratio:
                worst_ratio = ratio
                worst_note = (f"{r['event_name']} {r['segment']}/{r['payer_flag']}: "
                              f"{fnum(r[col]):,.0f} participants vs {pl:,.0f} players"
                              + ('' if per_instance
                                 else f" over {fnum(r.get('n_instances')):,.0f} instances"))
        verdicts[nm] = (('VARIANT' if worst_ratio <= 1.05 else 'WHOLE POPULATION'),
                        f'max participation rate = {worst_ratio:.2f} of the cohort ({worst_note})')

    # window
    if 'data_econ_daily' in available:
        hdr, rows = read_csv(available['data_econ_daily'])
        i_d = hdr.index('day_index')
        days = sorted({int(float(r[i_d])) for r in rows if r[i_d]})
        basis.append(('window (data_econ_daily)', f'{len(days)} days (day_index {days[0]}..{days[-1]})',
                      f'the workbook engine assumes {SIM_DAYS} days, {WINDOW[0]}..{WINDOW[1]}'))
        if len(days) != SIM_DAYS:
            problems.append(f'data_econ_daily carries {len(days)} day indices, not {SIM_DAYS}')

    for nm, (v, ev) in sorted(verdicts.items()):
        basis.append((f'basis: {nm}', v, ev))
        if v not in ('VARIANT', 'APPLIED'):
            problems.append(f'{nm} is {v.lower()} — {ev}')

    # ---------------------------------------------------------------- load
    for name in SHEETS:
        if name not in available:
            problems.append(f'{name} was not delivered')
            report.append((name, 'MISSING', ''))
            continue
        if name not in wb.sheetnames:
            problems.append(f'{name} has no sheet in the workbook')
            continue
        sh = wb[name]
        want = [str(c.value) for c in sh[1] if c.value is not None]
        hdr, rows = read_csv(available[name])
        missing = [c for c in want if c not in hdr]
        extra = [c for c in hdr if c not in want]
        if missing:
            problems.append(f'{name} header mismatch — missing {missing}')
            report.append((name, 'HEADER MISMATCH', f'{len(rows)} rows not loaded'))
            continue
        # EXTRA columns are additive, not a mismatch: the export gains an audit column (e.g.
        # data_gains' *_uncorrected trio in the 2026-08-17 drop) and refusing the whole sheet
        # would silently leave the most important table stale. Every reader here works by header
        # NAME, so an unknown column is inert — append it to the header and keep its values so the
        # provenance survives in the workbook rather than being dropped on the floor.
        for c in extra:
            want.append(c)
            sh.cell(row=1, column=len(want)).value = c
        # Map by NAME, not position: the engine reads these sheets by header name, so a reordered
        # export is harmless as long as every column is present — writing positionally would put
        # values in the wrong columns silently, which is the worst possible failure here.
        col_of = {c: want.index(c) + 1 for c in hdr}
        reordered = hdr != want
        for r in range(2, sh.max_row + 1):                      # clear old body
            for c in range(1, sh.max_column + 1):
                sh.cell(row=r, column=c).value = None
        for i, row in enumerate(rows, start=2):
            for j, v in enumerate(row):
                if j < len(hdr):
                    sh.cell(row=i, column=col_of[hdr[j]]).value = num(v)
        note = f'{len(rows)} rows x {len(hdr)} cols'
        if reordered:
            note += ' (columns reordered on load)'
        if extra:
            note += f' (+{len(extra)} new column(s): {", ".join(extra)})'
        report.append((name, 'loaded', note))

    for name in EXTRA_SHEETS:
        if name not in available:
            continue
        hdr, rows = read_csv(available[name])
        if name in wb.sheetnames:
            del wb[name]
        sh = wb.create_sheet(name)
        for j, h in enumerate(hdr, start=1):
            sh.cell(row=1, column=j).value = h
        for i, row in enumerate(rows, start=2):
            for j, v in enumerate(row, start=1):
                sh.cell(row=i, column=j).value = num(v)
        report.append((name, 'created (arm-split reference)', f'{len(rows)} rows x {len(hdr)} cols'))

    reported = {n for n, _, _ in report}
    for name in NEW_ASKS:
        if name in reported:
            continue                       # already covered by the load loop
        if name in wb.sheetnames:
            sh = wb[name]
            filled = sum(1 for r in sh.iter_rows(min_row=2, values_only=True)
                         if any(v is not None for v in r))
            if not filled:
                report.append((name, 'STILL EMPTY', 'not in the drop'))
                problems.append(f'{name} was requested and not delivered')

    # ---------------------------------------------------------------- provenance sheet
    if 'data_basis' in wb.sheetnames:
        del wb['data_basis']
    sh = wb.create_sheet('data_basis', 0)
    sh.column_dimensions['A'].width = 34
    sh.column_dimensions['B'].width = 46
    sh.column_dimensions['C'].width = 90
    rows_out = [['WHAT THE DATA IN THIS WORKBOOK ACTUALLY IS', '', ''], ['', '', '']]
    rows_out.append(['loaded from', os.path.basename(args.zip) if args.zip else 'liveops20_fixes/data_in/', ''])
    for k, v, note in basis:
        rows_out.append([k, v, note])
    rows_out += [['', '', ''], ['SHEET', 'STATUS', 'DETAIL']]
    for n, s, d in report:
        rows_out.append([n, s, d])
    if problems:
        rows_out += [['', '', ''], ['OPEN PROBLEMS', '', '']]
        for p in problems:
            rows_out.append(['', p, ''])
    for i, row in enumerate(rows_out, start=1):
        for j, v in enumerate(row, start=1):
            sh.cell(row=i, column=j).value = v

    # the workbook is often open in Excel, which locks it — fall back to a sibling file rather than
    # dying after all the work, and say plainly which file to use
    out_path = args.workbook
    try:
        wb.save(out_path)
    except PermissionError:
        stem, ext = os.path.splitext(args.workbook)
        out_path = f'{stem}__LOADED{ext}'      # deterministic, so re-runs overwrite one file
        wb.save(out_path)
        print(f'\n! {os.path.basename(args.workbook)} is open/locked — wrote {os.path.basename(out_path)} '
              f'instead. Close the original and re-run to write in place.')

    print(f'\nwritten {out_path}')
    print('\nBASIS')
    for k, v, note in basis:
        print(f'  {k:20s} {v}')
        if note:
            print(f'  {"":20s}   {note}')
    print('\nSHEETS')
    for n, s, d in report:
        print(f'  {n:26s} {s:28s} {d}')
    if problems:
        print('\nOPEN PROBLEMS')
        for p in problems:
            print(f'  ! {p}')
    # 'APPLIED' is informational (the export's own correction), not a cohort verdict — a basis
    # check answers VARIANT / NOT THE VARIANT ARM and only those two count toward the summary.
    wrong_basis = [n for n, (v, _) in verdicts.items() if v not in ('VARIANT', 'APPLIED')]
    missing = [n for n, s, _ in report if s in ('MISSING', 'STILL EMPTY')]
    print()
    if wrong_basis:
        print(f'BASIS: {len(wrong_basis)} sheet(s) are NOT on the variant cohort '
              f'({", ".join(sorted(wrong_basis))}) — R / D / T are priced off the wrong population.')
    else:
        print(f'BASIS: every delivered sheet ({len(verdicts)}) is on the variant cohort. '
              f'Levels and distributions agree.')
    if missing:
        print(f'GAPS:  {len(missing)} sheet(s) still absent: {", ".join(missing)}.')
        if 'data_RM' in missing:
            print('       Without data_RM the engine cannot price Rainbow Maker bottom-up; it '
                  'falls back to carrying the measured value (diff 0), so RM proposals are inert.')
    if not problems:
        print('\nno problems found — every sheet is on the variant basis')


if __name__ == '__main__':
    main()
