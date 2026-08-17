# Join + score: sim (wb16, offline harness dump) vs LiveOps v2 A/B actuals.
#
# DESIGN (locked with Garry 2026-08-13):
#   delta-vs-delta — r_sim = S̄/M̄ (windowed sim / windowed measured, payer-blended, per-earner
#   units cancel) vs r_act = (A_V/PD_V)/(A_C/PD_C) (per-active-player-day, FREE, corrected
#   ledger). RoR = r_sim/r_act. Window = cal_new days 5..13 mapped onto Aug 2-10 (phase corrected 2026-08-14).
#   Absolute view: dSim = a_C*(r_sim-1) vs dAct = a_V-a_C (both papd; anchor error quarantined
#   into the anchor-check boxes).
#   Class system prevents division-by-zero garbage: SCORED / KILL_AGREED / KILL_MISSED /
#   KILL_PHANTOM / NEW_BOTH / SIM_ONLY_NEW / ACT_ONLY_NEW / TRIVIAL.
#
# Core x SPT override: the resource ledger cannot see level-completion token grants, so the
# actual side of (Core, SPT) [and (Saga, SPT) which the telemetry folds into the same CORE
# lane] comes from liveops2_ab_daily_metrics avg_spt_gain_CORE (players-weighted, corrected).
#
# Output: analysis/out/comparison.json — the single source of truth for the HTML report.
# Every number in the report traces here.
#
# Usage: python analysis/_build_comparison.py   (after _dump_sim_matrix.js and _extract_actuals.py)
import csv, glob, json, math, os, re, sys
from collections import defaultdict

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, 'out')
WB = os.path.join(HERE, '..', 'workbooks')

sim = json.load(open(os.path.join(OUT, 'sim_matrix.json'), encoding='utf-8'))
aux = json.load(open(os.path.join(OUT, 'actuals_aux.json'), encoding='utf-8'))
mock = json.load(open(os.path.join(HERE, '..', 'harness', '_mockdata.json'), encoding='utf-8'))
actual_rows = list(csv.DictReader(open(os.path.join(OUT, 'actuals_long.csv'), encoding='utf-8')))

CATS = sim['meta']['categories']
RES = sim['meta']['resources']
SEGS = ['0-9', '10-19', '20-39', '40-99', '100+']          # A. 0 excluded from accuracy (appendix)
COMPARABLE_RES = ['HC', 'Slingshot', 'Shuffle', 'Comet', 'Red', 'Chuck', 'Bomb',
                  'UL Bomb', 'UL Chuck', 'UL Red', 'Unlimited Lives', 'SPT']
gates = {}

# ---------------------------------------------------------------- payer blend weights
# earner-share from data_gains resource_earners (per segment x payer x resource; repeats across
# category rows — first hit wins). Fallback: unique_players from data_seg_beh.
dg = mock['data_gains']['values']
dgi = {h: i for i, h in enumerate(dg[0])}
SEG_TO_GAINS = {'0-9': 'B. 1-9', '10-19': 'C. 10-19', '20-39': 'D. 20-39',
                '40-99': 'E. 40-99', '100+': 'F. 100+', 'A. 0': 'A. 0'}
earners = {}
for r in dg[1:]:
    k = (r[dgi['engagement_segment']], r[dgi['payer_flag']], r[dgi['resource']])
    if k not in earners:
        try: earners[k] = float(r[dgi['resource_earners']])
        except (TypeError, ValueError): pass
sb = mock['data_seg_beh']['values']
sbi = {h: i for i, h in enumerate(sb[0])}
seg_players = {(r[sbi['segment']], r[sbi['payer_flag']]): float(r[sbi['unique_players']])
               for r in sb[1:] if r[sbi['segment']]}
def payer_weight(seg, res):
    """w_PAYER for (segment, resource); returns (w_payer, source_tag)."""
    g = SEG_TO_GAINS[seg]
    np_, p_ = earners.get((g, 'NONPAYER', res)), earners.get((g, 'PAYER', res))
    if np_ and p_: return p_ / (np_ + p_), 'earners'
    np2, p2 = seg_players.get((seg, 'NONPAYER')), seg_players.get((seg, 'PAYER'))
    if np2 and p2: return p2 / (np2 + p2), 'players'
    return 0.0, 'nonpayer_only'

# ---------------------------------------------------------------- sim cells (blended)
simcell = defaultdict(lambda: {'cur_win': 0.0, 'new_win': 0.0, 'meas_33': 0.0, 'sim_33': 0.0})
raw_sim = {(r['payer'], r['segment'], r['category'], r['resource']): r for r in sim['rows']}
weight_census = defaultdict(int)
for seg in SEGS + ['A. 0']:
    for cat in CATS:
        for res in RES:
            w_p, wsrc = payer_weight(seg, res)
            weight_census[wsrc] += 1
            c = simcell[(seg, cat, res)]
            for payer, w in [('NONPAYER', 1 - w_p), ('PAYER', w_p)]:
                r = raw_sim.get((payer, seg, cat, res))
                if r:
                    c['cur_win'] += w * r['cur_win']; c['new_win'] += w * r['new_win']
                    c['meas_33'] += w * r['meas_33']; c['sim_33'] += w * r['sim_33']
gates['payer_weight_census'] = dict(weight_census)

# ---------------------------------------------------------------- actual cells
act = defaultdict(lambda: {'free': 0.0, 'papd': 0.0})
player_days = {}
for r in actual_rows:
    k = (r['segment'], r['ab_group'], r['category'], r['resource'])
    act[k]['free'] += float(r['amount_free']); act[k]['papd'] += float(r['papd_free'])
    player_days[(r['segment'], r['ab_group'])] = float(r['player_days'])

# Core x SPT override from the telemetry lane (the ledger cannot see per-level grants).
# The CORE lane covers level completions — the sim splits that anchor between 'Core' (the
# synthetic per-level SPT) and nothing else, so the override lands on Core alone; the tiny
# ledger-visible Core/Saga SPT rows (0 in this export) are replaced, not double-counted.
spt_lanes = aux['spt_lanes_papd']
core_spt_note = {}
for seg in SEGS:
    for arm in ('Control', 'Variant'):
        lane = spt_lanes.get(f'{seg}|{arm}')
        if lane:
            act[(seg, arm, 'Core', 'SPT')] = {'free': lane['CORE'] * player_days[(seg, arm)],
                                              'papd': lane['CORE']}
            core_spt_note[f'{seg}|{arm}'] = lane['CORE']
# gate: ledger event-lane DPT ~ telemetry non-CORE lanes (sanity that the two sources agree)
ev_check = {}
for seg in SEGS:
    for arm in ('Control', 'Variant'):
        ledger = sum(act[(seg, arm, c, 'SPT')]['papd'] for c in CATS if c != 'Core')
        lane = spt_lanes.get(f'{seg}|{arm}', {})
        telem = sum(lane.get(k, 0) for k in ('RAINBOW', 'EVENT', 'OFFER', 'OTHER'))
        ev_check[f'{seg}|{arm}'] = {'ledger': round(ledger, 3), 'telemetry': round(telem, 3),
                                    'ratio': round(ledger / telem, 3) if telem else None}
gates['spt_event_lane_check'] = ev_check

# ---------------------------------------------------------------- floors + classification
TAU_SHARE = 0.005
tau = {(seg, res): TAU_SHARE * sum(act[(seg, 'Control', c, res)]['papd'] for c in CATS)
       for seg in SEGS for res in COMPARABLE_RES}
mu = {(seg, res): TAU_SHARE * sum(simcell[(seg, c, res)]['cur_win'] for c in CATS)
      for seg in SEGS for res in COMPARABLE_RES}

def classify(M, S, m33, s33, aC, aV, t, m):
    """Returns (class, r_sim, r_act, window_fallback).
    Two floor tiers per side: PRESENCE (0.1x the materiality floor — enough to compute a ratio)
    and MATERIALITY (t / m — enough to matter). r_sim is the WINDOWED ratio except for
    WINDOW_MISS cells, where the 33-day ratio stands in (window_fallback=True)."""
    tp, mp = 0.1 * t, 0.1 * m                                 # presence floors
    r_sim = S / M if M > max(mp, 1e-12) else None
    r_sim33 = s33 / m33 if m33 > max(mp, 1e-12) else None
    r_act = aV / aC if aC > max(tp, 1e-12) else None
    sim_has, act_has = M > mp, aC > tp
    material = (M > m) or (S > m) or (aC > t) or (aV > t)
    if not material: return 'TRIVIAL', r_sim, r_act, False
    if not sim_has and not act_has:                           # neither window baseline pays
        if S > m and aV > t: return 'NEW_BOTH', r_sim, r_act, False
        if S > m: return 'SIM_ONLY_NEW', r_sim, r_act, False
        if aV > t:
            # live Variant launched it; does the sim model it anywhere in the 33 days?
            return ('WINDOW_MISS', r_sim33, r_act, True) if s33 > 1e-9 else \
                   ('ACT_ONLY_NEW', r_sim, r_act, False)
        return 'TRIVIAL', r_sim, r_act, False
    if not sim_has and act_has:
        # sim's window allocation is empty on a source live Control pays: placement artifact
        # (leaderboard END-day outside the mapped day list) if the 33-day model has it.
        return ('WINDOW_MISS', r_sim33, r_act, True) if m33 > 1e-9 else \
               ('NOT_MODELED', None, r_act, False)
    if sim_has and not act_has:
        if aV > t:
            # sim's measured baseline pays this in-window but the live CONTROL arm didn't run
            # it while the Variant does — the Night Sky framing mismatch (sim answers "config
            # change", the A/B answers "on vs off").
            return 'BASELINE_MISMATCH', r_sim, r_act, False
        # sim's baseline pays a source the live test-era world has in NEITHER arm
        return 'STALE_BASELINE', r_sim, r_act, False
    if r_act is not None and r_act < 0.1:
        if r_sim is not None and r_sim < 0.1: return 'KILL_AGREED', r_sim, r_act, False
        if r_sim is None or r_sim >= 0.5: return 'KILL_MISSED', r_sim, r_act, False
    if r_sim is not None and r_sim < 0.1 and r_act is not None and r_act >= 0.5:
        return 'KILL_PHANTOM', r_sim, r_act, False
    if r_sim is None or r_act is None: return 'TRIVIAL', r_sim, r_act, False
    return 'SCORED', r_sim, r_act, False

def status_for(cls, r_sim, r_act):
    if cls in ('SCORED', 'WINDOW_MISS') and r_sim is not None and r_act is not None:
        ror = r_sim / r_act
        sign_flip = (r_sim - 1) * (r_act - 1) < 0 and abs(r_sim - 1) > 0.05 and abs(r_act - 1) > 0.05
        if sign_flip: return 'red', ror, True
        if ror < 2/3 or ror > 1.5: return 'red', ror, False
        if ror < 0.8 or ror > 1.25: return 'amber', ror, False
        return 'green', ror, False
    if cls == 'KILL_AGREED': return 'green', 1.0, False
    if cls in ('KILL_MISSED', 'KILL_PHANTOM', 'SIM_ONLY_NEW', 'ACT_ONLY_NEW',
               'BASELINE_MISMATCH', 'NOT_MODELED', 'STALE_BASELINE'): return 'red', None, False
    if cls == 'NEW_BOTH': return 'info', None, False
    return 'na', None, False

cells = []
for seg in SEGS:
    for cat in CATS:
        for res in COMPARABLE_RES:
            sc = simcell[(seg, cat, res)]
            aC = act[(seg, 'Control', cat, res)]['papd']
            aV = act[(seg, 'Variant', cat, res)]['papd']
            M, S = sc['cur_win'], sc['new_win']
            if M < 1e-12 and S < 1e-12 and aC < 1e-12 and aV < 1e-12: continue
            if cat == 'Daily Night Sky Prize':
                # wb16 removed NS from cal_curr (baseline matches the live Control), but the sim's
                # MEASURED side still carries the May anchor flat/33 — a delta-vs-delta score would
                # re-litigate the on-vs-off framing Garry has closed. NS is compared DIRECTLY
                # (sim NEW-side prediction vs live Variant) in its own story panel.
                cells.append({'segment': seg, 'category': cat, 'resource': res, 'class': 'NS_DIRECT',
                              'status': 'info', 'sign_flip': False, 'window_fallback': False,
                              'M_win': M, 'S_win': S, 'a_C': aC, 'a_V': aV,
                              'r_sim': None, 'r_act': None, 'RoR': None, 'log2_RoR': None,
                              'd_act_papd': aV - aC, 'd_sim_papd': None, 'err_papd': None})
                continue
            cls, r_sim, r_act, win_fb = classify(M, S, sc['meas_33'], sc['sim_33'],
                                                 aC, aV, tau[(seg, res)], mu[(seg, res)])
            status, ror, flip = status_for(cls, r_sim, r_act)
            d_act = aV - aC
            d_sim = aC * (r_sim - 1) if (r_sim is not None and aC > 0) else None
            cells.append({'segment': seg, 'category': cat, 'resource': res, 'class': cls,
                          'status': status, 'sign_flip': flip, 'window_fallback': win_fb,
                          'M_win': M, 'S_win': S, 'a_C': aC, 'a_V': aV,
                          'r_sim': r_sim, 'r_act': r_act, 'RoR': ror,
                          'log2_RoR': (math.log2(ror) if ror and ror > 0 else None),
                          'd_act_papd': d_act, 'd_sim_papd': d_sim,
                          'err_papd': (d_act - d_sim) if d_sim is not None else None})
gates['class_census'] = dict(defaultdict(int))
cc = defaultdict(int)
for c in cells: cc[c['class']] += 1
gates['class_census'] = dict(cc)

# ---------------------------------------------------------------- PerSource reconciliation
import openpyxl
# Garry replaces this workbook rather than versioning it (the "(2)" copy was deleted 2026-08-17),
# so resolve it the same way _extract_ab_summary.py does: highest (N) suffix wins, then newest mtime.
def _newest_ab_summary():
    cands = glob.glob(os.path.join(WB, 'LiveOps_v2_AB_Summary*.xlsx'))
    if not cands:
        raise SystemExit('no LiveOps_v2_AB_Summary*.xlsx in workbooks/')
    def key(p):
        m = re.search(r'\((\d+)\)', os.path.basename(p))
        return (int(m.group(1)) if m else 0, os.path.getmtime(p))
    return max(cands, key=key)

AB_SUMMARY_PATH = _newest_ab_summary()
wb2 = openpyxl.load_workbook(AB_SUMMARY_PATH, data_only=True)
ps = wb2['PerSource']
PS_HDRS = {3: '0-9', 32: '10-19', 61: '20-39', 90: '40-99', 119: '100+', 148: 'A. 0'}
PS_RES = [ps.cell(4, c).value for c in range(3, 16)]        # HC..SPTx2 (13)
ps_actual = {}                                              # (seg, source, res) -> value
for hr, seg in PS_HDRS.items():
    for i in range(25):
        src = ps.cell(hr + 2 + i, 2).value
        for j, res in enumerate(PS_RES):
            v = ps.cell(hr + 2 + i, 17 + j).value
            if isinstance(v, (int, float)): ps_actual[(seg, src, res)] = float(v)
# The ACTUAL block's per-earner unit could not be reproduced from any repo CSV (candidates
# tested by hand: free/resource-gainers, free/gainer_days, free/player_days x{1,8,9}, both arms
# — all off by 1.3-20x on the Ads probe cell; Garry's per-source distinct-gainer query is not in
# the repo). Reconcile on STRUCTURE instead: each source's share of segment HC, which any
# per-source-constant denominator cancels out of... it does NOT cancel a per-source denominator,
# so treat this as a directional check, flagged in the report.
gpg = {(p['segment'], p['arm'], p['resource']): p for p in aux['player_dist']}
recon = []
for seg in SEGS:
    ps_tot = sum(v for (s, c, r), v in ps_actual.items() if s == seg and r == 'HC' and v > 0)
    my_tot = sum(act[(seg, 'Variant', c, 'HC')]['free'] for c in CATS)
    for cat in CATS:
        v_ps = ps_actual.get((seg, cat, 'HC'))
        if v_ps is None: continue
        share_ps = v_ps / ps_tot if ps_tot else 0
        share_my = act[(seg, 'Variant', cat, 'HC')]['free'] / my_tot if my_tot else 0
        if share_ps > 0.01 or share_my > 0.01:
            recon.append({'segment': seg, 'source': cat,
                          'persource_share': round(share_ps, 4), 'my_share': round(share_my, 4),
                          'delta_pp': round((share_my - share_ps) * 100, 2)})
gates['persource_unit'] = ('UNRESOLVED — per-earner-per-source basis from a query outside the repo; '
                           'structure (share-of-segment-HC) reconciliation only')
gates['persource_recon_structure'] = recon
gates['persource_max_share_delta_pp'] = max((abs(r['delta_pp']) for r in recon), default=None)

# ---------------------------------------------------------------- anchor check
# per-earner-9d units both sides: sim M̄_win (windowed measured per earner) vs actual Control
# free amount / resource-level 9-day gainers. Flagged: gainers count any-source gainers.
anchor = []
for seg, cat in [('20-39', 'Saga'), ('20-39', 'Daily Gift'), ('20-39', 'Core'),
                 ('100+', 'Saga'), ('100+', 'Daily Gift'), ('10-19', 'Team Race'),
                 ('10-19', 'Daily Night Sky Prize')]:
    g = gpg.get((seg, 'Control', 'HC'), {}).get('gainers', 0)
    freeC = act[(seg, 'Control', cat, 'HC')]['free']
    M = simcell[(seg, cat, 'HC')]['cur_win']
    if g and (freeC or M):
        anchor.append({'segment': seg, 'category': cat, 'resource': 'HC',
                       'sim_meas_win_per_earner': round(M, 2),
                       'actual_C_per_gainer_9d': round(freeC / g, 2),
                       'ratio': round(M / (freeC / g), 3) if freeC else None})

# ---------------------------------------------------------------- story data
story = {}
# NS — DIRECT comparison (wb16: cal_curr no longer runs NS, matching the live Control; the sim's
# NEW side = May-anchored measured x R(NS_v2/NS) x T=1 is a live prediction of the Variant's NS
# faucet). Levels bridged per-gainer (sim per-earner window vs Variant free / HC-gainers 9d).
ns = {'sim': {}, 'act': {}}
for seg in SEGS:
    sc = simcell[(seg, 'Daily Night Sky Prize', 'HC')]
    scS = simcell[(seg, 'Daily Night Sky Prize', 'SPT')]
    g_v = gpg.get((seg, 'Variant', 'HC'), {}).get('gainers', 0) if 'gpg' in dir() else 0
    ns['sim'][seg] = {'new_win': round(sc['new_win'], 2), 'meas_win': round(sc['cur_win'], 2),
                      'sim_33': round(sc['sim_33'], 2),
                      'R': round(sc['sim_33'] / sc['meas_33'], 4) if sc['meas_33'] else None,
                      'new_win_spt': round(scS['new_win'], 2)}
    ns['act'][seg] = {'a_C': round(act[(seg, 'Control', 'Daily Night Sky Prize', 'HC')]['papd'], 3),
                      'a_V': round(act[(seg, 'Variant', 'Daily Night Sky Prize', 'HC')]['papd'], 3),
                      'a_V_spt': round(act[(seg, 'Variant', 'Daily Night Sky Prize', 'SPT')]['papd'], 3)}
nse = wb2['NS_Config_Change_Est']
ns['ns_config_change_est'] = [[c.value for c in row[:10]] for row in nse.iter_rows(min_row=1, max_row=min(nse.max_row, 40))]
story['night_sky'] = ns
# Core-SPT R candidates
plr = aux['per_level_spt_rate']
by_bucket = defaultdict(dict)
for r in plr: by_bucket[r['play_bucket']][r['ab_group']] = float(r['spt_per_saga_level'] or 0)
ratios = [b['Variant'] / b['Control'] for b in by_bucket.values() if b.get('Control') and b.get('Variant')]
story['core_spt'] = {
    'sim_R_wb15': 0.81229, 'sim_E_base': 15.05, 'sim_E_v2': 12.225,
    'live_per_level_ratio_play_matched': round(sum(ratios) / len(ratios), 4) if ratios else None,
    'per_bucket_ratios': {k: round(b['Variant'] / b['Control'], 4)
                          for k, b in by_bucket.items() if b.get('Control') and b.get('Variant')},
    'telemetry_core_papd': core_spt_note,
    'note': 'live daily ratio ran ~0.51 to 29 Jul, ~0.85 from 31 Jul (prior report §07); '
            'wb16 SP_v2 models the late regime',
}
# RM: ladder conformance + which config ran
rm_cfg = aux['rm_config']
rm1 = mock['RM_1st']['values']; rm2 = mock['RM_2nd']['values']
story['rainbow_maker'] = {
    'live_config_rows': len(rm_cfg),
    'act_papd': {seg: {'C': round(act[(seg, 'Control', 'Rainbow Maker', 'HC')]['papd'], 3),
                       'V': round(act[(seg, 'Variant', 'Rainbow Maker', 'HC')]['papd'], 3),
                       'V_spt': round(act[(seg, 'Variant', 'Rainbow Maker', 'SPT')]['papd'], 3)}
                 for seg in SEGS},
    'sim_win': {seg: {'HC': round(simcell[(seg, 'Rainbow Maker', 'HC')]['new_win'], 2),
                      'SPT': round(simcell[(seg, 'Rainbow Maker', 'SPT')]['new_win'], 2)}
                for seg in SEGS},
    'spt_conformance': aux['rm_spt_conformance'][:5],
}
# killed events + chuck flip
story['killed'] = {}
for cat in ['Red Challenge', 'Bomb Challenge', 'Chuck Challenge']:
    story['killed'][cat] = {seg: {'a_C': round(act[(seg, 'Control', cat, 'HC')]['papd'], 3),
                                  'a_V': round(act[(seg, 'Variant', cat, 'HC')]['papd'], 3),
                                  'r_sim': (round(simcell[(seg, cat, 'HC')]['new_win'] /
                                                  simcell[(seg, cat, 'HC')]['cur_win'], 3)
                                            if simcell[(seg, cat, 'HC')]['cur_win'] > 1e-9 else None)}
                            for seg in SEGS}
# Other decomposition (single_collection & dream_peak live here)
audit = list(csv.DictReader(open(os.path.join(OUT, 'mapping_audit.csv'), encoding='utf-8')))
story['other_composition'] = [a for a in audit if a['category'] == 'Other'][:15]
# Season Pass milestones
story['sp_milestones'] = {'split': aux['sp_milestone_split'], 'summary': aux['sp_milestone_summary']}
# Saga / Daily Gift
for cat, key in [('Saga', 'saga'), ('Daily Gift', 'daily_gift')]:
    story[key] = {seg: {'a_C': round(act[(seg, 'Control', cat, 'HC')]['papd'], 3),
                        'a_V': round(act[(seg, 'Variant', cat, 'HC')]['papd'], 3),
                        'r_act': (round(act[(seg, 'Variant', cat, 'HC')]['papd'] /
                                        act[(seg, 'Control', cat, 'HC')]['papd'], 3)
                                  if act[(seg, 'Control', cat, 'HC')]['papd'] else None),
                        'r_sim': (round(simcell[(seg, cat, 'HC')]['new_win'] /
                                        simcell[(seg, cat, 'HC')]['cur_win'], 3)
                                  if simcell[(seg, cat, 'HC')]['cur_win'] > 1e-9 else None)}
                  for seg in SEGS}
# daily overlay (normalized shapes): sim ALL HC per day (blended) vs actual hc daily papd
overlay = {}
for seg in SEGS:
    w_p, _ = payer_weight(seg, 'HC')
    iHC = RES.index('HC')
    sim_days = []
    for d in sim['meta']['day_list']:
        v = 0.0
        for payer, w in [('NONPAYER', 1 - w_p), ('PAYER', w_p)]:
            ser = sim['daily'].get(f'{seg}|ALL|{payer}')
            if ser: v += w * ser['new'][d - 1][iHC]
        sim_days.append(v)
    act_days = sorted([r for r in aux['hc_daily'] if r['segment'] == seg and r['arm'] == 'Variant'],
                      key=lambda r: r['date'])
    overlay[seg] = {'dates': [r['date'] for r in act_days],
                    'cal_days': sim['meta']['day_list'],
                    'sim_new_hc': [round(v, 3) for v in sim_days],
                    'act_v_hc_papd': [round(r['hc_free_papd'], 3) for r in act_days]}
story['daily_overlay'] = overlay

# ---------------------------------------------------------------- headline aggregates
agg = {}
pdC = {seg: player_days[(seg, 'Control')] for seg in SEGS}
pdV = {seg: player_days[(seg, 'Variant')] for seg in SEGS}
PD_C, PD_V = sum(pdC.values()), sum(pdV.values())
for res in COMPARABLE_RES:
    # player-day-weighted overall papd (per segment rates weighted by that segment's exposure)
    sc_c = sum(c['a_C'] * pdC[c['segment']] for c in cells if c['resource'] == res) / PD_C
    sc_v = sum(c['a_V'] * pdV[c['segment']] for c in cells if c['resource'] == res) / PD_V
    pred = sum(c['d_sim_papd'] * pdC[c['segment']] for c in cells
               if c['resource'] == res and c['d_sim_papd'] is not None) / PD_C
    act_on_scored = sum(c['d_act_papd'] * pdC[c['segment']] for c in cells
                        if c['resource'] == res and c['d_sim_papd'] is not None) / PD_C
    per_seg = {seg: {'a_C': round(sum(c['a_C'] for c in cells if c['resource'] == res and c['segment'] == seg), 3),
                     'a_V': round(sum(c['a_V'] for c in cells if c['resource'] == res and c['segment'] == seg), 3),
                     'pred': round(sum(c['d_sim_papd'] for c in cells if c['resource'] == res
                                       and c['segment'] == seg and c['d_sim_papd'] is not None), 3)}
               for seg in SEGS}
    agg[res] = {'a_C_papd': round(sc_c, 3), 'a_V_papd': round(sc_v, 3),
                'actual_change_papd': round(sc_v - sc_c, 3),
                'predicted_change_scored_papd': round(pred, 3),
                'actual_change_on_scored_papd': round(act_on_scored, 3), 'per_segment': per_seg}

# ------------------------------------------------------- whole-faucet (full-scope) view
# The scored subset deliberately drops every cell that cannot be ratio-scored (Night Sky,
# Rainbow Maker, Chuck Challenge, the Other kills...). Garry reads the WHOLE-faucet number off
# `Sim per Segment`, so publish it on the same papd basis: per segment, the sim's full-scope
# windowed ratio over ALL categories, bridged onto that segment's live Control faucet, then
# player-day weighted. Same bridge as the per-cell d_sim_papd — just a wider numerator.
full = {}
for res in COMPARABLE_RES:
    seg_rows, num = {}, 0.0
    # additive per-source decomposition of the SAME total. The bridge a_C*(S/M - 1) is linear in
    # (S - M), so k_s = a_C_s / M_s converts any category's per-earner windowed delta onto the papd
    # axis and the parts sum to the whole. This is the only way to put a number on the sources with
    # no measured anchor (Rainbow Maker, and Night Sky on the wb16 baseline): a per-cell RATIO does
    # not exist there, but a per-cell CONTRIBUTION does. Weaker assumption than the scored cells'
    # (the segment's overall earner→active-player ratio is applied to that source), so it feeds the
    # whole-faucet narrative, not the accuracy scoring.
    kbr = {}
    for seg in SEGS:
        Ms = sum(simcell[(seg, c, res)]['cur_win'] for c in CATS)
        aCs = sum(act[(seg, 'Control', c, res)]['papd'] for c in CATS)
        kbr[seg] = (aCs / Ms) if Ms > 1e-12 else None
    by_cat, bridge_sum = {}, 0.0
    for cat in CATS:
        dsim = sum(kbr[seg] * (simcell[(seg, cat, res)]['new_win'] - simcell[(seg, cat, res)]['cur_win'])
                   * pdC[seg] for seg in SEGS if kbr[seg] is not None) / PD_C
        dact = sum((act[(seg, 'Variant', cat, res)]['papd'] - act[(seg, 'Control', cat, res)]['papd'])
                   * pdC[seg] for seg in SEGS) / PD_C
        bridge_sum += dsim
        if abs(dsim) > 1e-9 or abs(dact) > 1e-9:
            by_cat[cat] = {'d_sim_papd_bridge': round(dsim, 3), 'd_act_papd': round(dact, 3)}
    for seg in SEGS:
        M = sum(simcell[(seg, c, res)]['cur_win'] for c in CATS)
        S = sum(simcell[(seg, c, res)]['new_win'] for c in CATS)
        m33 = sum(simcell[(seg, c, res)]['meas_33'] for c in CATS)
        s33 = sum(simcell[(seg, c, res)]['sim_33'] for c in CATS)
        aC = sum(act[(seg, 'Control', c, res)]['papd'] for c in CATS)
        aV = sum(act[(seg, 'Variant', c, res)]['papd'] for c in CATS)
        r = (S / M) if M > 1e-12 else None
        d = (aC * (r - 1)) if r is not None else None
        if d is not None: num += d * pdC[seg]
        seg_rows[seg] = {'M_win': round(M, 3), 'S_win': round(S, 3),
                         'r_sim_full': (round(r, 4) if r is not None else None),
                         'r_sim_full_33': (round(s33 / m33, 4) if m33 > 1e-12 else None),
                         'a_C': round(aC, 3), 'd_act_papd': round(aV - aC, 3),
                         'd_sim_full_papd': (round(d, 3) if d is not None else None)}
    aC_tot = agg[res]['a_C_papd']
    full[res] = {'per_segment': seg_rows, 'by_category': by_cat,
                 'bridge_check_papd': round(bridge_sum - num / PD_C, 4),
                 'predicted_change_full_papd': round(num / PD_C, 3),
                 'r_sim_full_blended': (round(1 + (num / PD_C) / aC_tot, 4) if aC_tot else None),
                 'actual_change_papd': agg[res]['actual_change_papd'],
                 'miss_papd': round(agg[res]['actual_change_papd'] - num / PD_C, 3)}
# gate: the per-source decomposition must add back to the whole-faucet number it decomposes
gates['full_scope_bridge_max_resid'] = round(max(abs(full[r]['bridge_check_papd'])
                                                 for r in COMPARABLE_RES), 5)
assert gates['full_scope_bridge_max_resid'] < 1e-3, gates['full_scope_bridge_max_resid']

# ------------------------------------------------- which engine version is the workbook showing?
# The saga readers went header-driven 2026-08-13; until EcoGainsSim_v4.gs is re-pasted into the
# Apps Script project the workbook's display sheets keep the pre-fix numbers (chest IDs priced as
# coins -> Saga HC ~7x -> `Sim per Segment` shows a big POSITIVE HC uplift while this report,
# running the fixed reader, shows a decrease). Attribute the cached fill to a version instead of
# asserting it: harness/_dump_engine_versions.js runs both engines over the same mockdata.
def _newest_cal_workbook():
    best, bn = None, -1
    for p in glob.glob(os.path.join(WB, 'NEW_LIVEOPS_CALENDAR_ECO*.xlsx')):
        m = re.search(r'\((\d+)\)', p)
        n = int(m.group(1)) if m else 0
        if n > bn: best, bn = p, n
    return best, bn

wbfill = None
ev_path = os.path.join(OUT, 'engine_versions.json')
if os.path.exists(ev_path):
    ev = json.load(open(ev_path, encoding='utf-8'))['versions']
    calp, caln = _newest_cal_workbook()
    wbc = openpyxl.load_workbook(calp, data_only=True)
    rows_wb, overall_wb = {}, {}
    if 'Sim per Segment' in wbc.sheetnames:
        shs = wbc['Sim per Segment']
        # locate the '◆ HC' block, then its header row's two 'Total' columns (CURRENT, SIMULATED)
        hc_r = next((r for r in range(1, shs.max_row + 1)
                     if str(shs.cell(row=r, column=2).value or '').strip().startswith('◆ HC')), None)
        if hc_r:
            hdr = next((r for r in range(hc_r, min(hc_r + 6, shs.max_row + 1))
                        if str(shs.cell(row=r, column=2).value or '').strip() == 'Segment'), None)
            def _cols(name):
                return [c for c in range(3, shs.max_column + 1)
                        if str(shs.cell(row=hdr, column=c).value or '').strip() == name] if hdr else []
            tot_cols, core_cols = _cols('Total'), _cols('CORE')
            payer = None
            for r in range((hdr or hc_r) + 1, min((hdr or hc_r) + 20, shs.max_row + 1)):
                lab = str(shs.cell(row=r, column=2).value or '').strip()
                if lab in ('NONPAYER', 'PAYER'): payer = lab; continue
                if lab in SEGS + ['overall'] and payer and len(tot_cols) >= 2:
                    cur = shs.cell(row=r, column=tot_cols[0]).value
                    simv = shs.cell(row=r, column=tot_cols[1]).value
                    if isinstance(cur, (int, float)) and isinstance(simv, (int, float)):
                        rec = {'cur_total': float(cur), 'sim_total': float(simv)}
                        if len(core_cols) >= 2:
                            cc_, cs_ = (shs.cell(row=r, column=core_cols[0]).value,
                                        shs.cell(row=r, column=core_cols[1]).value)
                            if isinstance(cc_, (int, float)) and isinstance(cs_, (int, float)):
                                rec['cur_core'] = float(cc_); rec['sim_core'] = float(cs_)
                        (overall_wb if lab == 'overall' else rows_wb)[f'{payer}|{lab}'] = rec
    if rows_wb:
        def _fit(vkey):
            worst, det = 0.0, {}
            for k, w in rows_wb.items():
                s = ev[vkey]['totals'][k]['HC']['sim_33']
                rel = abs(s - w['sim_total']) / max(abs(w['sim_total']), 1e-9)
                worst = max(worst, rel)
                det[k] = {'workbook_sim': round(w['sim_total'], 1), 'engine_sim': round(s, 1),
                          'rel_err': round(rel, 4)}
            return worst, det
        worst_ref, det_ref = _fit('reference')
        worst_cur, det_cur = _fit('current')
        # measured side is version-independent — a mismatch there means the fill predates the data
        meas_worst = max(abs(ev['current']['totals'][k]['HC']['meas_33'] - w['cur_total'])
                         / max(abs(w['cur_total']), 1e-9) for k, w in rows_wb.items())
        matches = ('current' if worst_cur <= worst_ref else 'reference')
        wbfill = {
            'workbook': os.path.basename(calp), 'workbook_n': caln,
            'reference_label': ev['reference']['label'], 'current_label': ev['current']['label'],
            'reference_equals_current': ev.get('reference_equals_current', False),
            'matches': matches, 'stale': matches == 'reference' and worst_ref < 0.05,
            'max_rel_err_vs_reference': round(worst_ref, 4),
            'max_rel_err_vs_current': round(worst_cur, 4),
            'max_rel_err_measured_side': round(meas_worst, 4),
            'rows': rows_wb, 'overall_rows': overall_wb,
            # the sheet's own headline: overall HC uplift % it currently reports, per payer
            'overall_uplift_pct': {k: round((v['sim_total'] / v['cur_total'] - 1) * 100, 1)
                                   for k, v in overall_wb.items() if v['cur_total']},
            'detail_reference': det_ref, 'detail_current': det_cur,
            'saga_R_reference': {k: v['R'] for k, v in ev['reference']['saga'].items()},
            'saga_R_current': {k: v['R'] for k, v in ev['current']['saga'].items()},
        }
        gates['workbook_fill_matches'] = matches
        gates['workbook_fill_max_rel_err'] = {'reference': round(worst_ref, 4),
                                              'current': round(worst_cur, 4)}

out = {
    'full_scope': full, 'workbook_fill_check': wbfill,
    'meta': {'window': aux['window'], 'day_list': sim['meta']['day_list'],
             'ab_summary': os.path.basename(AB_SUMMARY_PATH),
             'sim_vintage': 'NEW_LIVEOPS_CALENDAR_ECO (16).xlsx via harness/_mockdata.json',
             'ns_simulate': sim['meta']['ns_simulate'],
             'fingerprint': sim['meta']['mockdata_fingerprint'],
             'segments': SEGS, 'comparable_resources': COMPARABLE_RES},
    'cells': cells, 'aggregates': agg, 'anchor_check': anchor, 'story': story,
    'gates': gates, 'tau': {f'{k[0]}|{k[1]}': round(v, 5) for k, v in tau.items()},
    'denominators': aux['denominators'], 'excluded_resources': aux['excluded_resources'],
    'extract_gates': aux['gates'],
}
with open(os.path.join(OUT, 'comparison.json'), 'w', encoding='utf-8') as f:
    json.dump(out, f, indent=1)

print('cells:', len(cells), '| classes:', dict(cc))
print('persource structure recon rows:', len(recon),
      '| max share delta pp:', gates['persource_max_share_delta_pp'])
print('anchor rows:', len(anchor))
n_red = sum(1 for c in cells if c['status'] == 'red')
n_amber = sum(1 for c in cells if c['status'] == 'amber')
n_green = sum(1 for c in cells if c['status'] == 'green')
print(f'status: {n_green} green / {n_amber} amber / {n_red} red')
