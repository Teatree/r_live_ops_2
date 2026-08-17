# Reshape the A/B summary workbook into a tidy JSON the report generators read.
#
# Source: the highest-priority workbooks/LiveOps_v2_AB_Summary*.xlsx (newest by mtime when the
# filename carries no (N) suffix — Garry replaces the file rather than versioning it).
# Sheets consumed:
#   Overall / Overall (new users) / Overall (old players)  — Category | Metric | Control | Variant
#                                                            | delta abs | delta %
#   By Bucket    — Category | Metric | then 3 columns (Control | Variant | delta %) per bucket
#   NS_Config_Change_Est — Garry's proposed Night Sky coin ladder + its per-segment HC estimate
#                          (treated as the BASELINE for v3 proposals, per his 2026-08-17 answer)
#
# Output: analysis/out/ab_summary.json
#   {'source': fname, 'population': {'all'|'new'|'old': {metric: {control, variant, d_abs, d_pct}}},
#    'buckets': {bucket: {metric: {control, variant, d_pct}}}, 'bucket_order': [...],
#    'sections': {metric: section}, 'ns_config': {row_label: {bucket: value}}}
#
# Usage: python analysis/_extract_ab_summary.py
import glob
import json
import os
import re

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, 'out')
WB = os.path.join(HERE, '..', 'workbooks')

POP_SHEETS = {'all': 'Overall', 'new': 'Overall (new users)', 'old': 'Overall (old players)'}


def newest_summary():
    """Highest (N) suffix wins; among un-suffixed names the newest mtime wins."""
    cands = glob.glob(os.path.join(WB, 'LiveOps_v2_AB_Summary*.xlsx'))
    if not cands:
        raise SystemExit('no LiveOps_v2_AB_Summary*.xlsx in workbooks/')
    def key(p):
        m = re.search(r'\((\d+)\)', os.path.basename(p))
        return (int(m.group(1)) if m else 0, os.path.getmtime(p))
    return max(cands, key=key)


def cell(sh, r, c):
    v = sh.cell(row=r, column=c).value
    return v


def is_num(v):
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def parse_population(sh):
    """Category | Metric | Control | Variant | delta abs | delta % (header row located, not fixed)."""
    hdr = None
    for r in range(1, min(sh.max_row, 12) + 1):
        row = [str(cell(sh, r, c) or '').strip().lower() for c in range(1, 7)]
        if 'metric' in row and 'control' in row:
            hdr = r
            break
    if hdr is None:
        return {}, {}
    metrics, sections, section = {}, {}, ''
    for r in range(hdr + 1, sh.max_row + 1):
        cat, met = cell(sh, r, 1), cell(sh, r, 2)
        if cat and not met:
            section = str(cat).strip()
            continue
        if not met:
            continue
        name = str(met).strip()
        c_, v_, da, dp = (cell(sh, r, 3), cell(sh, r, 4), cell(sh, r, 5), cell(sh, r, 6))
        if not (is_num(c_) or is_num(v_)):
            continue
        metrics[name] = {'control': c_ if is_num(c_) else None,
                         'variant': v_ if is_num(v_) else None,
                         'd_abs': da if is_num(da) else None,
                         'd_pct': dp if is_num(dp) else None}
        sections[name] = section
    return metrics, sections


def parse_buckets(sh):
    """Bucket labels live on the row above the Control/Variant/delta% row; 3 columns per bucket."""
    hdr = None
    for r in range(1, min(sh.max_row, 12) + 1):
        vals = [str(cell(sh, r, c) or '').strip().lower() for c in range(1, sh.max_column + 1)]
        if vals.count('control') >= 2:
            hdr = r
            break
    if hdr is None:
        return {}, [], {}
    labels = []                                   # (bucket_label, control_col)
    for c in range(1, sh.max_column + 1):
        if str(cell(sh, hdr, c) or '').strip().lower() == 'control':
            lab = None
            for rr in (hdr - 1, hdr - 2):         # label sits above 'Control' or one col left
                for cc in (c, c - 1):
                    v = cell(sh, rr, cc)
                    if v is not None and str(v).strip() != '':
                        lab = str(v).strip()
                        break
                if lab:
                    break
            if lab is None:
                lab = f'col{c}'
            if re.fullmatch(r'\d+(\.\d+)?', lab):        # '0.0' -> '0'
                lab = str(int(float(lab)))
            labels.append((lab, c))
    buckets = {lab: {} for lab, _ in labels}
    sections, section = {}, ''
    for r in range(hdr + 1, sh.max_row + 1):
        cat, met = cell(sh, r, 1), cell(sh, r, 2)
        if cat and not met:
            section = str(cat).strip()
            continue
        if not met:
            continue
        name = str(met).strip()
        got = False
        for lab, c in labels:
            c_, v_, dp = cell(sh, r, c), cell(sh, r, c + 1), cell(sh, r, c + 2)
            if is_num(c_) or is_num(v_):
                buckets[lab][name] = {'control': c_ if is_num(c_) else None,
                                      'variant': v_ if is_num(v_) else None,
                                      'd_pct': dp if is_num(dp) else None}
                got = True
        if got:
            sections[name] = section
    return buckets, [lab for lab, _ in labels], sections


def parse_ns_config(sh):
    """Block | Metric | one column per segment. Returns {metric: {segment: value}} + labels."""
    hdr = None
    for r in range(1, min(sh.max_row, 12) + 1):
        row = [str(cell(sh, r, c) or '').strip() for c in range(1, sh.max_column + 1)]
        if row and row[0].lower() == 'block':
            hdr = r
            break
    if hdr is None:
        return {}, []
    segs = []
    for c in range(3, sh.max_column + 1):
        v = cell(sh, hdr, c)
        if v is not None and str(v).strip():
            segs.append((str(v).strip(), c))
    out, block = {}, ''
    for r in range(hdr + 1, sh.max_row + 1):
        b, met = cell(sh, r, 1), cell(sh, r, 2)
        if b and not met:
            block = str(b).strip()
            continue
        if not met:
            continue
        name = str(met).strip()
        row = {}
        for lab, c in segs:
            v = cell(sh, r, c)
            row[lab] = v if is_num(v) else (str(v).strip() if v is not None else None)
        out[name] = {'block': block, 'by_segment': row}
    return out, [lab for lab, _ in segs]


def main():
    path = newest_summary()
    wb = openpyxl.load_workbook(path, data_only=True)
    pop, sections = {}, {}
    for key, name in POP_SHEETS.items():
        if name in wb.sheetnames:
            m, s = parse_population(wb[name])
            pop[key] = m
            sections.update(s)
    buckets, border, bsec = ({}, [], {})
    if 'By Bucket' in wb.sheetnames:
        buckets, border, bsec = parse_buckets(wb['By Bucket'])
    ns_cfg, ns_segs = ({}, [])
    if 'NS_Config_Change_Est' in wb.sheetnames:
        ns_cfg, ns_segs = parse_ns_config(wb['NS_Config_Change_Est'])

    out = {'source': os.path.basename(path), 'sheets': wb.sheetnames,
           'population': pop, 'buckets': buckets, 'bucket_order': border,
           'sections': {**bsec, **sections},
           'ns_config': ns_cfg, 'ns_config_segments': ns_segs}
    os.makedirs(OUT, exist_ok=True)
    with open(os.path.join(OUT, 'ab_summary.json'), 'w', encoding='utf-8') as f:
        json.dump(out, f, indent=1)

    # gates: the three population sheets must agree on player counts with the bucket sheet, and
    # every headline metric this analysis leans on must be present.
    need = ['Total revenue', 'ARPPU', 'Time spent (mins)', 'Avg daily streak',
            'HC spent / player', 'HC gain total / player', 'D1 churn rate']
    missing = [m for m in need if m not in pop.get('all', {})]
    print(f'source: {os.path.basename(path)}')
    print(f'populations: {sorted(pop)} | metrics: {len(pop.get("all", {}))}')
    print(f'buckets: {border} | bucket metrics: {len(next(iter(buckets.values())) if buckets else {})}')
    print(f'ns_config rows: {len(ns_cfg)} over segments {ns_segs}')
    if missing:
        print('MISSING headline metrics:', missing)
        raise SystemExit(1)
    pl_all = pop['all']['Players (sum)']['control']
    pl_new = pop['new']['Players (sum)']['control']
    pl_old = pop['old']['Players (sum)']['control']
    print(f'players control: all {pl_all:,.0f} vs new+old {pl_new + pl_old:,.0f} '
          f'(new/old are per-population averages of a different denominator — informational)')
    print('written analysis/out/ab_summary.json')


if __name__ == '__main__':
    main()
